from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode
import calendar

import hashlib
import json
import logging
import unicodedata
import urllib.error
import urllib.request

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Request, Response, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image, KeepInFrame, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session, joinedload
from starlette.datastructures import UploadFile as StarletteUploadFile

from app.core.config import get_settings
from app.core.admin_access import has_admin_access, is_super_admin_email
from app.core.csrf import validate_csrf
from app.core.deps import get_csrf_token, get_current_user_web
from app.core.security import get_password_hash, verify_password
from app.core.timezone import app_now, format_app_datetime, today_in_app_timezone
from app.core.user_context import persist_user_context, sync_user_context_from_preferences
from app.db.session import get_db
from app.models import (
    AgronomicProfile,
    CoffeeVariety,
    CropSeason,
    EquipmentAsset,
    EquipmentAssetAttachment,
    Farm,
    FinanceAccount,
    FinanceCustomBank,
    FinanceTransaction,
    FinanceTransactionAttachment,
    FinanceTransactionInstallment,
    FertilizationItem,
    FertilizationSchedule,
    FertilizationRecord,
    HarvestRecord,
    InputRecommendation,
    IrrigationRecord,
    PestIncident,
    Plot,
    PlotAttachment,
    PurchasedInput,
    PurchasedInputAttachment,
    RainfallRecord,
    SoilAnalysis,
    User,
)
from app.repositories.farm import FarmRepository
from app.services.backup_service import delete_backup_run, execute_backup
from app.services.dashboard import build_dashboard_context
from app.services.finance_overview import build_finance_extract_rows, build_finance_overview_context
from app.services.farm_preview_image import (
    ensure_farm_preview_thumb,
    farm_preview_fs_path,
    farm_preview_thumb_fs_path,
    generate_farm_preview_image,
    remove_farm_preview_image,
)
from app.services.plot_preview_image import (
    ensure_plot_preview_thumb,
    generate_plot_geometry_session_preview,
    generate_plot_preview_draft,
    generate_plot_preview_image,
    plot_preview_fs_path,
    plot_preview_thumb_fs_path,
    remove_plot_preview_draft,
    remove_plot_preview_image,
)
from app.services.forms import (
    calculate_geojson_area_hectares,
    calculate_irrigation_volume,
    calculate_soil_recommendations,
    create_agronomic_profile,
    create_crop_season,
    create_equipment_asset,
    create_farm,
    create_fertilization,
    create_fertilization_schedule,
    create_harvest,
    create_input_recommendation,
    create_irrigation,
    create_manual_stock_output,
    create_pest_incident,
    create_plot,
    create_purchased_input,
    create_rainfall,
    create_soil_analysis,
    create_user,
    create_variety,
    estimate_geojson_centroid,
    extract_geojson_file,
    normalize_geojson,
    update_agronomic_profile,
    update_crop_season,
    update_equipment_asset,
    update_farm,
    update_fertilization,
    update_fertilization_schedule,
    update_harvest,
    update_input_recommendation,
    update_irrigation,
    update_pest_incident,
    update_plot,
    update_purchased_input,
    update_rainfall,
    update_soil_analysis,
    update_user,
    update_variety,
    conclude_fertilization_schedule,
    delete_fertilization,
    delete_fertilization_schedule,
    fertilization_application_method_label,
    delete_manual_stock_output,
    validate_schedule_stock,
    update_manual_stock_output,
)
from app.services.openai_service import gerar_recomendacao_adubacao
from app.services.password_change import (
    get_active_password_change_verification,
    issue_password_change_verification,
    revoke_password_change_verifications,
    verify_password_change_code,
)
from app.services.password_reset import revoke_user_password_reset_tokens
from app.services.email_service import send_access_code_email, send_password_change_code_email
from app.services.trusted_browser import revoke_user_trusted_browsers
from app.services.two_factor import get_active_login_code, issue_login_verification_code, revoke_active_login_codes, verify_login_code

settings = get_settings()
logger = logging.getLogger(__name__)

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
templates.env.filters["datetime_sp"] = format_app_datetime
templates.env.filters["decimal_br"] = lambda value, places=2: _format_decimal_br(value, places)
templates.env.filters["quantity_br"] = lambda value, unit=None: _format_quantity_br(value, unit)

EQUIPMENT_ASSET_CATEGORY_OPTIONS = [
    "Benfeitoria",
    "Equipamento",
    "Máquina",
    "Pivô",
    "Silo",
    "Veículo",
]
PENDING_PASSWORD_CHANGE_SESSION_KEY = "pending_password_change_user_id"
PENDING_SUPER_ADMIN_2FA_DISABLE_SESSION_KEY = "pending_super_admin_2fa_disable"
PREVIOUS_CONTEXT_SESSION_KEY = "previous_context"
HISTORY_PAGE_SIZE = 10
SUPPLY_CATEGORY_OPTIONS = [
    "Mudas",
    "Materiais de manutenção",
    "Hidráulica e irrigação",
    "Peças e sobressalentes",
    "Almoxarifado",
    "Ferramentas e consumíveis",
    "EPIs",
    "Outros",
]
FINANCE_BANK_OPTIONS = [
    {"code": "001", "name": "Banco do Brasil S.A.", "mark": "BB", "bg": "#fde047", "fg": "#1d4ed8"},
    {"code": "237", "name": "Banco Bradesco S.A.", "mark": "B", "bg": "#e11d48", "fg": "#ffffff"},
    {"code": "341", "name": "Itaú Unibanco S.A.", "mark": "I", "bg": "#f97316", "fg": "#ffffff"},
    {"code": "104", "name": "Caixa Econômica Federal", "mark": "CA", "bg": "#2563eb", "fg": "#ffffff"},
    {"code": "033", "name": "Banco Santander (Brasil) S.A.", "mark": "S", "bg": "#dc2626", "fg": "#ffffff"},
    {"code": "756", "name": "Banco Cooperativo do Brasil S.A. (Sicoob)", "mark": "SC", "bg": "#16a34a", "fg": "#ffffff"},
    {"code": "748", "name": "Banco Cooperativo Sicredi S.A.", "mark": "SI", "bg": "#22c55e", "fg": "#ffffff"},
    {"code": "077", "name": "Banco Inter S.A.", "mark": "IN", "bg": "#f97316", "fg": "#ffffff"},
    {"code": "260", "name": "Nu Pagamentos S.A. (Nubank)", "mark": "NU", "bg": "#7c3aed", "fg": "#ffffff"},
    {"code": "336", "name": "Banco C6 S.A.", "mark": "C6", "bg": "#111827", "fg": "#ffffff"},
    {"code": "422", "name": "Banco Safra S.A.", "mark": "SA", "bg": "#0f766e", "fg": "#ffffff"},
]
FINANCE_TRANSACTION_EXPENSE_CATEGORIES = [
    "Acaricida",
    "Adjuvante",
    "Amortizaçao de Empréstimos",
    "Aquisiçao de Imóveis",
    "Aquisiçao de Máquinas e Implementos",
    "Aquisiçao de Veículos",
    "Arrendamento de Terra",
    "Combustíveis",
    "Comercializaçao",
    "Compra de Água",
    "Compra de Animais",
    "Comunicação (Telefone, Intenert)",
    "Construção de Benfeitorias",
    "Corretivos de Solo",
    "Dedução de Receita de Venda",
    "Despesas com Viagens",
    "Despesas Diversas",
    "Devoluçao/Cancelamento de Venda",
    "Encargos Financeiros (Juros, Taxas e Multas)",
    "Encargos Sociais",
    "Energia Elétrica",
    "Feritlizantes",
    "Financiamentos",
    "Fretes",
    "Fungicida",
    "Herbicida",
    "Impostos (IPTU, ITR, etc.)",
    "Inseticida",
    "IRPF, IRPJ ou CSLL",
    "Irrigação",
    "Lubrificantes",
    "Manutenção de Máquinas e Equipamentos",
    "Manutenção de Benfeitorias",
    "Outras Despesas Administrativas",
    "Outros Custos de Insumos",
    "Outros Custos de Máquinas",
    "Outros Defensivos",
    "Outros Investimentos",
    "Outros Itens",
    "Peças de Máquinas e Equipamentos",
    "Secagem de Grãos",
    "Seguros",
    "Sementes e Mudas",
    "Serviços Terceirizados",
    "Taxas",
    "Transportes Internos",
]
FINANCE_TRANSACTION_REVENUE_CATEGORIES = [
    "Outras Receitas Operacionais",
    "Receita de Aluguel ou Arrendamento",
    "Recita de Juros, Dividendos e Lucros",
    "Receitas de Estoque",
    "Venda Agrícola",
    "Venda Pecuária",
]
FINANCE_TRANSACTION_PAYMENT_METHODS = [
    "PIX",
    "TED",
    "DOC",
    "Boleto",
    "Cartão",
    "Dinheiro",
    "Débito automático",
    "Cheque",
    "Outro",
]
MENU_ITEM_VISIBILITY_RULES = {
    "users": has_admin_access,
    "backups": has_admin_access,
}


def _redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=status.HTTP_303_SEE_OTHER)


def _is_modal_request(request: Request) -> bool:
    return str(request.query_params.get("modal") or "") == "1"


def _redirect_with_query(path: str, **params) -> RedirectResponse:
    filtered = {}
    for key, value in params.items():
        if value in (None, "", False):
            continue
        if isinstance(value, (list, tuple)):
            if value:
                filtered[key] = value
            continue
        filtered[key] = value
    return _redirect(f"{path}?{urlencode(filtered, doseq=True)}" if filtered else path)


def _redirect_for_request(request: Request, path: str, **params) -> RedirectResponse:
    if _is_modal_request(request):
        params["modal"] = "1"
    return _redirect_with_query(path, **params)


def _base_context(request: Request, user: User, csrf_token: str, page: str, **kwargs):
    modal_mode = _is_modal_request(request)
    flash = request.session.get("flash") if modal_mode else request.session.pop("flash", None)
    repo = kwargs.pop("_repo", None)
    context = {
        "request": request,
        "user": user,
        "csrf_token": csrf_token,
        "page": page,
        "flash": flash,
        "modal_mode": modal_mode,
        "menu_visibility": _build_menu_visibility(user),
    }
    if repo:
        scope_context = _global_scope_context(request, repo, user)
        context.update(scope_context)
        context["context_lock_exempt"] = page in {"farms", "seasons", "profile"}
        context["context_selection_blocking"] = scope_context["context_selection_required"] and not context["context_lock_exempt"]
        context["context_previous_available"] = bool(scope_context.get("previous_context_available"))
    context.update(kwargs)
    return context


def _repository(db: Session) -> FarmRepository:
    return FarmRepository(db)


def _build_menu_visibility(user: User | None) -> dict[str, bool]:
    visibility: dict[str, bool] = {}
    for key, rule in MENU_ITEM_VISIBILITY_RULES.items():
        try:
            visibility[key] = bool(rule(user))
        except Exception:
            visibility[key] = False
    return visibility


def _render_profile_page(
    request: Request,
    user: User,
    csrf_token: str,
    repo: FarmRepository,
    pending_password_change=None,
    active_profile_tab: str = "profile",
):
    return templates.TemplateResponse(
        "profile.html",
        _base_context(
            request,
            user,
            csrf_token,
            "profile",
            title="Meu Perfil",
            _repo=repo,
            profile_gender_options=_profile_gender_options(),
            profile_gender_label=_profile_gender_label(user.gender),
            pending_password_change=pending_password_change,
            active_profile_tab=active_profile_tab,
        ),
    )


def _flash(request: Request, kind: str, message: str) -> None:
    request.session["flash"] = {"kind": kind, "message": message}


def _float_or_none(value: str | None):
    return float(value) if value not in (None, "") else None


def _int_or_none(value: str | None):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _positive_int(value: str | None, default: int = 1) -> int:
    parsed = _int_or_none(value)
    if parsed is None or parsed < 1:
        return default
    return parsed


def _backup_details(value: str | None) -> dict:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def _sort_collection_desc(items, *getters):
    def sort_key(item):
        values = []
        for getter in getters:
            value = getter(item)
            values.append((value is not None, value))
        return tuple(values)

    return sorted(items, key=sort_key, reverse=True)


def _url_with_query(request: Request, **updates) -> str:
    params = [(key, value) for key, value in request.query_params.multi_items() if key not in updates]
    for key, value in updates.items():
        if value in (None, "", False):
            continue
        params.append((key, str(value)))
    query = urlencode(params, doseq=True)
    return f"{request.url.path}?{query}" if query else request.url.path


def _fertilization_schedule_tab_url(request: Request, tab: str) -> str:
    """Apenas schedule_tab: ao trocar Ativos/Concluidos, zera busca, periodo e paginacao."""
    safe_tab = tab if tab in {"active", "completed"} else "active"
    return f"{request.url.path}?{urlencode({'schedule_tab': safe_tab})}"


def _paginate_collection(
    request: Request,
    items,
    page_param: str,
    per_page: int = HISTORY_PAGE_SIZE,
):
    total_items = len(items)
    total_pages = max(1, (total_items + per_page - 1) // per_page)
    page = min(_positive_int(request.query_params.get(page_param), 1), total_pages)
    start = (page - 1) * per_page
    end = start + per_page
    return {
        "items": items[start:end],
        "page": page,
        "total_items": total_items,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "first_url": _url_with_query(request, **{page_param: None}),
        "prev_url": _url_with_query(request, **{page_param: page - 1 if page > 2 else None}),
        "next_url": _url_with_query(request, **{page_param: page + 1}),
        "last_url": _url_with_query(request, **{page_param: total_pages if total_pages > 1 else None}),
    }


def _mark_password_change_pending(request: Request, user_id: int) -> None:
    request.session[PENDING_PASSWORD_CHANGE_SESSION_KEY] = user_id


def _has_password_change_pending(request: Request, user_id: int) -> bool:
    return request.session.get(PENDING_PASSWORD_CHANGE_SESSION_KEY) == user_id


def _clear_password_change_pending(request: Request) -> None:
    request.session.pop(PENDING_PASSWORD_CHANGE_SESSION_KEY, None)


def _discard_password_change_pending(request: Request, db: Session, user_id: int) -> None:
    revoke_password_change_verifications(db, user_id)
    _clear_password_change_pending(request)


def _mark_super_admin_2fa_disable_pending(
    request: Request,
    *,
    target_user_id: int,
    actor_user_id: int,
    confirmed: bool = False,
) -> None:
    request.session[PENDING_SUPER_ADMIN_2FA_DISABLE_SESSION_KEY] = {
        "target_user_id": int(target_user_id),
        "actor_user_id": int(actor_user_id),
        "confirmed": bool(confirmed),
    }


def _get_super_admin_2fa_disable_pending(request: Request) -> dict | None:
    payload = request.session.get(PENDING_SUPER_ADMIN_2FA_DISABLE_SESSION_KEY)
    return payload if isinstance(payload, dict) else None


def _has_super_admin_2fa_disable_pending(request: Request, *, target_user_id: int, actor_user_id: int) -> bool:
    payload = _get_super_admin_2fa_disable_pending(request)
    if not payload:
        return False
    return (
        int(payload.get("target_user_id") or 0) == int(target_user_id)
        and int(payload.get("actor_user_id") or 0) == int(actor_user_id)
    )


def _clear_super_admin_2fa_disable_pending(request: Request) -> None:
    request.session.pop(PENDING_SUPER_ADMIN_2FA_DISABLE_SESSION_KEY, None)


def _super_admin_2fa_disable_pending_confirmed(request: Request, *, target_user_id: int, actor_user_id: int) -> bool:
    payload = _get_super_admin_2fa_disable_pending(request)
    if not payload:
        return False
    return (
        int(payload.get("target_user_id") or 0) == int(target_user_id)
        and int(payload.get("actor_user_id") or 0) == int(actor_user_id)
        and bool(payload.get("confirmed"))
    )


def _mark_super_admin_2fa_disable_confirmed(request: Request, *, target_user_id: int, actor_user_id: int) -> None:
    _mark_super_admin_2fa_disable_pending(
        request,
        target_user_id=target_user_id,
        actor_user_id=actor_user_id,
        confirmed=True,
    )


def _discard_super_admin_2fa_disable_pending(request: Request, db: Session, actor_user_id: int | None = None) -> None:
    if actor_user_id:
        revoke_active_login_codes(db, actor_user_id)
    _clear_super_admin_2fa_disable_pending(request)


def _resolve_fertilization_output_context(
    repo: FarmRepository,
    db: Session,
    output,
) -> tuple[FertilizationItem | None, FertilizationRecord | None]:
    if output.reference_type == "fertilization_item" and output.reference_id:
        fertilization_item = db.query(FertilizationItem).filter(FertilizationItem.id == output.reference_id).first()
        fertilization = (
            repo.get_fertilization(fertilization_item.fertilization_record_id)
            if fertilization_item and fertilization_item.fertilization_record_id
            else None
        )
        return fertilization_item, fertilization

    if output.reference_type == "fertilization_record" and output.reference_id:
        fertilization = repo.get_fertilization(output.reference_id)
        if not fertilization:
            return None, None

        candidates = [
            item for item in fertilization.items
            if output.input_id is not None and item.input_id == output.input_id
        ]
        if not candidates and len(fertilization.items) == 1:
            candidates = list(fertilization.items)
        if len(candidates) > 1 and output.quantity is not None:
            target_quantity = float(output.quantity or 0)
            exact_matches = [
                item for item in candidates
                if abs(float(item.total_quantity or 0) - target_quantity) < 0.01
            ]
            if exact_matches:
                candidates = exact_matches
        return (candidates[0] if candidates else None), fertilization

    return None, None


def _int_list(values: list[str]) -> list[int]:
    parsed: list[int] = []
    for value in values:
        if value in (None, ""):
            continue
        parsed.append(int(value))
    return parsed


def _date_or_none(value: str | None):
    if value in (None, ""):
        return None
    try:
        return date.fromisoformat(value.strip())
    except (TypeError, ValueError):
        return None


def _clean_text(value: str | None) -> str | None:
    normalized = " ".join((value or "").strip().split())
    return normalized or None


def _profile_gender_options() -> list[tuple[str, str]]:
    return [
        ("feminino", "Feminino"),
        ("masculino", "Masculino"),
        ("nao_informar", "Prefiro nao informar"),
        ("outro", "Outro"),
    ]


def _profile_gender_label(value: str | None) -> str:
    for option_value, label in _profile_gender_options():
        if option_value == value:
            return label
    return "Nao informado"


def _parse_equipment_asset_manufacture_year(value: str | None) -> int | None:
    normalized = (value or "").strip()
    if not normalized:
        return None
    if len(normalized) != 4 or not normalized.isdigit():
        raise ValueError("Informe o ano de fabricacao no formato YYYY.")
    parsed = int(normalized)
    current_year = today_in_app_timezone().year
    if parsed < 1900 or parsed > current_year:
        raise ValueError(f"Informe um ano de fabricacao entre 1900 e {current_year}.")
    return parsed


async def _read_avatar_upload(avatar: UploadFile | None) -> tuple[dict | None, str | None]:
    if not avatar or not avatar.filename:
        return None, None

    if not (avatar.content_type or "").startswith("image/"):
        await avatar.close()
        return None, "Envie uma imagem valida para o avatar."

    avatar_bytes = await avatar.read()
    await avatar.close()
    if not avatar_bytes:
        return None, "A imagem selecionada esta vazia."

    return {
        "avatar_filename": Path(avatar.filename).name[:255],
        "avatar_content_type": (avatar.content_type or "image/jpeg")[:120],
        "avatar_data": avatar_bytes,
    }, None


def _active_farm_id(request: Request) -> int | None:
    return _int_or_none(request.session.get("active_farm_id"))


def _active_season_id(request: Request) -> int | None:
    return _int_or_none(request.session.get("active_season_id"))


def _global_scope_context(request: Request, repo: FarmRepository, user: User | None = None) -> dict:
    if user:
        sync_user_context_from_preferences(request, repo.db, user)

    farms = repo.list_farms()
    active_farm_id = _active_farm_id(request)
    active_season_id = _active_season_id(request)
    active_farm = repo.get_farm(active_farm_id) if active_farm_id else None
    active_season = repo.get_crop_season(active_season_id) if active_season_id else None

    # Contexto anterior (para permitir cancelar o modal obrigatório quando houver algo para voltar).
    previous_payload = request.session.get(PREVIOUS_CONTEXT_SESSION_KEY)
    prev_farm_id = _int_or_none(previous_payload.get("farm_id")) if isinstance(previous_payload, dict) else None
    prev_season_id = _int_or_none(previous_payload.get("season_id")) if isinstance(previous_payload, dict) else None
    prev_farm = repo.get_farm(prev_farm_id) if prev_farm_id else None
    prev_season = repo.get_crop_season(prev_season_id) if prev_season_id else None
    previous_context_available = bool(prev_farm and prev_season and prev_season.farm_id == prev_farm.id)

    if active_farm_id and not active_farm:
        if active_farm_id and active_season_id:
            request.session[PREVIOUS_CONTEXT_SESSION_KEY] = {"farm_id": active_farm_id, "season_id": active_season_id}
        request.session.pop("active_farm_id", None)
        active_farm_id = None
    if active_season_id and not active_season:
        if active_farm_id and active_season_id:
            request.session[PREVIOUS_CONTEXT_SESSION_KEY] = {"farm_id": active_farm_id, "season_id": active_season_id}
        request.session.pop("active_season_id", None)
        active_season_id = None

    if active_season and active_farm_id and active_season.farm_id != active_farm_id:
        if active_farm_id and active_season_id:
            request.session[PREVIOUS_CONTEXT_SESSION_KEY] = {"farm_id": active_farm_id, "season_id": active_season_id}
        request.session.pop("active_season_id", None)
        active_season = None
        active_season_id = None

    all_seasons = repo.list_crop_seasons()
    context_seasons = [
        season for season in all_seasons if active_farm_id and season.farm_id == active_farm_id
    ]
    context_season_options = [
        {"id": season.id, "farm_id": season.farm_id, "name": season.name}
        for season in all_seasons
    ]
    current_url = request.url.path
    if request.url.query:
        current_url = f"{current_url}?{request.url.query}"

    return {
        "context_farms": farms,
        "context_seasons": context_seasons,
        "context_season_options": context_season_options,
        "active_farm_id": active_farm_id,
        "active_season_id": active_season_id,
        "active_farm": active_farm,
        "active_season": active_season,
        "context_selection_required": not active_farm_id or not active_season_id,
        "context_missing_farm": not active_farm_id,
        "context_missing_season": not active_season_id,
        "current_url": current_url,
        "previous_context_available": previous_context_available,
        "previous_context_farm_id": prev_farm_id if previous_context_available else None,
        "previous_context_season_id": prev_season_id if previous_context_available else None,
    }


def _scoped_plot_filters(request: Request, active_season: CropSeason | None) -> tuple[list[int] | None, list[int] | None]:
    farm_ids = [_active_farm_id(request)] if _active_farm_id(request) else None
    variety_ids = [active_season.variety_id] if active_season and active_season.variety_id else None
    return farm_ids, variety_ids


def _scoped_dates(active_season: CropSeason | None) -> tuple[date | None, date | None]:
    if not active_season:
        return None, None
    return active_season.start_date, active_season.end_date


def _within_scope(value: date | None, start_date: date | None, end_date: date | None) -> bool:
    if start_date and value and value < start_date:
        return False
    if end_date and value and value > end_date:
        return False
    return True


def _schedule_filter_date_bounds(
    request: Request,
    active_season,
    *,
    flash_invalid: bool = False,
) -> tuple[date | None, date | None, str, str]:
    """Restringe o intervalo da safra ativa com start_date/end_date opcionais na query."""
    scope_start, scope_end = _scoped_dates(active_season)
    raw_start = (request.query_params.get("start_date") or "").strip()
    raw_end = (request.query_params.get("end_date") or "").strip()
    user_start: date | None = None
    user_end: date | None = None
    if raw_start:
        try:
            user_start = date.fromisoformat(raw_start)
        except ValueError:
            raw_start = ""
    if raw_end:
        try:
            user_end = date.fromisoformat(raw_end)
        except ValueError:
            raw_end = ""
    if user_start and user_end and user_start > user_end:
        if flash_invalid:
            _flash(request, "error", "A data inicial do periodo nao pode ser posterior a data final.")
        raw_start = ""
        raw_end = ""
        user_start = None
        user_end = None
    eff_start = scope_start
    eff_end = scope_end
    if user_start:
        eff_start = user_start if eff_start is None else max(eff_start, user_start)
    if user_end:
        eff_end = user_end if eff_end is None else min(eff_end, user_end)
    return eff_start, eff_end, raw_start, raw_end


def _planning_filter_range_preset(raw_preset: str | None, raw_start: str, raw_end: str) -> str:
    valid_presets = {"next_10_days", "next_20_days", "next_month", "custom"}
    preset = (raw_preset or "").strip()
    if preset in valid_presets:
        return preset
    if raw_start or raw_end:
        return "custom"
    return "next_10_days"


def _fertilization_filter_range_preset(raw_preset: str | None, raw_start: str, raw_end: str) -> str:
    valid_presets = {"last_10_days", "last_20_days", "last_month", "custom"}
    preset = (raw_preset or "").strip()
    if preset in valid_presets:
        return preset
    if raw_start or raw_end:
        return "custom"
    return "last_10_days"


def _period_filter_explicit_in_query(request: Request) -> bool:
    """True se o usuario passou periodo na URL (igual convencao da pagina estoque)."""
    qp = request.query_params
    return bool((qp.get("schedule_range") or "").strip() or (qp.get("start_date") or "").strip() or (qp.get("end_date") or "").strip())


def _finance_extract_period_bounds(
    request: Request,
    *,
    flash_invalid: bool = False,
) -> tuple[date | None, date | None, str, str]:
    """Intervalo do extrato financeiro: apenas datas da query, sem limitar à safra ativa."""
    raw_start = (request.query_params.get("start_date") or "").strip()
    raw_end = (request.query_params.get("end_date") or "").strip()
    user_start: date | None = None
    user_end: date | None = None
    if raw_start:
        try:
            user_start = date.fromisoformat(raw_start)
        except ValueError:
            raw_start = ""
    if raw_end:
        try:
            user_end = date.fromisoformat(raw_end)
        except ValueError:
            raw_end = ""
    if user_start and user_end and user_start > user_end:
        if flash_invalid:
            _flash(request, "error", "A data inicial do periodo nao pode ser posterior a data final.")
        return None, None, "", ""
    return user_start, user_end, raw_start, raw_end


def _finance_extract_apply_season_bounds(
    repo: FarmRepository,
    farm_id: int | None,
    *,
    period_start: date | None,
    period_end: date | None,
    extract_season_id: int | None,
) -> tuple[date | None, date | None, int | None, bool]:
    """Restringe o extrato ao intervalo da safra (interseção com datas do filtro). Retorna (start, end, id válido ou None, vazio)."""
    if not farm_id or not extract_season_id:
        return period_start, period_end, None, False
    season = repo.get_crop_season(extract_season_id)
    if not season or season.farm_id != farm_id:
        return period_start, period_end, None, False
    ss, se = season.start_date, season.end_date
    if period_start is None and period_end is None:
        return ss, se, extract_season_id, False
    low = ss if period_start is None else max(period_start, ss)
    high = se if period_end is None else min(period_end, se)
    if low > high:
        return period_start, period_end, extract_season_id, True
    return low, high, extract_season_id, False


def _schedule_tab_filter_range_preset(
    raw_preset: str | None,
    raw_start: str,
    raw_end: str,
    *,
    schedule_tab: str,
) -> str:
    """Aba Ativos: presets futuros; aba Concluídos: presets passados."""
    preset = (raw_preset or "").strip()
    past_ok = {"last_10_days", "last_20_days", "last_month", "custom"}
    future_ok = {"next_10_days", "next_20_days", "next_month", "custom"}
    if schedule_tab == "completed":
        if preset in past_ok:
            return preset
        if raw_start or raw_end:
            return "custom"
        return "last_10_days"
    if preset in future_ok:
        return preset
    if raw_start or raw_end:
        return "custom"
    return "next_10_days"


def _normalize_search_value(value: object) -> str:
    return (
        unicodedata.normalize("NFD", str(value or ""))
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )


def _filter_fertilization_records(
    repo: FarmRepository,
    plot_ids: set[int],
    start_date: date | None,
    end_date: date | None,
    *,
    search: str | None = None,
) -> list[FertilizationRecord]:
    fertilizations = [
        item
        for item in repo.list_fertilizations()
        if item.plot_id in plot_ids and _within_scope(item.application_date, start_date, end_date)
    ]
    query = _normalize_search_value((search or "").strip())
    if query:
        fertilizations = [
            item
            for item in fertilizations
            if query in _normalize_search_value(
                f"{item.plot.name if item.plot else ''} "
                f"{item.application_date or ''} "
                f"{item.product or ''} "
                f"{item.notes or ''} "
                + " ".join(detail.name for detail in item.items)
            )
        ]
    fertilizations = _sort_collection_desc(
        fertilizations,
        lambda item: item.application_date,
        lambda item: item.id,
    )
    return fertilizations


def _launch_scope_or_redirect(
    request: Request,
    repo: FarmRepository,
    redirect_path: str,
    require_season: bool = True,
):
    scope = _global_scope_context(request, repo)
    if scope["active_farm_id"] and (scope["active_season_id"] or not require_season):
        return scope, None
    _flash(request, "error", "Selecione a fazenda e a safra ativas no topo antes de continuar.")
    return None, _redirect(redirect_path)


def _plot_matches_scope(plot: Plot | None, scope: dict) -> bool:
    if not plot:
        return False
    if scope.get("active_farm_id") and plot.farm_id != scope["active_farm_id"]:
        return False
    active_season = scope.get("active_season")
    if active_season and active_season.variety_id and plot.variety_id != active_season.variety_id:
        return False
    return True


def _resolve_plot_in_scope(
    request: Request,
    repo: FarmRepository,
    plot_id: int | None,
    redirect_path: str,
):
    scope, denied = _launch_scope_or_redirect(request, repo, redirect_path)
    if denied:
        return None, None, denied
    plot = repo.get_plot(plot_id) if plot_id else None
    if not plot:
        _flash(request, "error", "Setor nao encontrado para o contexto atual.")
        return None, scope, _redirect(redirect_path)
    if not _plot_matches_scope(plot, scope):
        _flash(request, "error", "O setor informado nao pertence ao contexto ativo de fazenda e safra.")
        return None, scope, _redirect(redirect_path)
    return plot, scope, None


def _resolve_optional_plot_in_scope(
    request: Request,
    repo: FarmRepository,
    plot_id: int | None,
    redirect_path: str,
):
    scope, denied = _launch_scope_or_redirect(request, repo, redirect_path)
    if denied:
        return None, scope, denied
    if not plot_id:
        return None, scope, None
    plot = repo.get_plot(plot_id)
    if not plot:
        _flash(request, "error", "Setor nao encontrado para o contexto atual.")
        return None, scope, _redirect(redirect_path)
    if not _plot_matches_scope(plot, scope):
        _flash(request, "error", "O setor informado nao pertence ao contexto ativo de fazenda e safra.")
        return None, scope, _redirect(redirect_path)
    return plot, scope, None


def _farm_matches_scope(farm_id: int | None, scope: dict) -> bool:
    return farm_id in (None, scope.get("active_farm_id"))


def _page_number(value: str | int | None, default: int = 1) -> int:
    try:
        return max(int(value or default), 1)
    except (TypeError, ValueError):
        return default


def _build_stock_context(
    repo: FarmRepository,
    farm_id: int | None = None,
    input_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    movement_type: str = "all",
    item_type: str | None = None,
):
    def _stock_priority(row: dict) -> tuple[int, float, str]:
        threshold = float(row.get("low_stock_threshold") or 0)
        available = float(row.get("available_quantity") or 0)
        if threshold > 0:
            if available <= threshold * 0.4:
                return (0, available, row["name"].lower())
            if available <= threshold:
                return (1, available, row["name"].lower())
        return (2, available, row["name"].lower())

    def _is_inventory_item(catalog_item) -> bool:
        return bool(catalog_item and getattr(catalog_item, "item_type", None) in {"insumo_agricola", "combustivel"})

    catalog_inputs = repo.list_input_catalog(item_type=item_type)
    if not item_type:
        catalog_inputs = [item for item in catalog_inputs if _is_inventory_item(item)]
    if input_id:
        catalog_inputs = [item for item in catalog_inputs if item.id == input_id]

    purchase_entries = repo.list_purchased_inputs(item_type=item_type)
    if not item_type:
        purchase_entries = [entry for entry in purchase_entries if _is_inventory_item(entry.input_catalog)]
    if input_id:
        purchase_entries = [entry for entry in purchase_entries if entry.input_id == input_id]
    if farm_id:
        purchase_entries = [entry for entry in purchase_entries if entry.farm_id in (None, farm_id)]

    stock_outputs = repo.list_stock_outputs()
    if input_id:
        stock_outputs = [output for output in stock_outputs if output.input_id == input_id]
    if farm_id:
        stock_outputs = [output for output in stock_outputs if output.farm_id in (None, farm_id)]
    if item_type:
        stock_outputs = [output for output in stock_outputs if output.input_catalog and output.input_catalog.item_type == item_type]
    else:
        stock_outputs = [output for output in stock_outputs if _is_inventory_item(output.input_catalog)]

    input_stock = {
        item.id: {
            "available": round(
                sum(float(entry.available_quantity or 0) for entry in purchase_entries if entry.input_id == item.id),
                2,
            ),
            "total": round(
                sum(float(entry.total_quantity or 0) for entry in purchase_entries if entry.input_id == item.id),
                2,
            ),
            "unit": item.default_unit,
        }
        for item in catalog_inputs
    }

    stock_catalog_rows = []
    extract_rows = []
    for item in catalog_inputs:
        related_entries = [entry for entry in purchase_entries if entry.input_id == item.id]
        related_outputs = [output for output in stock_outputs if output.input_id == item.id]
        if not related_entries:
            continue
        total_quantity = sum(float(entry.total_quantity or 0) for entry in related_entries)
        total_value = sum(float(entry.total_cost or 0) for entry in related_entries)
        row = {
            "id": item.id,
            "name": item.name,
            "item_type": item.item_type,
            "unit": item.default_unit,
            "available_quantity": input_stock[item.id]["available"],
            "total_quantity": input_stock[item.id]["total"],
            "low_stock_threshold": float(item.low_stock_threshold or 0),
            "average_cost": round(total_value / total_quantity, 2) if total_quantity else 0,
            "entries_count": len(related_entries),
            "outputs_count": len(related_outputs),
            "last_movement_date": (
                max(
                    [entry.purchase_date for entry in related_entries if entry.purchase_date]
                    + [output.movement_date for output in related_outputs if output.movement_date]
                )
                if related_entries or related_outputs
                else None
            ),
        }
        stock_catalog_rows.append(row)

        events = []
        for entry in related_entries:
            unit_cost = round(float(entry.total_cost or 0) / max(float(entry.total_quantity or 0), 1), 4)
            events.append(
                {
                    "kind": "entrada",
                    "date": entry.purchase_date,
                    "quantity": float(entry.total_quantity or 0),
                    "unit": entry.package_unit,
                    "farm": entry.farm,
                    "plot": None,
                    "origin": "compra",
                    "reference": f"Lote #{entry.id}",
                    "value": float(entry.total_cost or 0),
                    "unit_cost": unit_cost,
                    "notes": entry.notes,
                    "sort_key": (entry.purchase_date or today_in_app_timezone(), 0, entry.id),
                }
            )
        for output in related_outputs:
            events.append(
                {
                    "kind": "saida",
                    "date": output.movement_date,
                    "quantity": float(output.quantity or 0),
                    "unit": output.unit,
                    "farm": output.farm,
                    "plot": output.plot,
                    "origin": output.origin,
                    "reference": f"{output.reference_type or 'movimento'}#{output.reference_id}" if output.reference_id else (output.reference_type or "movimento manual"),
                    "value": float(output.total_cost or 0),
                    "unit_cost": float(output.unit_cost or 0),
                    "notes": output.notes,
                    "sort_key": (output.movement_date or today_in_app_timezone(), 1, output.id),
                }
            )

        running_balance = 0.0
        for event in sorted(events, key=lambda item_: item_["sort_key"]):
            delta = event["quantity"] if event["kind"] == "entrada" else -event["quantity"]
            running_balance = round(running_balance + delta, 2)
            extract_rows.append(
                {
                    "input_id": item.id,
                    "input_name": item.name,
                    "kind": event["kind"],
                    "date": event["date"],
                    "quantity": event["quantity"],
                    "unit": event["unit"] or item.default_unit,
                    "farm": event["farm"],
                    "plot": event["plot"],
                    "origin": event["origin"],
                    "reference": event["reference"],
                    "unit_cost": event["unit_cost"],
                    "total_cost": event["value"],
                    "balance_after": running_balance,
                    "notes": event["notes"],
                    "sort_key": event["sort_key"],
                }
            )

    stock_catalog_rows.sort(key=_stock_priority)
    filtered_entries = purchase_entries
    filtered_outputs = stock_outputs
    filtered_extract_rows = extract_rows
    if start_date:
        filtered_entries = [entry for entry in filtered_entries if entry.purchase_date and entry.purchase_date >= start_date]
        filtered_outputs = [output for output in filtered_outputs if output.movement_date and output.movement_date >= start_date]
        filtered_extract_rows = [row for row in filtered_extract_rows if row["date"] and row["date"] >= start_date]
    if end_date:
        filtered_entries = [entry for entry in filtered_entries if entry.purchase_date and entry.purchase_date <= end_date]
        filtered_outputs = [output for output in filtered_outputs if output.movement_date and output.movement_date <= end_date]
        filtered_extract_rows = [row for row in filtered_extract_rows if row["date"] and row["date"] <= end_date]
    if movement_type in {"entrada", "saida"}:
        filtered_extract_rows = [row for row in filtered_extract_rows if row["kind"] == movement_type]
        if movement_type == "entrada":
            filtered_outputs = []
        else:
            filtered_entries = []

    filtered_extract_rows.sort(key=lambda row: row["sort_key"], reverse=True)
    return {
        "catalog_inputs": catalog_inputs,
        "purchase_entries": filtered_entries,
        "stock_outputs": filtered_outputs,
        "input_stock": input_stock,
        "stock_catalog_rows": stock_catalog_rows,
        "extract_rows": filtered_extract_rows,
    }


def _stock_export_query(
    farm_id: int | None = None,
    input_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    movement_type: str = "all",
    item_type: str | None = None,
    stock_tab: str | None = None,
) -> str:
    params = {
        "farm_id": farm_id,
        "input_id": input_id,
        "start_date": start_date.isoformat() if start_date else None,
        "end_date": end_date.isoformat() if end_date else None,
        "movement_type": movement_type if movement_type and movement_type != "all" else None,
        "item_type": item_type,
        "stock_tab": stock_tab if stock_tab in {"entries", "outputs", "extract"} else None,
    }
    clean = {key: value for key, value in params.items() if value not in (None, "", "all")}
    return urlencode(clean)


def _stock_page_filters_active(
    request: Request,
    *,
    active_farm_id: int | None,
    selected_input_id: int | None,
    movement_type: str,
    normalized_item_type: str | None,
    edit_output_id: int | None,
) -> bool:
    if edit_output_id:
        return True
    qp = request.query_params
    if selected_input_id:
        return True
    if (qp.get("start_date") or "").strip() or (qp.get("end_date") or "").strip():
        return True
    if movement_type and movement_type != "all":
        return True
    if normalized_item_type:
        return True
    if (qp.get("schedule_range") or "").strip():
        return True
    raw_farm = qp.get("farm_id")
    if raw_farm is not None and str(raw_farm).strip() != "":
        try:
            if int(raw_farm) != int(active_farm_id or 0):
                return True
        except (TypeError, ValueError):
            return True
    return False


def _plots_farm_query_is_user_filter(request: Request, active_farm_id: int | None) -> bool:
    raw_farm = request.query_params.get("farm_id")
    if raw_farm is None or str(raw_farm).strip() == "":
        return False
    try:
        return int(raw_farm) != int(active_farm_id or 0)
    except (TypeError, ValueError):
        return True


def _plots_variety_query_is_user_filter(request: Request, active_season_variety_id: int | None) -> bool:
    raw_variety = request.query_params.get("variety_id")
    if raw_variety is None or str(raw_variety).strip() == "":
        return False
    try:
        vid = int(raw_variety)
    except (TypeError, ValueError):
        return True
    return active_season_variety_id is None or vid != int(active_season_variety_id)


def _plots_page_filters_active(
    request: Request,
    *,
    active_farm_id: int | None,
    active_season_variety_id: int | None,
    q: str | None,
    sort: str,
) -> bool:
    """Igual ao estoque: verde / 'filtro aplicado' só com parâmetros na URL que alteram o padrão do contexto."""
    if (q or "").strip():
        return True
    if sort != "name":
        return True
    if _plots_farm_query_is_user_filter(request, active_farm_id):
        return True
    if _plots_variety_query_is_user_filter(request, active_season_variety_id):
        return True
    return False


def _assets_export_query(farm_id: int | None = None, status: str | None = None) -> str:
    params = {"farm_id": farm_id, "status": status}
    clean = {key: value for key, value in params.items() if value not in (None, "", "all")}
    return urlencode(clean)


def _purchased_inputs_export_query(
    farm_id: int | None = None,
    item_type: str | None = None,
    purchased_tab: str | None = None,
) -> str:
    params = {
        "farm_id": farm_id,
        "item_type": item_type,
        "purchased_tab": purchased_tab if purchased_tab in {"entries", "outputs", "extract"} else None,
    }
    clean = {key: value for key, value in params.items() if value not in (None, "", "all")}
    return urlencode(clean)


def _export_purchased_tab_param(request: Request) -> str:
    raw = (request.query_params.get("purchased_tab") or "").strip().lower()
    if raw in {"entries", "outputs", "extract"}:
        return raw
    return "entries"


def _export_stock_tab_param(request: Request, movement_type: str) -> str:
    raw = (request.query_params.get("stock_tab") or "").strip().lower()
    if raw in {"entries", "outputs", "extract"}:
        return raw
    if movement_type == "saida":
        return "outputs"
    return "entries"


def _catalog_item_type_label_pt(item_catalog) -> str:
    if item_catalog and item_catalog.item_type == "combustivel":
        return "Combustível"
    return "Insumo agrícola"


def _xlsx_apply_column_widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _xlsx_write_purchase_entries(sheet, purchase_entries) -> None:
    sheet.append(["Data", "Insumo", "Tipo", "Fazenda", "Quantidade", "Saldo", "Valor", "Observações"])
    for item in purchase_entries:
        sheet.append(
            [
                item.purchase_date.isoformat() if item.purchase_date else "",
                item.input_catalog.name if item.input_catalog else item.name,
                _catalog_item_type_label_pt(item.input_catalog),
                item.farm.name if item.farm else "",
                float(item.total_quantity or 0),
                float(item.available_quantity or 0),
                float(item.total_cost or 0),
                item.notes or "",
            ]
        )
    _xlsx_apply_column_widths(sheet, [14, 30, 18, 24, 16, 16, 16, 40])


def _xlsx_write_stock_outputs(sheet, stock_outputs) -> None:
    sheet.append(["Data", "Insumo", "Tipo", "Origem", "Fazenda / Setor", "Quantidade", "Custo", "Observações"])
    for output in stock_outputs:
        sheet.append(
            [
                output.movement_date.isoformat() if output.movement_date else "",
                output.input_catalog.name if output.input_catalog else "Insumo removido",
                _catalog_item_type_label_pt(output.input_catalog),
                output.origin or "",
                f"{output.farm.name if output.farm else ''}{(' / ' + output.plot.name) if output.plot else ''}",
                float(output.quantity or 0),
                float(output.total_cost or 0),
                output.notes or "",
            ]
        )
    _xlsx_apply_column_widths(sheet, [14, 30, 18, 18, 28, 16, 16, 40])


def _xlsx_write_extract_rows(sheet, extract_rows: list[dict]) -> None:
    sheet.append(
        [
            "Data",
            "Insumo",
            "Tipo da movimentacao",
            "Origem",
            "Quantidade",
            "Unidade",
            "Custo unitario",
            "Custo total",
            "Saldo apos movimentacao",
            "Observacoes",
        ]
    )
    for row in extract_rows:
        sheet.append(
            [
                row["date"].isoformat() if row["date"] else "",
                row["input_name"],
                row["kind"],
                row["origin"],
                float(row["quantity"] or 0),
                row["unit"],
                float(row.get("unit_cost") or 0),
                float(row.get("total_cost") or 0),
                float(row.get("balance_after") or 0),
                row.get("notes") or "",
            ]
        )
    _xlsx_apply_column_widths(sheet, [14, 30, 18, 20, 14, 12, 16, 16, 20, 42])


def _format_currency(value) -> str:
    numeric = float(value or 0)
    return f"R$ {numeric:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_decimal_br(value, places: int = 2) -> str:
    numeric = float(value or 0)
    return f"{numeric:,.{places}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_quantity_br(value, unit: str | None = None) -> str:
    discrete_units = {"un", "un.", "peça", "peca", "par", "kit", "cx", "caixa", "rolo", "saco"}
    places = 0 if (unit or "").strip().lower() in discrete_units else 2
    return _format_decimal_br(value, places)


def _format_iso_date_br(value: str | None) -> str:
    raw = (value or "").strip()
    if not raw:
        return "--"
    try:
        return date.fromisoformat(raw).strftime("%d/%m/%Y")
    except ValueError:
        return raw


def _format_duration_hours_minutes(total_minutes: int) -> str:
    t = max(0, int(total_minutes or 0))
    hours, minutes = divmod(t, 60)
    if hours and minutes:
        return f"{hours} h {minutes} min"
    if hours:
        return f"{hours} h"
    return f"{minutes} min"


def _stock_report_totals(rows: list[dict]) -> dict:
    entries_total = sum(float(row.get("total_cost") or 0) for row in rows if row.get("kind") == "entrada")
    outputs_total = sum(float(row.get("total_cost") or 0) for row in rows if row.get("kind") == "saida")
    return {
        "entries_total": round(entries_total, 2),
        "outputs_total": round(outputs_total, 2),
        "grand_total": round(entries_total + outputs_total, 2),
        "movements_count": len(rows),
    }


def _pdf_flowables_extract_detail_table(
    doc,
    rows: list[dict],
    totals: dict,
    *,
    cell_style,
    cell_muted_style,
    cell_numeric_style,
    meta_value_style,
    summary_value_style,
) -> list:
    data = [
        [
            "Data",
            "Insumo",
            "Mov.",
            "Origem / Ref.",
            "Qtd.",
            "Un.",
            "Custo un.",
            "Custo total",
            "Saldo apos",
            "Observacoes",
        ]
    ]
    for row in rows:
        movement_color = "#166534" if row["kind"] == "entrada" else "#be123c"
        data.append(
            [
                Paragraph(row["date"].strftime("%d/%m/%Y") if row["date"] else "-", cell_style),
                Paragraph(row["input_name"], cell_style),
                Paragraph(f'<font color="{movement_color}"><b>{str(row["kind"]).title()}</b></font>', cell_style),
                Paragraph(f'{row["origin"]} • {row["reference"]}', cell_muted_style),
                Paragraph(_format_decimal_br(row["quantity"], 2), cell_numeric_style),
                Paragraph(str(row["unit"] or "-"), cell_style),
                Paragraph(_format_currency(row.get("unit_cost") or 0), cell_numeric_style),
                Paragraph(_format_currency(row.get("total_cost") or 0), cell_numeric_style),
                Paragraph(_format_decimal_br(row.get("balance_after") or 0, 2), cell_numeric_style),
                Paragraph((row.get("notes") or "-")[:90], cell_muted_style),
            ]
        )
    data.append(
        [
            "",
            "",
            "",
            Paragraph("<b>Total geral</b>", cell_style),
            "",
            "",
            "",
            Paragraph(f"<b>{_format_currency(totals['grand_total'])}</b>", cell_numeric_style),
            "",
            "",
        ]
    )
    column_weights = [6, 14, 6, 19, 7, 5, 9, 10, 8, 16]
    weight_total = sum(column_weights)
    table_col_widths = [doc.width * (weight / weight_total) for weight in column_weights[:-1]]
    table_col_widths.append(doc.width - sum(table_col_widths))
    table = Table(data, repeatRows=1, colWidths=table_col_widths, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#36552a")),
                ("LINEBELOW", (0, 1), (-1, -2), 0.35, colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f8fafc")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.2),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                ("ALIGN", (6, 1), (8, -1), "RIGHT"),
            ]
        )
    )
    footer_col_widths = [
        sum(table_col_widths[:4]),
        sum(table_col_widths[4:7]),
        sum(table_col_widths[7:]),
    ]
    footer_summary = Table(
        [[
            Paragraph(f"<b>Entradas:</b> {_format_currency(totals['entries_total'])}", meta_value_style),
            Paragraph(f"<b>Saídas:</b> {_format_currency(totals['outputs_total'])}", meta_value_style),
            Paragraph(f"<b>Total geral:</b> {_format_currency(totals['grand_total'])}", summary_value_style),
        ]],
        colWidths=footer_col_widths,
        hAlign="LEFT",
    )
    footer_summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef6ee")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cfe1d0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return [table, Spacer(1, 12), footer_summary]


def _filter_equipment_assets_by_status(assets: list[EquipmentAsset], status_value: str | None) -> list[EquipmentAsset]:
    normalized_status = status_value if status_value in {"ativo", "em_manutencao", "baixado"} else None
    if not normalized_status:
        return assets
    return [asset for asset in assets if asset.status == normalized_status]


def _bool_from_form(value) -> bool:
    return str(value or "").lower() in {"1", "true", "on", "yes"}


def _meters_from_cm(value: str | None):
    centimeters = _float_or_none(value)
    return round(centimeters / 100, 4) if centimeters is not None else None


def _normalize_irrigation_type(value: str | None) -> str:
    normalized = (value or "none").strip().lower()
    return normalized if normalized in {"none", "gotejo", "aspersor"} else "none"


def _resolve_geojson(
    upload: UploadFile | None,
    fallback_text: str | None,
    current_value: str | None = None,
) -> tuple[str | None, bool, bytes | None, str | None]:
    """
    Retorna (geometria texto, sucesso parcial conforme legado, bytes do upload para anexo, nome do arquivo).
    Quando o upload falha na validação, retorna (None, False, None, None).
    """
    if upload and upload.filename:
        raw_bytes = upload.file.read()
        parsed = extract_geojson_file(raw_bytes)
        if parsed is None:
            return None, False, None, None
        return parsed, True, raw_bytes, _clean_attachment_filename(upload.filename)
    normalized = normalize_geojson(fallback_text)
    if normalized is not None:
        return normalized, True, None, None
    return current_value, False, None, None


def _farm_boundary_for_plot_preview(plot) -> str | None:
    """Perímetro da fazenda para sobrepor na imagem de prévia do setor (plot.farm deve estar carregado)."""
    farm = getattr(plot, "farm", None)
    if not farm:
        return None
    raw = getattr(farm, "boundary_geojson", None)
    return raw if raw and str(raw).strip() else None


def _parse_fertilization_items(values) -> list[dict]:
    input_ids = values.getlist("input_id")
    names = values.getlist("item_name")
    units = values.getlist("item_unit")
    quantities = values.getlist("item_quantity")
    items: list[dict] = []
    for index, name in enumerate(names):
        input_id = input_ids[index] if index < len(input_ids) else ""
        unit = units[index] if index < len(units) else ""
        quantity = quantities[index] if index < len(quantities) else ""
        if not (name or "").strip():
            continue
        items.append(
            {
                "input_id": _int_or_none(input_id),
                "purchased_input_id": None,
                "name": name,
                "unit": unit,
                "quantity": quantity,
            }
        )
    return items


def _legacy_fertilization_items(record: FertilizationRecord | None, plot_area: float | None) -> list[dict]:
    if not record:
        return [{"input_id": "", "purchased_input_id": "", "name": "", "unit": "kg", "quantity": "", "total_quantity": "", "available": ""}]
    if record.items:
        return [
            {
                "input_id": item.input_id or "",
                "purchased_input_id": item.purchased_input_id or "",
                "name": item.name,
                "unit": item.unit,
                "quantity": float(item.quantity_per_hectare),
                "total_quantity": float(item.total_quantity),
                "available": float(item.purchased_input.available_quantity) if item.purchased_input and item.purchased_input.available_quantity is not None else "",
            }
            for item in record.items
        ]
    quantity = ""
    if record.dose:
        quantity = record.dose.split(" ")[0].replace(",", ".")
    quantity_value = _float_or_none(quantity)
    return [
        {
            "input_id": "",
            "purchased_input_id": "",
            "name": record.product,
            "unit": "kg",
            "quantity": quantity_value if quantity_value is not None else "",
            "total_quantity": quantity_value if quantity_value is not None else "",
            "available": "",
        }
    ]


def _parse_recommendation_items(values) -> list[dict]:
    input_ids = values.getlist("input_id")
    units = values.getlist("unit")
    quantities = values.getlist("quantity")
    items: list[dict] = []
    for index, input_id in enumerate(input_ids):
        quantity = quantities[index] if index < len(quantities) else ""
        unit = units[index] if index < len(units) else ""
        if input_id in ("", None) or quantity in ("", None):
            continue
        items.append(
            {
                "input_id": int(input_id),
                "unit": unit,
                "quantity": float(quantity),
            }
        )
    return items


def _build_plot_payload(
    farm: Farm,
    name: str,
    area_hectares: float,
    planting_date: str | None,
    plant_count: int,
    spacing_row_meters: str | None,
    spacing_plant_centimeters: str | None,
    estimated_yield_sacks: str | None,
    variety_id: str | None,
    centroid_lat: str | None,
    centroid_lng: str | None,
    boundary_geojson: str | None,
    notes: str | None,
    irrigation_type: str,
    irrigation_line_count: str | None,
    irrigation_line_length_meters: str | None,
    drip_spacing_centimeters: str | None,
    drip_liters_per_hour: str | None,
    sprinkler_count: str | None,
    sprinkler_liters_per_hour: str | None,
) -> dict:
    auto_lat, auto_lng = estimate_geojson_centroid(boundary_geojson)
    normalized_irrigation_type = _normalize_irrigation_type(irrigation_type)
    payload = {
        "name": name,
        "area_hectares": area_hectares,
        "location": farm.location,
        "planting_date": planting_date,
        "plant_count": plant_count,
        "spacing_row_meters": _float_or_none(spacing_row_meters),
        "spacing_plant_meters": _meters_from_cm(spacing_plant_centimeters),
        "estimated_yield_sacks": _float_or_none(estimated_yield_sacks),
        "variety_id": _int_or_none(variety_id),
        "centroid_lat": _float_or_none(centroid_lat) if centroid_lat not in (None, "") else auto_lat,
        "centroid_lng": _float_or_none(centroid_lng) if centroid_lng not in (None, "") else auto_lng,
        "boundary_geojson": boundary_geojson,
        "notes": notes,
        "farm_id": farm.id,
        "irrigation_type": normalized_irrigation_type,
        "irrigation_line_count": _int_or_none(irrigation_line_count),
        "irrigation_line_length_meters": _float_or_none(irrigation_line_length_meters),
        "drip_spacing_meters": _meters_from_cm(drip_spacing_centimeters),
        "drip_liters_per_hour": _float_or_none(drip_liters_per_hour),
        "sprinkler_count": _int_or_none(sprinkler_count),
        "sprinkler_liters_per_hour": _float_or_none(sprinkler_liters_per_hour),
    }
    if normalized_irrigation_type != "gotejo":
        payload["irrigation_line_length_meters"] = None
        payload["drip_spacing_meters"] = None
        payload["drip_liters_per_hour"] = None
    if normalized_irrigation_type != "aspersor":
        payload["sprinkler_count"] = None
        payload["sprinkler_liters_per_hour"] = None
    if normalized_irrigation_type == "none":
        payload["irrigation_line_count"] = None
    return payload


def _read_upload(upload: UploadFile | None) -> tuple[str | None, str | None, bytes | None]:
    if not upload or not upload.filename:
        return None, None, None
    try:
        upload.file.seek(0)
    except Exception:
        pass
    payload = upload.file.read()
    if not payload:
        return None, None, None
    return upload.filename, upload.content_type or "application/octet-stream", payload


_ALLOWED_ATTACHMENT_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
    ".bmp",
    ".pdf",
    ".txt",
    ".csv",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".odt",
    ".ods",
    ".json",
    ".geojson",
}
_MAX_ATTACHMENT_SIZE_BYTES = 10 * 1024 * 1024


def _clean_attachment_filename(filename: str | None) -> str:
    cleaned = Path((filename or "arquivo").strip()).name
    return cleaned[:255] or "arquivo"


def _read_attachments(uploads: list[UploadFile] | None) -> list[tuple[str, str, bytes]]:
    attachments: list[tuple[str, str, bytes]] = []
    for upload in uploads or []:
        filename, content_type, payload = _read_upload(upload)
        if not filename or payload is None:
            continue
        extension = Path(filename).suffix.lower()
        if extension not in _ALLOWED_ATTACHMENT_EXTENSIONS and not (content_type or "").startswith("image/"):
            raise ValueError("Envie apenas imagens, PDF ou documentos comuns.")
        if len(payload) > _MAX_ATTACHMENT_SIZE_BYTES:
            raise ValueError("Cada anexo deve ter no maximo 10 MB.")
        attachments.append((_clean_attachment_filename(filename), content_type or "application/octet-stream", payload))
    return attachments


def _save_purchased_input_attachments(
    repo: FarmRepository,
    item: PurchasedInput,
    attachments: list[tuple[str, str, bytes]],
) -> int:
    if not attachments:
        return 0
    repo.db.add_all(
        [
            PurchasedInputAttachment(
                purchased_input_id=item.id,
                filename=filename,
                content_type=content_type,
                file_data=payload,
            )
            for filename, content_type, payload in attachments
        ]
    )
    try:
        repo.db.commit()
    except Exception:
        repo.db.rollback()
        raise
    repo.db.refresh(item)
    return len(attachments)


def _save_equipment_asset_attachments(
    repo: FarmRepository,
    asset: EquipmentAsset,
    attachments: list[tuple[str, str, bytes]],
) -> int:
    if not attachments:
        return 0
    repo.db.add_all(
        [
            EquipmentAssetAttachment(
                equipment_asset_id=asset.id,
                filename=filename,
                content_type=content_type,
                file_data=payload,
            )
            for filename, content_type, payload in attachments
        ]
    )
    try:
        repo.db.commit()
    except Exception:
        repo.db.rollback()
        raise
    repo.db.refresh(asset)
    return len(attachments)


def _pretty_geojson_file_bytes(geojson_text: str) -> bytes:
    try:
        obj = json.loads(geojson_text)
        return json.dumps(obj, ensure_ascii=False, indent=2).encode("utf-8")
    except (json.JSONDecodeError, TypeError):
        return (geojson_text or "").encode("utf-8")


def _persist_plot_geojson_attachment(
    repo: FarmRepository,
    plot: Plot,
    payload: bytes,
    filename: str,
    content_type: str = "application/geo+json",
) -> bool:
    if not payload or len(payload) > _MAX_ATTACHMENT_SIZE_BYTES:
        return False
    try:
        repo.db.add(
            PlotAttachment(
                plot_id=plot.id,
                filename=_clean_attachment_filename(filename),
                content_type=(content_type or "application/geo+json")[:120],
                file_data=payload,
            )
        )
        repo.db.commit()
        repo.db.refresh(plot)
        return True
    except Exception:
        repo.db.rollback()
        logging.getLogger(__name__).exception("Falha ao gravar anexo GeoJSON do setor %s", plot.id)
        return False


def _maybe_record_plot_boundary_attachments(
    repo: FarmRepository,
    plot: Plot,
    geometry: str | None,
    upload_payload: bytes | None,
    upload_filename: str | None,
    old_geometry: str | None,
    *,
    is_new_plot: bool,
) -> None:
    if not geometry or not str(geometry).strip():
        return
    if upload_payload and upload_filename:
        ok = _persist_plot_geojson_attachment(repo, plot, upload_payload, upload_filename, "application/geo+json")
        if not ok:
            logging.getLogger(__name__).warning("Anexo GeoJSON enviado pelo usuario nao foi persistido (setor %s)", plot.id)
        return
    old_n = normalize_geojson(old_geometry) if (old_geometry or "").strip() else None
    new_n = normalize_geojson(geometry) if geometry else None
    if not is_new_plot and old_n == new_n:
        return
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    fn = f"perimetro-setor-{plot.id}-{stamp}.geojson"
    body = _pretty_geojson_file_bytes(geometry)
    ok = _persist_plot_geojson_attachment(repo, plot, body, fn, "application/geo+json")
    if not ok:
        logging.getLogger(__name__).warning("GeoJSON gerado automaticamente nao foi persistido (setor %s)", plot.id)


async def _request_attachments(request: Request, field_name: str = "attachments") -> list[UploadFile]:
    form = await request.form()
    uploads: list[UploadFile] = []
    for value in form.getlist(field_name):
        if isinstance(value, (UploadFile, StarletteUploadFile)) and getattr(value, "filename", None):
            uploads.append(value)
    return uploads


def _attachment_response(filename: str, content_type: str, payload: bytes) -> Response:
    safe_name = _clean_attachment_filename(filename).replace('"', "")
    return Response(
        content=payload,
        media_type=content_type or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{safe_name}"'},
    )


def _require_admin(request: Request, user: User) -> RedirectResponse | None:
    if has_admin_access(user):
        return None
    _flash(request, "error", "Apenas administradores podem acessar este modulo.")
    return _redirect("/dashboard")


def _soil_payload(
    farm_id: int,
    plot_id: int,
    analysis_date: str,
    laboratory: str,
    ph: str | None,
    organic_matter: str | None,
    phosphorus: str | None,
    potassium: str | None,
    calcium: str | None,
    magnesium: str | None,
    aluminum: str | None,
    h_al: str | None,
    ctc: str | None,
    base_saturation: str | None,
    observations: str | None,
    pdf_filename: str | None,
    pdf_content_type: str | None,
    pdf_data: bytes | None,
    current_analysis: SoilAnalysis | None = None,
) -> dict:
    payload = {
        "farm_id": farm_id,
        "plot_id": plot_id,
        "analysis_date": analysis_date,
        "laboratory": laboratory,
        "ph": _float_or_none(ph),
        "organic_matter": _float_or_none(organic_matter),
        "phosphorus": _float_or_none(phosphorus),
        "potassium": _float_or_none(potassium),
        "calcium": _float_or_none(calcium),
        "magnesium": _float_or_none(magnesium),
        "aluminum": _float_or_none(aluminum),
        "h_al": _float_or_none(h_al),
        "ctc": _float_or_none(ctc),
        "base_saturation": _float_or_none(base_saturation),
        "observations": observations,
        "pdf_filename": pdf_filename if pdf_filename is not None else (current_analysis.pdf_filename if current_analysis else None),
        "pdf_content_type": pdf_content_type if pdf_content_type is not None else (current_analysis.pdf_content_type if current_analysis else None),
        "pdf_data": pdf_data if pdf_data is not None else (current_analysis.pdf_data if current_analysis else None),
    }
    payload.update(calculate_soil_recommendations(payload))
    return payload


def _apply_soil_ai_recommendation(repo: FarmRepository, analysis: SoilAnalysis) -> SoilAnalysis:
    refreshed = repo.get_soil_analysis(analysis.id) or analysis
    ai_result = gerar_recomendacao_adubacao(refreshed)
    return repo.update(
        refreshed,
        {
            "ai_recommendation": ai_result["recommendation"],
            "ai_status": ai_result["status"],
            "ai_model": ai_result["model"],
            "ai_error": ai_result["error"],
            "ai_generated_at": ai_result["generated_at"],
        },
    )


def _soil_history_chart(analyses: list[SoilAnalysis]) -> str:
    ordered = list(reversed(sorted(analyses, key=lambda item: (item.analysis_date, item.id))))
    return json.dumps(
        {
            "labels": [item.analysis_date.isoformat() for item in ordered],
            "ph": [float(item.ph or 0) for item in ordered],
            "phosphorus": [float(item.phosphorus or 0) for item in ordered],
            "potassium": [float(item.potassium or 0) for item in ordered],
            "calcium": [float(item.calcium or 0) for item in ordered],
            "magnesium": [float(item.magnesium or 0) for item in ordered],
            "base_saturation": [float(item.base_saturation or 0) for item in ordered],
        }
    )


@router.get("/")
def home():
    return _redirect("/dashboard")


@router.post("/contexto")
def update_global_context(
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str | None = Form(None),
    season_id: str | None = Form(None),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    validate_csrf(request, csrf_token)
    # Guarda o contexto atual completo como "anterior" antes de alterar.
    current_farm_id = _active_farm_id(request)
    current_season_id = _active_season_id(request)
    if current_farm_id and current_season_id:
        request.session[PREVIOUS_CONTEXT_SESSION_KEY] = {"farm_id": current_farm_id, "season_id": current_season_id}
    persist_user_context(request, db, user, _int_or_none(farm_id), _int_or_none(season_id))

    if not redirect_to or not redirect_to.startswith("/"):
        redirect_to = "/dashboard"
    return _redirect(redirect_to)


@router.post("/contexto/reverter")
def revert_global_context(
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    validate_csrf(request, csrf_token)
    payload = request.session.get(PREVIOUS_CONTEXT_SESSION_KEY)
    farm_id = _int_or_none(payload.get("farm_id")) if isinstance(payload, dict) else None
    season_id = _int_or_none(payload.get("season_id")) if isinstance(payload, dict) else None
    if farm_id and season_id:
        persist_user_context(request, db, user, farm_id, season_id)
    if not redirect_to or not redirect_to.startswith("/"):
        redirect_to = "/dashboard"
    return _redirect(redirect_to)


@router.get("/dashboard")
def dashboard(
    request: Request,
    rain_start_date: str | None = None,
    rain_end_date: str | None = None,
    irrigations_page: int = 1,
    rainfalls_page: int = 1,
    fertilizations_page: int = 1,
    incidents_page: int = 1,
    harvests_page: int = 1,
    forecast_page: int = 1,
    timeline_page: int = 1,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    data = build_dashboard_context(
        repo,
        rain_start_date=_date_or_none(rain_start_date),
        rain_end_date=_date_or_none(rain_end_date),
        farm_id=scope["active_farm_id"],
        season=scope["active_season"],
        pages={
            "irrigations": _page_number(irrigations_page),
            "rainfalls": _page_number(rainfalls_page),
            "fertilizations": _page_number(fertilizations_page),
            "incidents": _page_number(incidents_page),
            "harvests": _page_number(harvests_page),
            "forecast": _page_number(forecast_page),
            "timeline": _page_number(timeline_page),
        },
    )
    return templates.TemplateResponse(
        "dashboard.html",
        _base_context(
            request,
            user,
            csrf_token,
            "dashboard",
            rain_filters={
                "start_date": rain_start_date or "",
                "end_date": rain_end_date or "",
            },
            dashboard_page_values={
                "irrigations_page": _page_number(irrigations_page),
                "rainfalls_page": _page_number(rainfalls_page),
                "fertilizations_page": _page_number(fertilizations_page),
                "incidents_page": _page_number(incidents_page),
                "harvests_page": _page_number(harvests_page),
                "forecast_page": _page_number(forecast_page),
                "timeline_page": _page_number(timeline_page),
            },
            _repo=repo,
            **data,
        ),
    )

def _finance_export_query(request: Request) -> str:
    allowed_keys = {"start_date", "end_date", "schedule_range", "extract_season_id", "extract_finance_account_id"}
    params = [
        (key, value)
        for key, value in request.query_params.multi_items()
        if key in allowed_keys and str(value).strip()
    ]
    return urlencode(params, doseq=True)


def _finance_management_dataset(
    request: Request,
    repo: FarmRepository,
    *,
    flash_invalid: bool,
) -> dict:
    scope = _global_scope_context(request, repo)
    farm_id = scope.get("active_farm_id")
    active_season = scope.get("active_season")
    period_start, period_end, filter_start_str, filter_end_str = _finance_extract_period_bounds(
        request,
        flash_invalid=flash_invalid,
    )
    selected_finance_range = (
        _fertilization_filter_range_preset(
            request.query_params.get("schedule_range"),
            filter_start_str,
            filter_end_str,
        )
        if _period_filter_explicit_in_query(request)
        else ""
    )
    extract_season_q = _int_or_none(request.query_params.get("extract_season_id"))
    extract_finance_account_id = _int_or_none(request.query_params.get("extract_finance_account_id"))
    period_start_for_extract, period_end_for_extract, finance_filter_season_id, extract_range_empty = (
        _finance_extract_apply_season_bounds(
            repo,
            farm_id,
            period_start=period_start,
            period_end=period_end,
            extract_season_id=extract_season_q,
        )
    )
    finance_filters_active = _period_filter_explicit_in_query(request) or bool(finance_filter_season_id) or bool(extract_finance_account_id)
    finance_filter_clear_url = _url_with_query(
        request,
        start_date=None,
        end_date=None,
        schedule_range=None,
        extract_season_id=None,
        extract_finance_account_id=None,
    )
    finance_season_options = repo.list_crop_seasons(farm_id=farm_id) if farm_id else []
    finance_filter_season = repo.get_crop_season(finance_filter_season_id) if finance_filter_season_id else None
    finance_account_options = repo.list_finance_accounts(farm_id=farm_id) if farm_id else []
    finance_filter_account = next((item for item in finance_account_options if item.id == extract_finance_account_id), None)
    if extract_finance_account_id and not finance_filter_account:
        extract_finance_account_id = None
    finance_data = build_finance_overview_context(
        repo,
        farm_id=farm_id,
        active_season=active_season,
    )
    finance_data["finance_currency"] = {
        "inventory_total": _format_currency(finance_data["inventory_value_total"]),
        "inventory_insumo": _format_currency(finance_data["inventory_value_insumo"]),
        "inventory_suprimento": _format_currency(finance_data["inventory_value_suprimento"]),
        "purchase_insumo": _format_currency(finance_data["historical_purchase_cost_insumo"]),
        "purchase_suprimento": _format_currency(finance_data["historical_purchase_cost_suprimento"]),
        "purchase_total": _format_currency(finance_data["historical_purchase_cost_total"]),
        "stock_out_season": _format_currency(finance_data["stock_output_cost_season"]),
        "fertilization_season": _format_currency(finance_data["fertilization_cost_season"]),
        "operational_season": _format_currency(finance_data["operational_cost_season"]),
        "assets": _format_currency(finance_data["assets_acquisition_total"]),
    }
    if extract_range_empty:
        extract_rows, extract_truncated = [], False
    else:
        extract_rows, extract_truncated = build_finance_extract_rows(
            repo,
            farm_id=farm_id,
            period_start=period_start_for_extract,
            period_end=period_end_for_extract,
            finance_account_id=extract_finance_account_id,
        )
    summary_credit_total = round(sum(float(row.get("credit") or 0) for row in extract_rows), 2)
    summary_debit_total = round(sum(float(row.get("debit") or 0) for row in extract_rows), 2)
    summary_balance_total = round(summary_credit_total - summary_debit_total, 2)
    finance_data["finance_extract_rows"] = []
    for r in extract_rows:
        bal = float(r["balance"] or 0)
        if bal < 0:
            balance_class = "text-rose-600"
        elif bal > 0:
            balance_class = "text-blue-600"
        else:
            balance_class = "text-slate-600"
        finance_data["finance_extract_rows"].append(
            {
                "date_label": r["date"].strftime("%d/%m/%Y"),
                "module": r["module"],
                "description": r["description"],
                "detail": (r.get("detail") or "").strip(),
                "debit_fmt": _format_currency(r["debit"]) if r.get("debit") else "—",
                "credit_fmt": _format_currency(r["credit"]) if r.get("credit") else "—",
                "balance_fmt": _format_currency(r["balance"]),
                "balance_class": balance_class,
            }
        )
    finance_data["finance_extract_truncated"] = extract_truncated
    finance_data["finance_filter_start_date"] = filter_start_str
    finance_data["finance_filter_end_date"] = filter_end_str
    finance_data["selected_finance_range"] = selected_finance_range
    finance_data["finance_filters_active"] = finance_filters_active
    finance_data["finance_filter_clear_url"] = finance_filter_clear_url
    finance_data["finance_season_options"] = finance_season_options
    finance_data["finance_filter_season_id"] = finance_filter_season_id
    finance_data["finance_filter_season"] = finance_filter_season
    finance_data["finance_account_options"] = finance_account_options
    finance_data["finance_filter_account_id"] = extract_finance_account_id
    finance_data["finance_filter_account"] = finance_filter_account
    finance_data["finance_export_query"] = _finance_export_query(request)
    finance_data["finance_extract_rows_raw"] = extract_rows
    finance_data["finance_summary"] = {
        "balance_total": summary_balance_total,
        "credit_total": summary_credit_total,
        "debit_total": summary_debit_total,
        "future_total": 0.0,
        "balance_fmt": _format_currency(summary_balance_total),
        "credit_fmt": _format_currency(summary_credit_total),
        "debit_fmt": _format_currency(summary_debit_total),
        "future_fmt": _format_currency(0),
        "balance_tone": "positive" if summary_balance_total >= 0 else "negative",
    }
    finance_data["scope"] = scope
    return finance_data


def _finance_bank_option_map() -> dict[str, dict]:
    return {item["code"]: item for item in FINANCE_BANK_OPTIONS}


def _finance_builtin_bank_codes() -> set[str]:
    return {str(item["code"]).strip() for item in FINANCE_BANK_OPTIONS}


def _finance_bank_choice_options(repo: FarmRepository) -> list[dict]:
    options = [dict(item, source="builtin") for item in FINANCE_BANK_OPTIONS]
    for custom in repo.list_finance_custom_banks():
        options.append(
            {
                "code": custom.bank_code,
                "name": custom.bank_name,
                "mark": "OT",
                "bg": "#94a3b8",
                "fg": "#ffffff",
                "source": "custom",
                "custom_id": custom.id,
            }
        )
    options.append({"code": "__other__", "name": "Outro banco", "mark": "+", "bg": "#e2e8f0", "fg": "#334155", "source": "other"})
    return options


def _finance_accounts_modal_query(request: Request, *, edit_id: int | None = None, launch: bool | None = None) -> str:
    params = dict(request.query_params)
    if edit_id is None:
        params.pop("edit_id", None)
    else:
        params["edit_id"] = str(edit_id)
    if launch:
        params["launch"] = "1"
    else:
        params.pop("launch", None)
    if params:
        return f"/gestao-financeira/contas?{urlencode(params)}"
    return "/gestao-financeira/contas"


def _finance_transactions_modal_query(
    request: Request,
    *,
    edit_id: int | None = None,
    launch: bool | None = None,
) -> str:
    params = dict(request.query_params)
    if edit_id is None:
        params.pop("transaction_edit_id", None)
    else:
        params["transaction_edit_id"] = str(edit_id)
    if launch:
        params["transaction_launch"] = "1"
    else:
        params.pop("transaction_launch", None)
    if params:
        return f"/gestao-financeira/contas?{urlencode(params)}"
    return "/gestao-financeira/contas"


def _finance_accounts_set_default(repo: FarmRepository, farm_id: int, keep_id: int | None = None) -> None:
    for account in repo.list_finance_accounts(farm_id=farm_id):
        if keep_id is not None and account.id == keep_id:
            continue
        if account.is_default:
            account.is_default = False
            repo.db.add(account)


def _save_finance_transaction_attachments(
    repo: FarmRepository,
    transaction: FinanceTransaction,
    attachments: list[tuple[str, str, bytes]],
) -> int:
    if not attachments:
        return 0
    repo.db.add_all(
        [
            FinanceTransactionAttachment(
                finance_transaction_id=transaction.id,
                filename=filename,
                content_type=content_type,
                file_data=payload,
            )
            for filename, content_type, payload in attachments
        ]
    )
    try:
        repo.db.commit()
    except Exception:
        repo.db.rollback()
        raise
    repo.db.refresh(transaction)
    return len(attachments)


def _finance_transaction_category_options(operation_type: str | None) -> list[str]:
    normalized = (operation_type or "").strip().lower()
    if normalized == "receita":
        return FINANCE_TRANSACTION_REVENUE_CATEGORIES
    return FINANCE_TRANSACTION_EXPENSE_CATEGORIES


def _parse_finance_transaction_amount(raw_value: str | None) -> float:
    try:
        amount = round(abs(float(raw_value or 0)), 2)
    except (TypeError, ValueError):
        raise ValueError("Informe um valor válido para o lançamento.")
    if amount <= 0:
        raise ValueError("Informe um valor maior que zero para o lançamento.")
    return amount


def _resolve_finance_transaction_account(
    repo: FarmRepository,
    *,
    active_farm: Farm,
    account_id: str | None,
) -> FinanceAccount:
    resolved_id = int(account_id) if str(account_id or "").strip().isdigit() else None
    account = repo.get_finance_account(resolved_id) if resolved_id else None
    if not account:
        account = next((item for item in repo.list_finance_accounts(farm_id=active_farm.id) if item.is_default), None)
    if not account:
        raise ValueError("Selecione a conta de lançamento.")
    if account.farm_id != active_farm.id:
        raise ValueError("A conta selecionada não pertence à fazenda ativa.")
    return account


def _resolve_optional_finance_account(
    repo: FarmRepository,
    *,
    active_farm: Farm | None,
    account_id: str | int | None,
) -> FinanceAccount | None:
    if not active_farm:
        return None
    resolved_id = _int_or_none(account_id)
    if resolved_id:
        account = repo.get_finance_account(resolved_id)
        if not account or account.farm_id != active_farm.id:
            raise ValueError("A conta selecionada não pertence à fazenda ativa.")
        return account
    return next((item for item in repo.list_finance_accounts(farm_id=active_farm.id) if item.is_default), None)


def _finance_transaction_payload(
    repo: FarmRepository,
    *,
    active_farm: Farm,
    operation_type: str,
    launch_date: str,
    amount: str,
    finance_account_id: str | None,
    payment_condition: str | None,
    installment_count: str | None,
    installment_frequency: str | None,
    first_installment_date: str | None,
    category: str,
    product_service: str,
    description: str | None,
    counterparty_name: str | None,
    document_number: str | None,
    payment_method: str | None,
    notes: str | None,
) -> dict:
    normalized_type = (operation_type or "").strip().lower()
    if normalized_type not in {"despesa", "receita"}:
        raise ValueError("Selecione o tipo de operação financeira.")
    if not launch_date:
        raise ValueError("Informe a data do lançamento.")
    try:
        parsed_launch_date = date.fromisoformat(launch_date)
    except ValueError:
        raise ValueError("Informe uma data válida para o lançamento.")
    category_value = _clean_text(category)
    if category_value not in _finance_transaction_category_options(normalized_type):
        raise ValueError("Selecione uma categoria válida para o lançamento.")
    product_service_value = _clean_text(product_service)
    if not product_service_value:
        raise ValueError("Informe o produto ou serviço.")
    account = _resolve_finance_transaction_account(repo, active_farm=active_farm, account_id=finance_account_id)
    normalized_condition = (payment_condition or "a_vista").strip().lower()
    if normalized_condition not in {"a_vista", "a_prazo"}:
        raise ValueError("Selecione uma condição de pagamento válida.")
    resolved_installment_count = 1
    resolved_frequency = None
    resolved_first_installment_date = None
    if normalized_condition == "a_prazo":
        try:
            resolved_installment_count = int(str(installment_count or "").strip())
        except ValueError:
            raise ValueError("Informe a quantidade de parcelas para pagamento a prazo.")
        if resolved_installment_count < 2:
            raise ValueError("Para pagamento a prazo, informe ao menos 2 parcelas.")
        resolved_frequency = (installment_frequency or "").strip().lower()
        if resolved_frequency not in {"mensal", "anual"}:
            raise ValueError("Selecione a periodicidade das parcelas.")
        if not first_installment_date:
            raise ValueError("Informe a data da primeira parcela.")
        try:
            resolved_first_installment_date = date.fromisoformat(first_installment_date)
        except ValueError:
            raise ValueError("Informe uma data válida para a primeira parcela.")
    else:
        resolved_first_installment_date = parsed_launch_date
    return {
        "farm_id": active_farm.id,
        "finance_account_id": account.id,
        "operation_type": normalized_type,
        "launch_date": parsed_launch_date,
        "amount": _parse_finance_transaction_amount(amount),
        "payment_condition": normalized_condition,
        "installment_count": resolved_installment_count,
        "installment_frequency": resolved_frequency,
        "first_installment_date": resolved_first_installment_date,
        "category": category_value,
        "product_service": product_service_value,
        "description": _clean_text(description),
        "counterparty_name": _clean_text(counterparty_name),
        "document_number": _clean_text(document_number),
        "payment_method": _clean_text(payment_method),
        "notes": _clean_text(notes),
    }


def _finance_add_months(base_date: date, months: int) -> date:
    year = base_date.year + ((base_date.month - 1 + months) // 12)
    month = ((base_date.month - 1 + months) % 12) + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _finance_add_years(base_date: date, years: int) -> date:
    target_year = base_date.year + years
    day = min(base_date.day, calendar.monthrange(target_year, base_date.month)[1])
    return date(target_year, base_date.month, day)


def _build_finance_transaction_installments(
    *,
    amount: float,
    payment_condition: str,
    installment_count: int,
    installment_frequency: str | None,
    first_installment_date: date | None,
) -> list[dict]:
    if payment_condition != "a_prazo" or installment_count <= 1 or not first_installment_date:
        return []
    total_cents = int((Decimal(str(amount)) * 100).quantize(Decimal("1")))
    base_cents = total_cents // installment_count
    remainder = total_cents % installment_count
    installments: list[dict] = []
    for index in range(installment_count):
        current_cents = base_cents + (1 if index < remainder else 0)
        if installment_frequency == "anual":
            due_date = _finance_add_years(first_installment_date, index)
        else:
            due_date = _finance_add_months(first_installment_date, index)
        installments.append(
            {
                "installment_number": index + 1,
                "due_date": due_date,
                "amount": round(current_cents / 100, 2),
                "status": "pendente",
            }
        )
    return installments


def _reject_finance_schedule_change_if_installment_paid(
    transaction: FinanceTransaction,
    payload: dict,
) -> None:
    """Impede alterar tipo, valor, condição ou cronograma via POST quando já existe parcela paga (além da UI)."""
    any_paid = any(
        (getattr(i, "status", None) or "").strip().lower() == "pago"
        for i in (transaction.installments or [])
    )
    if not any_paid:
        return
    if (payload["operation_type"] or "").strip().lower() != (transaction.operation_type or "").strip().lower():
        raise ValueError(
            "Com parcela já quitada, o tipo de operação não pode ser alterado. Estorne os pagamentos das parcelas antes."
        )
    tr_amt = Decimal(str(transaction.amount))
    pl_amt = Decimal(str(payload["amount"]))
    if tr_amt.quantize(Decimal("0.01")) != pl_amt.quantize(Decimal("0.01")):
        raise ValueError(
            "Com parcela já quitada, o valor não pode ser alterado. Estorne os pagamentos das parcelas antes."
        )
    prior_pc = (transaction.payment_condition or "a_vista").strip().lower()
    if payload["payment_condition"] != prior_pc:
        raise ValueError(
            "Com parcela já paga, a condição de pagamento não pode ser alterada. Estorne os pagamentos das parcelas antes."
        )
    if prior_pc != "a_prazo":
        return
    if payload["installment_count"] != transaction.installment_count:
        raise ValueError(
            "Com parcela já paga, a quantidade e o cronograma de parcelas não podem ser alterados. Estorne os pagamentos antes."
        )
    tr_freq = (transaction.installment_frequency or "mensal").strip().lower()
    pl_freq = (payload["installment_frequency"] or "mensal").strip().lower()
    if tr_freq != pl_freq:
        raise ValueError(
            "Com parcela já paga, a periodicidade das parcelas não pode ser alterada. Estorne os pagamentos antes."
        )
    tr_first = transaction.first_installment_date
    pl_first = payload["first_installment_date"]
    if tr_first != pl_first:
        raise ValueError(
            "Com parcela já paga, a data da primeira parcela não pode ser alterada. Estorne os pagamentos antes."
        )


def _snapshot_finance_installments_for_edit(transaction: FinanceTransaction) -> list[dict]:
    """Cópia dos dados de parcelas antes de recriar linhas (evita perder pagamento ao editar)."""
    out: list[dict] = []
    for inst in sorted(transaction.installments or [], key=lambda x: (x.installment_number or 0, x.id)):
        out.append(
            {
                "installment_number": inst.installment_number,
                "status": inst.status,
                "paid_at": inst.paid_at,
                "payment_notes": inst.payment_notes,
            }
        )
    return out


def _validate_installment_edit_against_prior(
    new_rows: list[dict],
    prior_snapshot: list[dict],
) -> None:
    """
    Impede remover parcelas já pagas ou 'apagar' parcelamento com histórico de pagamento
    (ex.: mudar para à vista) sem estornar antes — padrão em ERPs.
    """
    if not prior_snapshot:
        return
    any_paid = any((row.get("status") or "").strip().lower() == "pago" for row in prior_snapshot)
    if not new_rows:
        if any_paid:
            raise ValueError(
                "Não é possível alterar para à vista ou remover o parcelamento enquanto houver parcelas pagas. "
                "Estorne os pagamentos no extrato de contas a pagar ou mantenha o lançamento à prazo."
            )
        return
    new_nums = {row["installment_number"] for row in new_rows}
    for prev in prior_snapshot:
        if (prev.get("status") or "").strip().lower() != "pago":
            continue
        num = prev.get("installment_number")
        if num not in new_nums:
            raise ValueError(
                "A nova configuração remove parcelas que já estão pagas. "
                "Mantenha quantidade de parcelas suficiente para incluir todas as parcelas já quitadas ou estorne pagamentos antes."
            )


def _merge_paid_installment_state_into_rows(new_rows: list[dict], prior_snapshot: list[dict]) -> None:
    """Preserva status, data e observações de parcelas pagas quando o número da parcela ainda existe."""
    prior_by_num = {row["installment_number"]: row for row in prior_snapshot}
    for row in new_rows:
        prior = prior_by_num.get(row["installment_number"])
        if not prior:
            continue
        if (prior.get("status") or "").strip().lower() != "pago":
            continue
        row["status"] = "pago"
        row["paid_at"] = prior.get("paid_at")
        row["payment_notes"] = prior.get("payment_notes")


def _replace_finance_transaction_installments(
    repo: FarmRepository,
    transaction: FinanceTransaction,
    installment_rows: list[dict],
) -> None:
    for installment in list(transaction.installments or []):
        repo.delete(installment)
    for row in installment_rows:
        repo.create(
            FinanceTransactionInstallment(
                finance_transaction_id=transaction.id,
                installment_number=row["installment_number"],
                due_date=row["due_date"],
                amount=row["amount"],
                status=row["status"],
                paid_at=row.get("paid_at"),
                payment_notes=row.get("payment_notes"),
            )
        )


def _finance_payables_period_ui_mode(payables_status: str) -> str:
    """Modo do picker de período na aba A pagar: past | future | custom_only."""
    if payables_status == "open":
        return "future"
    if payables_status == "all":
        return "custom_only"
    return "past"


def _payables_redirect_url(request: Request) -> str:
    """URL atual sem payables_partial (redirect após POST de cancelar pagamento)."""
    items = [(k, v) for k, v in request.query_params.multi_items() if k != "payables_partial"]
    path = str(request.url.path)
    if not items:
        return path
    return f"{path}?{urlencode(items)}"


def _finance_transaction_balance_amount_chunks(transaction: FinanceTransaction) -> list[float]:
    """
    Montantes absolutos a aplicar no saldo do card da conta: uma parcela paga por vez se a prazo;
    valor integral do lançamento se à vista. Alinhado ao extrato (finance_overview).
    """
    if (transaction.payment_condition or "").strip().lower() == "a_prazo":
        chunks: list[float] = []
        for inst in sorted(
            transaction.installments or [],
            key=lambda x: (x.installment_number or 0, x.id),
        ):
            if (inst.status or "").strip().lower() != "pago":
                continue
            amt = abs(float(inst.amount or 0))
            if amt > 0:
                chunks.append(round(amt, 2))
        return chunks
    amt = abs(float(transaction.amount or 0))
    if amt <= 0:
        return []
    return [round(amt, 2)]


def _finance_payables_period_bounds(request: Request) -> tuple[date | None, date | None, str, str, str]:
    """Retorna (start, end, raw_start, raw_end, selected_range) para filtro de vencimento em A pagar."""
    qp = request.query_params
    selected_range = (qp.get("schedule_range") or "").strip()
    raw_start = (qp.get("start_date") or "").strip()
    raw_end = (qp.get("end_date") or "").strip()
    payables_status = (qp.get("payables_status") or "").strip().lower()
    if payables_status not in {"", "open", "paid", "overdue", "all"}:
        payables_status = ""
    today = today_in_app_timezone()

    past_only = {"last_10_days", "last_20_days", "last_month"}
    future_only = {"next_10_days", "next_20_days", "next_month"}
    if payables_status == "open" and selected_range in past_only:
        selected_range = "current_month"
    if payables_status in ("paid", "overdue", "") and selected_range in future_only:
        selected_range = "current_month"
    if payables_status == "all" and selected_range and selected_range != "custom":
        selected_range = "custom"
        raw_start, raw_end = "", ""

    if payables_status == "all":
        if not selected_range and not raw_start and not raw_end:
            return None, None, "", "", "custom"
        if selected_range == "custom" or raw_start or raw_end:
            start_date, end_date, raw_start, raw_end = _finance_extract_period_bounds(request, flash_invalid=False)
            return start_date, end_date, raw_start, raw_end, "custom"
        return None, None, "", "", "custom"

    if not selected_range and not raw_start and not raw_end:
        month_end = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1) - timedelta(days=1)
        return date(today.year, today.month, 1), month_end, "", "", "current_month"
    if selected_range == "current_month":
        month_end = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1) - timedelta(days=1)
        return date(today.year, today.month, 1), month_end, "", "", "current_month"
    if selected_range == "last_10_days":
        return today - timedelta(days=10), today, "", "", "last_10_days"
    if selected_range == "last_20_days":
        return today - timedelta(days=20), today, "", "", "last_20_days"
    if selected_range == "last_month":
        current_month_start = date(today.year, today.month, 1)
        previous_month_end = current_month_start - timedelta(days=1)
        previous_month_start = date(previous_month_end.year, previous_month_end.month, 1)
        return previous_month_start, previous_month_end, "", "", "last_month"
    if selected_range == "next_10_days":
        return today, today + timedelta(days=10), "", "", "next_10_days"
    if selected_range == "next_20_days":
        return today, today + timedelta(days=20), "", "", "next_20_days"
    if selected_range == "next_month":
        if today.month == 12:
            nm_start = date(today.year + 1, 1, 1)
        else:
            nm_start = date(today.year, today.month + 1, 1)
        nm_end = date(nm_start.year + (1 if nm_start.month == 12 else 0), 1 if nm_start.month == 12 else nm_start.month + 1, 1) - timedelta(days=1)
        return nm_start, nm_end, "", "", "next_month"
    start_date, end_date, raw_start, raw_end = _finance_extract_period_bounds(request, flash_invalid=False)
    return start_date, end_date, raw_start, raw_end, (selected_range or ("custom" if raw_start or raw_end else "current_month"))


def _cleanup_custom_bank_if_unused(repo: FarmRepository, custom_bank_id: int | None) -> None:
    if not custom_bank_id:
        return
    custom_bank = repo.get_finance_custom_bank(custom_bank_id)
    if not custom_bank:
        return
    remaining = [item for item in repo.list_finance_accounts() if item.custom_bank_id == custom_bank_id]
    if not remaining:
        repo.delete(custom_bank)


def _normalize_finance_account_identity(value: str | None) -> str:
    return "".join(char for char in str(value or "").strip() if char.isalnum()).upper()


def _validate_finance_account_uniqueness(
    repo: FarmRepository,
    *,
    farm_id: int,
    branch_number: str | None,
    account_number: str | None,
    ignore_id: int | None = None,
) -> None:
    normalized_branch = _normalize_finance_account_identity(branch_number)
    normalized_account = _normalize_finance_account_identity(account_number)
    if not normalized_branch or not normalized_account:
        return
    for account in repo.list_finance_accounts(farm_id=farm_id):
        if ignore_id is not None and account.id == ignore_id:
            continue
        if (
            _normalize_finance_account_identity(account.branch_number) == normalized_branch
            and _normalize_finance_account_identity(account.account_number) == normalized_account
        ):
            raise ValueError("Já existe uma conta cadastrada com a mesma agência e número da conta.")


def _find_matching_custom_bank(repo: FarmRepository, bank_code: str, bank_name: str) -> FinanceCustomBank | None:
    normalized_code = (bank_code or "").strip()
    normalized_name = unicodedata.normalize("NFD", (bank_name or "").strip())
    normalized_name = "".join(char for char in normalized_name if unicodedata.category(char) != "Mn").lower()
    for custom_bank in repo.list_finance_custom_banks():
        current_name = unicodedata.normalize("NFD", custom_bank.bank_name or "")
        current_name = "".join(char for char in current_name if unicodedata.category(char) != "Mn").lower()
        if (custom_bank.bank_code or "").strip() == normalized_code and current_name == normalized_name:
            return custom_bank
    return None


def _finance_accounts_parse_form(
    account_name: str,
    initial_balance_date: str,
    initial_balance: str,
    bank_code: str,
    bank_name: str,
    custom_bank_code: str | None,
    custom_bank_name: str | None,
    branch_number: str | None,
    account_number: str | None,
    is_default: bool,
    repo: FarmRepository,
) -> dict:
    bank_map = _finance_bank_option_map()
    normalized_bank_code = (bank_code or "").strip()
    selected_bank = bank_map.get(normalized_bank_code)
    if not account_name.strip():
        raise ValueError("Informe o nome da conta bancária.")
    if not initial_balance_date:
        raise ValueError("Informe a data do saldo inicial.")
    try:
        parsed_initial_balance = round(float(initial_balance or 0), 2)
    except (TypeError, ValueError):
        raise ValueError("Informe um saldo inicial válido.")
    payload = {
        "account_name": account_name.strip(),
        "initial_balance_date": date.fromisoformat(initial_balance_date),
        "initial_balance": parsed_initial_balance,
        "branch_number": (branch_number or "").strip(),
        "account_number": (account_number or "").strip(),
        "is_default": bool(is_default),
        "custom_bank": None,
    }
    if normalized_bank_code == "__other__":
        resolved_custom_code = (custom_bank_code or "").strip()
        resolved_custom_name = (custom_bank_name or "").strip()
        if not resolved_custom_code:
            raise ValueError("Informe o código do outro banco.")
        if not resolved_custom_name:
            raise ValueError("Informe o nome do outro banco.")
        if resolved_custom_code in _finance_builtin_bank_codes():
            raise ValueError("O código informado já existe na lista de bancos padrão.")
        existing_custom_bank = _find_matching_custom_bank(repo, resolved_custom_code, resolved_custom_name)
        if existing_custom_bank:
            payload.update(
                {
                    "bank_code": existing_custom_bank.bank_code,
                    "bank_name": existing_custom_bank.bank_name,
                    "custom_bank": existing_custom_bank,
                }
            )
            return payload
        payload.update(
            {
                "bank_code": resolved_custom_code,
                "bank_name": resolved_custom_name,
                "custom_bank": {"bank_code": resolved_custom_code, "bank_name": resolved_custom_name},
            }
        )
        return payload

    if selected_bank:
        payload.update({"bank_code": selected_bank["code"], "bank_name": selected_bank["name"]})
        return payload

    normalized_bank_name = (bank_name or "").strip()
    custom_bank = _find_matching_custom_bank(repo, normalized_bank_code, normalized_bank_name)
    if not custom_bank:
        raise ValueError("Selecione um banco da lista.")
    return {
        **payload,
        "bank_code": custom_bank.bank_code,
        "bank_name": custom_bank.bank_name,
        "custom_bank": custom_bank,
    }


@router.get("/gestao-financeira")
def finance_management_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    finance_data = _finance_management_dataset(request, repo, flash_invalid=True)
    return templates.TemplateResponse(
        "finance_management.html",
        _base_context(
            request,
            user,
            csrf_token,
            "finance_management",
            title="Extrato",
            _repo=repo,
            **finance_data,
        ),
    )


@router.get("/gestao-financeira/exportar.xlsx")
def export_finance_management_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    finance_data = _finance_management_dataset(request, repo, flash_invalid=False)
    rows = finance_data["finance_extract_rows_raw"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extrato financeiro"
    sheet.append(["Data", "Origem", "Historico", "Detalhe", "Debito", "Credito", "Acumulado"])
    for row in rows:
        sheet.append(
            [
                row["date"].isoformat() if row.get("date") else "",
                row.get("module") or "",
                row.get("description") or "",
                (row.get("detail") or "").strip(),
                float(row.get("debit") or 0),
                float(row.get("credit") or 0),
                float(row.get("balance") or 0),
            ]
        )
    for index, width in enumerate([14, 18, 36, 34, 16, 16, 16], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="gestao_financeira.xlsx"'},
    )


@router.get("/gestao-financeira/exportar.pdf")
def export_finance_management_pdf(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    finance_data = _finance_management_dataset(request, repo, flash_invalid=False)
    rows = finance_data["finance_extract_rows_raw"]
    scope = finance_data["scope"]

    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    farm_name = scope["active_farm"].name if scope.get("active_farm") else "Fazenda Bela Vista"
    season_label = finance_data["finance_filter_season"].name if finance_data.get("finance_filter_season") else "Todas as safras"
    period_label = "Histórico completo"
    if finance_data["finance_filter_start_date"] or finance_data["finance_filter_end_date"]:
        period_label = f"{_format_iso_date_br(finance_data['finance_filter_start_date'])} a {_format_iso_date_br(finance_data['finance_filter_end_date'])}"
    total_debit = sum(float(row.get("debit") or 0) for row in rows)
    total_credit = sum(float(row.get("credit") or 0) for row in rows)
    final_balance = float(rows[-1]["balance"]) if rows else 0.0

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle("FinancePdfFarmHeader", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, alignment=TA_RIGHT, textColor=colors.HexColor("#1e293b"))
    meta_label_style = ParagraphStyle("FinancePdfMetaLabel", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#446a36"), spaceAfter=2)
    meta_value_style = ParagraphStyle("FinancePdfMetaValue", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13, textColor=colors.HexColor("#334155"))
    cell_style = ParagraphStyle("FinancePdfCell", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.2, leading=10.4, textColor=colors.HexColor("#0f172a"))
    cell_muted_style = ParagraphStyle("FinancePdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    cell_numeric_style = ParagraphStyle("FinancePdfCellNumeric", parent=cell_style, alignment=TA_RIGHT)
    summary_value_style = ParagraphStyle("FinancePdfSummaryValue", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#0f172a"))

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
    summary_table = Table([
        [
            [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
            [Paragraph("SAFRA (FILTRO)", meta_label_style), Paragraph(season_label, meta_value_style)],
            [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
        ],
        [
            [Paragraph("LANÇAMENTOS", meta_label_style), Paragraph(str(len(rows)), summary_value_style)],
            [Paragraph("DÉBITOS", meta_label_style), Paragraph(_format_currency(total_debit), summary_value_style)],
            [Paragraph("SALDO FINAL", meta_label_style), Paragraph(_format_currency(final_balance), summary_value_style)],
        ],
    ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    summary_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")), ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")), ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))

    title_style = ParagraphStyle("FinancePdfTitle", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=colors.HexColor("#0f172a"), spaceAfter=6)
    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 12), Paragraph("Extrato financeiro", title_style), Spacer(1, 8)]
    if rows:
        data_rows = [["Data", "Origem", "Histórico", "Débito", "Crédito", "Acumulado"]]
        for row in rows:
            description = row.get("description") or "-"
            detail = (row.get("detail") or "").strip()
            if detail:
                description = f"{description}<br/><font color='#64748b'>{detail}</font>"
            data_rows.append([
                Paragraph(row["date"].strftime("%d/%m/%Y") if row.get("date") else "-", cell_style),
                Paragraph(row.get("module") or "-", cell_style),
                Paragraph(description, cell_muted_style),
                Paragraph(_format_currency(row.get("debit") or 0) if row.get("debit") else "—", cell_numeric_style),
                Paragraph(_format_currency(row.get("credit") or 0) if row.get("credit") else "—", cell_numeric_style),
                Paragraph(_format_currency(row.get("balance") or 0), cell_numeric_style),
            ])
        table = Table(data_rows, colWidths=[64, 88, doc.width - 64 - 88 - 74 - 74 - 82, 74, 74, 82], repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbe5dd")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
    else:
        empty_style = ParagraphStyle("FinancePdfEmpty", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=14, textColor=colors.HexColor("#475569"))
        elements.append(Paragraph("Nenhum lançamento encontrado para o período selecionado.", empty_style))

    elements.append(Spacer(1, 14))
    footer_summary = Table([
        ["Créditos", _format_currency(total_credit)],
        ["Débitos", _format_currency(total_debit)],
        ["Saldo final", _format_currency(final_balance)],
    ], colWidths=[doc.width * 0.28, doc.width * 0.18], hAlign="LEFT")
    footer_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(footer_summary)
    generated_at_label = format_app_datetime(generated_at)

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="gestao_financeira.pdf"'},
    )


@router.get("/gestao-financeira/contas")
def finance_accounts_page(
    request: Request,
    edit_id: int | None = None,
    launch: int | None = None,
    transaction_edit_id: int | None = None,
    transaction_launch: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    accounts = repo.list_finance_accounts(farm_id=active_farm.id) if active_farm else []
    transactions = repo.list_finance_transactions(farm_id=active_farm.id) if active_farm else []
    transactions_pagination = _paginate_collection(request, transactions, "transactions_page") if active_farm else _paginate_collection(request, [], "transactions_page")
    payables_start_date, payables_end_date, payables_filter_start_date, payables_filter_end_date, selected_payables_range = _finance_payables_period_bounds(request)
    payables_search = (request.query_params.get("payables_search") or "").strip()
    payables_status = (request.query_params.get("payables_status") or "").strip().lower()
    if payables_status not in {"", "open", "paid", "overdue", "all"}:
        payables_status = ""
    finance_payables_filters_active = bool(
        payables_search
        or request.query_params.get("schedule_range")
        or request.query_params.get("start_date")
        or request.query_params.get("end_date")
        or bool(payables_status)
    )
    payables_clear_params = dict(request.query_params)
    for key in ("payables_search", "schedule_range", "start_date", "end_date", "payables_page", "payables_status"):
        payables_clear_params.pop(key, None)
    payables_clear_params["finance_tab"] = "payables"
    finance_payables_clear_url = f"/gestao-financeira/contas?{urlencode(payables_clear_params)}" if payables_clear_params else "/gestao-financeira/contas?finance_tab=payables"
    payable_rows: list[dict] = []
    today = today_in_app_timezone()
    for transaction in transactions:
        if (transaction.operation_type or "").strip().lower() != "despesa":
            continue
        for installment in sorted(transaction.installments or [], key=lambda item: (item.due_date or date.max, item.installment_number or 0)):
            status = (installment.status or "pendente").strip().lower()
            due_date = installment.due_date
            payable_status = "open"
            payable_label = "Em aberto"
            if status == "pago":
                payable_status = "paid"
                payable_label = "Pago"
            elif due_date and due_date < today:
                payable_status = "overdue"
                payable_label = "Atrasado"
            elif due_date and due_date == today:
                payable_status = "today"
                payable_label = "Vence hoje"
            if not payables_status and payable_status == "paid":
                continue
            if payables_status == "open" and payable_status not in {"open", "today"}:
                continue
            if payables_status == "paid" and payable_status != "paid":
                continue
            if payables_status == "overdue" and payable_status != "overdue":
                continue
            if payables_start_date and due_date and due_date < payables_start_date:
                continue
            if payables_end_date and due_date and due_date > payables_end_date:
                continue
            searchable = " ".join(
                [
                    transaction.finance_account.account_name if transaction.finance_account else "",
                    transaction.category or "",
                    transaction.product_service or "",
                    transaction.description or "",
                    installment.due_date.strftime("%d/%m/%Y") if installment.due_date else "",
                    payable_label,
                ]
            ).strip()
            if payables_search and _normalize_search_value(payables_search) not in _normalize_search_value(searchable):
                continue
            payable_rows.append(
                {
                    "transaction": transaction,
                    "installment": installment,
                    "status": payable_status,
                    "status_label": payable_label,
                    "status_chip_class": {
                        "paid": "stock-movement-chip-entry",
                        "overdue": "stock-movement-chip-output",
                        "today": "stock-movement-chip-warning",
                        "open": "stock-movement-chip-info",
                    }.get(payable_status, "stock-movement-chip-info"),
                    "row_kind": {
                        "paid": "entrada",
                        "overdue": "saida",
                        "today": "warning",
                        "open": "info",
                    }.get(payable_status, "info"),
                    "installment_label": f"{installment.installment_number}/{transaction.installment_count or 1}",
                    "search_text": searchable,
                }
            )
    payable_rows.sort(key=lambda row: (row["installment"].due_date or date.max, row["transaction"].id))
    payables_pagination = _paginate_collection(request, payable_rows, "payables_page") if active_farm else _paginate_collection(request, [], "payables_page")
    if (request.query_params.get("payables_partial") or "").strip() == "1":
        if not active_farm:
            return HTMLResponse(content="", status_code=204)
        tab = (request.query_params.get("finance_tab") or "").strip()
        if tab not in ("", "payables"):
            return HTMLResponse(content="Invalid tab", status_code=400)
        return templates.TemplateResponse(
            "partials/finance_payables_list.html",
            _base_context(
                request,
                user,
                csrf_token,
                "finance_accounts",
                title="Contas",
                _repo=repo,
                finance_payables=payables_pagination["items"],
                finance_payables_pagination=payables_pagination,
                finance_payables_redirect_to=_payables_redirect_url(request),
                today=today,
            ),
        )
    edit_account = repo.get_finance_account(edit_id) if edit_id else None
    edit_transaction = repo.get_finance_transaction(transaction_edit_id) if transaction_edit_id else None
    bank_options = _finance_bank_choice_options(repo)
    if edit_account and (not active_farm or edit_account.farm_id != active_farm.id):
        _flash(request, "error", "Esta conta não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    if edit_transaction and (not active_farm or edit_transaction.farm_id != active_farm.id):
        _flash(request, "error", "Este lançamento não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    account_balances: dict[int, float] = {item.id: round(float(item.initial_balance or 0), 2) for item in accounts}
    for transaction in transactions:
        acc_id = transaction.finance_account_id
        is_revenue = (transaction.operation_type or "").strip().lower() == "receita"
        for chunk in _finance_transaction_balance_amount_chunks(transaction):
            if is_revenue:
                account_balances[acc_id] = round(account_balances.get(acc_id, 0.0) + chunk, 2)
            else:
                account_balances[acc_id] = round(account_balances.get(acc_id, 0.0) - chunk, 2)
    return templates.TemplateResponse(
        "finance_accounts.html",
        _base_context(
            request,
            user,
            csrf_token,
            "finance_accounts",
            title="Contas",
            _repo=repo,
            active_farm=active_farm,
            finance_accounts=accounts,
            finance_account_total=len(accounts),
            finance_account_default=next((item for item in accounts if item.is_default), None),
            edit_finance_account=edit_account,
            finance_transactions=transactions_pagination["items"],
            finance_transaction_total=len(transactions),
            finance_transactions_pagination=transactions_pagination,
            finance_payables=payables_pagination["items"],
            finance_payables_total=len(payable_rows),
            finance_payables_pagination=payables_pagination,
            finance_payables_filter_start_date=payables_filter_start_date,
            finance_payables_filter_end_date=payables_filter_end_date,
            finance_payables_selected_range=selected_payables_range,
            finance_payables_period_ui_mode=_finance_payables_period_ui_mode(payables_status),
            finance_payables_search=payables_search,
            finance_payables_status=payables_status,
            finance_payables_filters_active=finance_payables_filters_active,
            finance_payables_clear_url=finance_payables_clear_url,
            finance_payables_redirect_to=_payables_redirect_url(request),
            edit_finance_transaction=edit_transaction,
            finance_bank_options=bank_options,
            finance_open_launch_modal=bool(launch or edit_account),
            finance_open_transaction_modal=bool(transaction_launch or edit_transaction),
            finance_transaction_expense_categories=FINANCE_TRANSACTION_EXPENSE_CATEGORIES,
            finance_transaction_revenue_categories=FINANCE_TRANSACTION_REVENUE_CATEGORIES,
            finance_transaction_payment_methods=FINANCE_TRANSACTION_PAYMENT_METHODS,
            finance_account_balances=account_balances,
            today=today,
        ),
    )


@router.post("/gestao-financeira/contas")
def create_finance_account_action(
    request: Request,
    csrf_token: str = Form(...),
    account_name: str = Form(...),
    initial_balance_date: str = Form(...),
    initial_balance: str = Form("0"),
    bank_code: str = Form(...),
    bank_name: str = Form(""),
    custom_bank_code: str | None = Form(None),
    custom_bank_name: str | None = Form(None),
    branch_number: str | None = Form(None),
    account_number: str | None = Form(None),
    is_default: bool = Form(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    if not active_farm:
        _flash(request, "error", "Selecione uma fazenda ativa antes de cadastrar uma conta.")
        return _redirect("/gestao-financeira/contas")
    try:
        payload = _finance_accounts_parse_form(
            account_name,
            initial_balance_date,
            initial_balance,
            bank_code,
            bank_name,
            custom_bank_code,
            custom_bank_name,
            branch_number,
            account_number,
            is_default,
            repo,
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_accounts_modal_query(request, launch=True))
    try:
        _validate_finance_account_uniqueness(
            repo,
            farm_id=active_farm.id,
            branch_number=payload["branch_number"],
            account_number=payload["account_number"],
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_accounts_modal_query(request, launch=True))

    if payload["is_default"]:
        _finance_accounts_set_default(repo, active_farm.id)

    custom_bank = payload.get("custom_bank")
    custom_bank_id = None
    if isinstance(custom_bank, dict):
        custom_bank_record = FinanceCustomBank(
            bank_code=custom_bank["bank_code"],
            bank_name=custom_bank["bank_name"],
            created_at=app_now(),
        )
        repo.create(custom_bank_record)
        custom_bank_id = custom_bank_record.id

    account = FinanceAccount(
        farm_id=active_farm.id,
        account_name=payload["account_name"],
        initial_balance_date=payload["initial_balance_date"],
        initial_balance=payload["initial_balance"],
        bank_code=payload["bank_code"],
        bank_name=payload["bank_name"],
        custom_bank_id=custom_bank_id if isinstance(custom_bank, dict) else (custom_bank.id if custom_bank else None),
        branch_number=payload["branch_number"],
        account_number=payload["account_number"],
        is_default=payload["is_default"],
        created_at=app_now(),
    )
    repo.db.add(account)
    repo.db.commit()
    _flash(request, "success", "Conta bancária cadastrada com sucesso.")
    return _redirect("/gestao-financeira/contas")


@router.post("/gestao-financeira/contas/{account_id}/editar")
def update_finance_account_action(
    request: Request,
    account_id: int,
    csrf_token: str = Form(...),
    account_name: str = Form(...),
    initial_balance_date: str = Form(...),
    initial_balance: str = Form("0"),
    bank_code: str = Form(...),
    bank_name: str = Form(""),
    custom_bank_code: str | None = Form(None),
    custom_bank_name: str | None = Form(None),
    branch_number: str | None = Form(None),
    account_number: str | None = Form(None),
    is_default: bool = Form(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    account = repo.get_finance_account(account_id)
    if not account:
        _flash(request, "error", "Conta bancária não encontrada.")
        return _redirect("/gestao-financeira/contas")
    if not active_farm or account.farm_id != active_farm.id:
        _flash(request, "error", "Esta conta não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    try:
        payload = _finance_accounts_parse_form(
            account_name,
            initial_balance_date,
            initial_balance,
            bank_code,
            bank_name,
            custom_bank_code,
            custom_bank_name,
            branch_number,
            account_number,
            is_default,
            repo,
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_accounts_modal_query(request, edit_id=account_id))
    try:
        _validate_finance_account_uniqueness(
            repo,
            farm_id=active_farm.id,
            branch_number=payload["branch_number"],
            account_number=payload["account_number"],
            ignore_id=account.id,
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_accounts_modal_query(request, edit_id=account_id))

    if payload["is_default"]:
        _finance_accounts_set_default(repo, active_farm.id, keep_id=account.id)

    previous_custom_bank_id = account.custom_bank_id
    custom_bank = payload.get("custom_bank")
    next_custom_bank_id = None
    if isinstance(custom_bank, dict):
        custom_bank_record = FinanceCustomBank(
            bank_code=custom_bank["bank_code"],
            bank_name=custom_bank["bank_name"],
            created_at=app_now(),
        )
        repo.create(custom_bank_record)
        next_custom_bank_id = custom_bank_record.id
    elif custom_bank:
        next_custom_bank_id = custom_bank.id

    repo.update(
        account,
        {
            "account_name": payload["account_name"],
            "initial_balance_date": payload["initial_balance_date"],
            "initial_balance": payload["initial_balance"],
            "bank_code": payload["bank_code"],
            "bank_name": payload["bank_name"],
            "custom_bank_id": next_custom_bank_id,
            "branch_number": payload["branch_number"],
            "account_number": payload["account_number"],
            "is_default": payload["is_default"],
        },
    )
    if previous_custom_bank_id and previous_custom_bank_id != next_custom_bank_id:
        _cleanup_custom_bank_if_unused(repo, previous_custom_bank_id)
    _flash(request, "success", "Conta bancária atualizada com sucesso.")
    return _redirect("/gestao-financeira/contas")


@router.post("/gestao-financeira/contas/{account_id}/excluir")
def delete_finance_account_action(
    request: Request,
    account_id: int,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    account = repo.get_finance_account(account_id)
    if not account:
        _flash(request, "error", "Conta bancária não encontrada.")
        return _redirect("/gestao-financeira/contas")
    if not active_farm or account.farm_id != active_farm.id:
        _flash(request, "error", "Esta conta não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    previous_custom_bank_id = account.custom_bank_id
    repo.delete(account)
    _cleanup_custom_bank_if_unused(repo, previous_custom_bank_id)
    _flash(request, "success", "Conta bancária excluída com sucesso.")
    return _redirect("/gestao-financeira/contas")


@router.post("/gestao-financeira/contas/{account_id}/definir-padrao")
def set_default_finance_account_action(
    request: Request,
    account_id: int,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    account = repo.get_finance_account(account_id)
    if not account:
        _flash(request, "error", "Conta bancária não encontrada.")
        return _redirect("/gestao-financeira/contas")
    if not active_farm or account.farm_id != active_farm.id:
        _flash(request, "error", "Esta conta não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    if account.is_default:
        return _redirect("/gestao-financeira/contas")
    _finance_accounts_set_default(repo, active_farm.id, keep_id=account.id)
    repo.update(account, {"is_default": True})
    _flash(request, "success", f"{account.account_name} definida como conta padrão.")
    return _redirect("/gestao-financeira/contas")


@router.post("/gestao-financeira/contas/lancamentos")
async def create_finance_transaction_action(
    request: Request,
    csrf_token: str = Form(...),
    operation_type: str = Form(...),
    launch_date: str = Form(...),
    amount: str = Form(...),
    finance_account_id: str | None = Form(None),
    payment_condition: str | None = Form("a_vista"),
    installment_count: str | None = Form(None),
    installment_frequency: str | None = Form(None),
    first_installment_date: str | None = Form(None),
    category: str = Form(...),
    product_service: str = Form(...),
    description: str | None = Form(None),
    counterparty_name: str | None = Form(None),
    document_number: str | None = Form(None),
    payment_method: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    if not active_farm:
        _flash(request, "error", "Selecione uma fazenda ativa antes de cadastrar um lançamento.")
        return _redirect("/gestao-financeira/contas")
    try:
        payload = _finance_transaction_payload(
            repo,
            active_farm=active_farm,
            operation_type=operation_type,
            launch_date=launch_date,
            amount=amount,
            finance_account_id=finance_account_id,
            payment_condition=payment_condition,
            installment_count=installment_count,
            installment_frequency=installment_frequency,
            first_installment_date=first_installment_date,
            category=category,
            product_service=product_service,
            description=description,
            counterparty_name=counterparty_name,
            document_number=document_number,
            payment_method=payment_method,
            notes=notes,
        )
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_transactions_modal_query(request, launch=True))

    transaction = repo.create(
        FinanceTransaction(
            farm_id=payload["farm_id"],
            finance_account_id=payload["finance_account_id"],
            operation_type=payload["operation_type"],
            launch_date=payload["launch_date"],
            amount=payload["amount"],
            category=payload["category"],
            product_service=payload["product_service"],
            description=payload["description"],
            counterparty_name=payload["counterparty_name"],
            document_number=payload["document_number"],
            payment_method=payload["payment_method"],
            payment_condition=payload["payment_condition"],
            installment_count=payload["installment_count"],
            installment_frequency=payload["installment_frequency"],
            first_installment_date=payload["first_installment_date"],
            notes=payload["notes"],
            created_at=app_now(),
        )
    )
    _replace_finance_transaction_installments(
        repo,
        transaction,
        _build_finance_transaction_installments(
            amount=payload["amount"],
            payment_condition=payload["payment_condition"],
            installment_count=payload["installment_count"],
            installment_frequency=payload["installment_frequency"],
            first_installment_date=payload["first_installment_date"],
        ),
    )
    try:
        saved_attachments = _save_finance_transaction_attachments(repo, transaction, attachment_payloads)
    except Exception:
        _flash(request, "error", "O lançamento foi salvo, mas não foi possível gravar os anexos agora.")
        return _redirect(_finance_transactions_modal_query(request, edit_id=transaction.id))
    if saved_attachments:
        _flash(request, "success", f"Lançamento salvo com sucesso. {saved_attachments} anexo(s) salvo(s).")
        return _redirect("/gestao-financeira/contas?finance_tab=transactions")
    _flash(request, "success", "Lançamento salvo com sucesso.")
    return _redirect("/gestao-financeira/contas?finance_tab=transactions")


@router.post("/gestao-financeira/contas/lancamentos/{transaction_id}/editar")
async def update_finance_transaction_action(
    request: Request,
    transaction_id: int,
    csrf_token: str = Form(...),
    operation_type: str = Form(...),
    launch_date: str = Form(...),
    amount: str = Form(...),
    finance_account_id: str | None = Form(None),
    payment_condition: str | None = Form("a_vista"),
    installment_count: str | None = Form(None),
    installment_frequency: str | None = Form(None),
    first_installment_date: str | None = Form(None),
    category: str = Form(...),
    product_service: str = Form(...),
    description: str | None = Form(None),
    counterparty_name: str | None = Form(None),
    document_number: str | None = Form(None),
    payment_method: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    transaction = repo.get_finance_transaction(transaction_id)
    if not transaction:
        _flash(request, "error", "Lançamento não encontrado.")
        return _redirect("/gestao-financeira/contas")
    if not active_farm or transaction.farm_id != active_farm.id:
        _flash(request, "error", "Este lançamento não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    prior_installment_snapshot = _snapshot_finance_installments_for_edit(transaction)
    try:
        payload = _finance_transaction_payload(
            repo,
            active_farm=active_farm,
            operation_type=operation_type,
            launch_date=launch_date,
            amount=amount,
            finance_account_id=finance_account_id,
            payment_condition=payment_condition,
            installment_count=installment_count,
            installment_frequency=installment_frequency,
            first_installment_date=first_installment_date,
            category=category,
            product_service=product_service,
            description=description,
            counterparty_name=counterparty_name,
            document_number=document_number,
            payment_method=payment_method,
            notes=notes,
        )
        _reject_finance_schedule_change_if_installment_paid(transaction, payload)
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_transactions_modal_query(request, edit_id=transaction_id))

    repo.update(transaction, payload)
    new_installment_rows = _build_finance_transaction_installments(
        amount=payload["amount"],
        payment_condition=payload["payment_condition"],
        installment_count=payload["installment_count"],
        installment_frequency=payload["installment_frequency"],
        first_installment_date=payload["first_installment_date"],
    )
    try:
        _validate_installment_edit_against_prior(new_installment_rows, prior_installment_snapshot)
        _merge_paid_installment_state_into_rows(new_installment_rows, prior_installment_snapshot)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(_finance_transactions_modal_query(request, edit_id=transaction_id))

    _replace_finance_transaction_installments(repo, transaction, new_installment_rows)
    try:
        saved_attachments = _save_finance_transaction_attachments(repo, transaction, attachment_payloads)
    except Exception:
        _flash(request, "error", "As alterações foram salvas, mas não foi possível incluir os novos anexos.")
        return _redirect(_finance_transactions_modal_query(request, edit_id=transaction_id))
    if saved_attachments:
        _flash(request, "success", f"Lançamento atualizado com sucesso. {saved_attachments} novo(s) anexo(s) adicionado(s).")
        return _redirect("/gestao-financeira/contas?finance_tab=transactions")
    _flash(request, "success", "Lançamento atualizado com sucesso.")
    return _redirect("/gestao-financeira/contas?finance_tab=transactions")


@router.post("/gestao-financeira/contas/lancamentos/{transaction_id}/excluir")
def delete_finance_transaction_action(
    request: Request,
    transaction_id: int,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    transaction = repo.get_finance_transaction(transaction_id)
    if not transaction:
        _flash(request, "error", "Lançamento não encontrado.")
        return _redirect("/gestao-financeira/contas")
    if not active_farm or transaction.farm_id != active_farm.id:
        _flash(request, "error", "Este lançamento não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas")
    repo.delete(transaction)
    _flash(request, "success", "Lançamento excluído com sucesso.")
    return _redirect("/gestao-financeira/contas?finance_tab=transactions")


@router.post("/gestao-financeira/contas/parcelas/{installment_id}/quitar")
def settle_finance_transaction_installment_action(
    request: Request,
    installment_id: int,
    csrf_token: str = Form(...),
    paid_at: str = Form(...),
    payment_notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    installment = repo.db.query(FinanceTransactionInstallment).options(joinedload(FinanceTransactionInstallment.transaction)).filter(FinanceTransactionInstallment.id == installment_id).first()
    if not installment or not installment.transaction:
        _flash(request, "error", "Parcela não encontrada.")
        return _redirect("/gestao-financeira/contas?finance_tab=payables")
    if not active_farm or installment.transaction.farm_id != active_farm.id:
        _flash(request, "error", "Esta parcela não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas?finance_tab=payables")
    try:
        paid_date = date.fromisoformat((paid_at or "").strip())
    except ValueError:
        _flash(request, "error", "Informe uma data válida para o pagamento.")
        return _redirect("/gestao-financeira/contas?finance_tab=payables")
    repo.update(
        installment,
        {
            "status": "pago",
            "paid_at": paid_date,
            "payment_notes": _clean_text(payment_notes),
        },
    )
    _flash(request, "success", "Parcela marcada como paga.")
    return _redirect("/gestao-financeira/contas?finance_tab=payables")


@router.post("/gestao-financeira/contas/parcelas/{installment_id}/cancelar-pagamento")
def revert_finance_transaction_installment_payment_action(
    request: Request,
    installment_id: int,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    installment = (
        repo.db.query(FinanceTransactionInstallment)
        .options(joinedload(FinanceTransactionInstallment.transaction))
        .filter(FinanceTransactionInstallment.id == installment_id)
        .first()
    )
    if not installment or not installment.transaction:
        _flash(request, "error", "Parcela não encontrada.")
        return _redirect("/gestao-financeira/contas?finance_tab=payables")
    if not active_farm or installment.transaction.farm_id != active_farm.id:
        _flash(request, "error", "Esta parcela não pertence à fazenda ativa.")
        return _redirect("/gestao-financeira/contas?finance_tab=payables")
    repo.update(
        installment,
        {
            "status": "pendente",
            "paid_at": None,
            "payment_notes": None,
        },
    )
    _flash(request, "success", "Pagamento cancelado. A parcela voltou para pendente.")
    safe_redirect = (redirect_to or "").strip()
    if safe_redirect.startswith("/"):
        return _redirect(safe_redirect)
    return _redirect("/gestao-financeira/contas?finance_tab=payables")


@router.get("/gestao-financeira/contas/parcelas/{installment_id}/observacao-pagamento")
def read_installment_payment_notes(
    request: Request,
    installment_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    installment = (
        repo.db.query(FinanceTransactionInstallment)
        .options(joinedload(FinanceTransactionInstallment.transaction))
        .filter(FinanceTransactionInstallment.id == installment_id)
        .first()
    )
    if not installment or not installment.transaction:
        return JSONResponse({"ok": False, "error": "Parcela não encontrada."}, status_code=404)
    if not active_farm or installment.transaction.farm_id != active_farm.id:
        return JSONResponse({"ok": False, "error": "Acesso negado."}, status_code=403)
    if (installment.status or "").strip().lower() != "pago":
        return JSONResponse({"ok": False, "error": "Somente parcelas pagas possuem observações de pagamento."}, status_code=400)
    return JSONResponse({"ok": True, "payment_notes": installment.payment_notes or ""})


@router.post("/gestao-financeira/contas/parcelas/{installment_id}/observacao-pagamento")
def update_installment_payment_notes_action(
    request: Request,
    installment_id: int,
    csrf_token: str = Form(...),
    payment_notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm = scope.get("active_farm")
    installment = (
        repo.db.query(FinanceTransactionInstallment)
        .options(joinedload(FinanceTransactionInstallment.transaction))
        .filter(FinanceTransactionInstallment.id == installment_id)
        .first()
    )
    if not installment or not installment.transaction:
        return JSONResponse({"ok": False, "error": "Parcela não encontrada."}, status_code=404)
    if not active_farm or installment.transaction.farm_id != active_farm.id:
        return JSONResponse({"ok": False, "error": "Acesso negado."}, status_code=403)
    if (installment.status or "").strip().lower() != "pago":
        return JSONResponse({"ok": False, "error": "Observações só podem ser editadas em parcelas já pagas."}, status_code=400)
    cleaned = _clean_text(payment_notes)
    repo.update(installment, {"payment_notes": cleaned})
    return JSONResponse({"ok": True, "payment_notes": cleaned or ""})


@router.get("/gestao-financeira/contas/lancamentos/anexos/{attachment_id}")
def open_finance_transaction_attachment(
    request: Request,
    attachment_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    attachment = repo.get_finance_transaction_attachment(attachment_id)
    if not attachment or not attachment.transaction:
        _flash(request, "error", "Anexo não encontrado.")
        return _redirect("/gestao-financeira/contas")
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(attachment.transaction.farm_id, scope):
        _flash(request, "error", "Este anexo não pertence ao contexto ativo.")
        return _redirect("/gestao-financeira/contas")
    return _attachment_response(attachment.filename, attachment.content_type, attachment.file_data)


@router.post("/gestao-financeira/contas/lancamentos/{transaction_id}/anexos/{attachment_id}/excluir")
def delete_finance_transaction_attachment_action(
    request: Request,
    transaction_id: int,
    attachment_id: int,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    transaction = repo.get_finance_transaction(transaction_id)
    if not transaction:
        _flash(request, "error", "Lançamento não encontrado.")
        return _redirect("/gestao-financeira/contas")
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(transaction.farm_id, scope):
        _flash(request, "error", "Este lançamento não pertence ao contexto ativo.")
        return _redirect(_finance_transactions_modal_query(request))
    attachment = repo.get_finance_transaction_attachment(attachment_id)
    if not attachment or attachment.finance_transaction_id != transaction.id:
        _flash(request, "error", "Anexo não encontrado.")
        return _redirect(_finance_transactions_modal_query(request, edit_id=transaction_id))
    repo.delete(attachment)
    _flash(request, "success", "Anexo excluído com sucesso.")
    return _redirect(_finance_transactions_modal_query(request, edit_id=transaction_id))


@router.get("/talhoes", include_in_schema=False)
@router.get("/setores")
def plots_page(
    request: Request,
    background_tasks: BackgroundTasks,
    q: str | None = None,
    sort: str = "name",
    edit_id: int | None = None,
    plot_map: int | None = None,
    selected_farm_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids = _int_list(request.query_params.getlist("farm_id"))
    variety_ids = _int_list(request.query_params.getlist("variety_id"))
    if not farm_ids and scope["active_farm_id"]:
        farm_ids = [scope["active_farm_id"]]
    if not variety_ids and scope["active_season"] and scope["active_season"].variety_id:
        variety_ids = [scope["active_season"].variety_id]
    edit_plot = repo.get_plot(edit_id) if edit_id else None
    open_plot_map_modal = bool(plot_map and edit_plot)
    farms, varieties = repo.list_plot_filter_options(farm_ids or None, variety_ids or None)
    # Para o cadastro/edição do setor, a lista de variedades não deve ser limitada pelo filtro da página.
    # Caso contrário, o select "Variedade" fica preso à variedade da safra ativa (ex.: apenas 1 opção).
    form_varieties = repo.list_varieties()
    plots_list = repo.list_plots(search=q, farm_ids=farm_ids, variety_ids=variety_ids, sort=sort)
    plot_preview_ready: dict[int, bool] = {}
    plot_preview_fingerprint: dict[int, str] = {}
    for plot in plots_list:
        if not plot.boundary_geojson:
            continue
        plot_preview_fingerprint[plot.id] = hashlib.sha256(plot.boundary_geojson.encode("utf-8")).hexdigest()[:14]
        full_path = plot_preview_fs_path(plot.id)
        thumb_path = plot_preview_thumb_fs_path(plot.id)
        full_ok = full_path.is_file() and full_path.stat().st_size > 0
        thumb_ok = thumb_path.is_file() and thumb_path.stat().st_size > 0
        plot_preview_ready[plot.id] = full_ok and thumb_ok
        if not full_ok:
            farm_gj = _farm_boundary_for_plot_preview(plot)
            background_tasks.add_task(generate_plot_preview_image, plot.id, plot.boundary_geojson, farm_gj)
        elif not thumb_ok:
            background_tasks.add_task(ensure_plot_preview_thumb, plot.id)
    farm_preview_ready: dict[int, bool] = {}
    farm_preview_fingerprint: dict[int, str] = {}
    plots_farm_preview_urls: dict[str, str] = {}
    for farm in farms:
        if not farm.boundary_geojson:
            continue
        farm_preview_fingerprint[farm.id] = hashlib.sha256(farm.boundary_geojson.encode("utf-8")).hexdigest()[:14]
        full_path = farm_preview_fs_path(farm.id)
        thumb_path = farm_preview_thumb_fs_path(farm.id)
        full_ok = full_path.is_file() and full_path.stat().st_size > 0
        thumb_ok = thumb_path.is_file() and thumb_path.stat().st_size > 0
        farm_preview_ready[farm.id] = full_ok and thumb_ok
        if not full_ok:
            background_tasks.add_task(generate_farm_preview_image, farm.id, farm.boundary_geojson)
        elif not thumb_ok:
            background_tasks.add_task(ensure_farm_preview_thumb, farm.id)
        if farm_preview_ready[farm.id]:
            plots_farm_preview_urls[str(farm.id)] = (
                f"/static/generated/farm_previews/{farm.id}.png?v={farm_preview_fingerprint[farm.id]}"
            )
    plot_farm_boundary_by_id: dict[str, object] = {}
    for farm in farms:
        if not farm.boundary_geojson:
            continue
        try:
            plot_farm_boundary_by_id[str(farm.id)] = json.loads(farm.boundary_geojson)
        except json.JSONDecodeError:
            pass
    plot_farm_boundaries_json = json.dumps(plot_farm_boundary_by_id)
    farm_ids_for_plot_context = [f.id for f in farms]
    plots_context_by_farm: dict[str, list[dict[str, object]]] = {}
    if farm_ids_for_plot_context:
        for plot in repo.list_plots_with_boundary_geojson(farm_ids=farm_ids_for_plot_context):
            if not plot.farm_id:
                continue
            try:
                gj_obj = json.loads(plot.boundary_geojson)
            except json.JSONDecodeError:
                continue
            plots_context_by_farm.setdefault(str(plot.farm_id), []).append({"id": plot.id, "geometry": gj_obj})
    plots_context_boundaries_json = json.dumps(plots_context_by_farm)
    edit_plot_geometry_json = "null"
    if edit_plot and edit_plot.boundary_geojson:
        try:
            edit_plot_geometry_json = json.dumps(json.loads(edit_plot.boundary_geojson))
        except json.JSONDecodeError:
            edit_plot_geometry_json = "null"
    google_maps_web_key = (get_settings().google_maps_api_key or "").strip()
    _plot_scope_season = scope.get("active_season")
    _plots_active_season_variety_id = (
        _plot_scope_season.variety_id if _plot_scope_season and _plot_scope_season.variety_id else None
    )
    plots_filters_active = _plots_page_filters_active(
        request,
        active_farm_id=scope["active_farm_id"],
        active_season_variety_id=_plots_active_season_variety_id,
        q=q,
        sort=sort,
    )
    plots_scope_active_farm_id = scope.get("active_farm_id")
    plots_farm_scope_locked = plots_scope_active_farm_id is not None
    plots_locked_farm_name = None
    if plots_farm_scope_locked:
        # A lista `farms` pode vir filtrada por safra/variedade e não incluir a fazenda do contexto.
        # Para evitar o fallback "Fazenda #id", busca sempre o nome real direto do repositório.
        _locked = repo.get_farm(plots_scope_active_farm_id)
        plots_locked_farm_name = _locked.name if _locked else f"Fazenda #{plots_scope_active_farm_id}"
    plots_field_q_filtered = bool((q or "").strip())
    plots_field_sort_filtered = sort != "name"
    plots_field_variety_filtered = _plots_variety_query_is_user_filter(request, _plots_active_season_variety_id)
    return templates.TemplateResponse(
        "plots.html",
        _base_context(
            request,
            user,
            csrf_token,
            "plots",
            plots=plots_list,
            farms=farms,
            varieties=varieties,
            form_varieties=form_varieties,
            filters={"q": q or "", "farm_ids": farm_ids, "variety_ids": variety_ids, "sort": sort},
            plots_filters_active=plots_filters_active,
            plots_farm_scope_locked=plots_farm_scope_locked,
            plots_scope_active_farm_id=plots_scope_active_farm_id,
            plots_locked_farm_name=plots_locked_farm_name,
            plots_field_q_filtered=plots_field_q_filtered,
            plots_field_variety_filtered=plots_field_variety_filtered,
            plots_field_sort_filtered=plots_field_sort_filtered,
            edit_plot=edit_plot,
            open_plot_map_modal=open_plot_map_modal,
            selected_farm_id=selected_farm_id or scope["active_farm_id"],
            plot_farm_boundaries_json=plot_farm_boundaries_json,
            plots_context_boundaries_json=plots_context_boundaries_json,
            edit_plot_geometry_json=edit_plot_geometry_json,
            google_maps_web_key=google_maps_web_key,
            plot_preview_ready=plot_preview_ready,
            plot_preview_fingerprint=plot_preview_fingerprint,
            farm_preview_ready=farm_preview_ready,
            farm_preview_fingerprint=farm_preview_fingerprint,
            plots_farm_preview_urls=plots_farm_preview_urls,
            filter_links=[
                {"farm_id": plot.farm_id, "variety_id": plot.variety_id}
                for plot in repo.list_plots(farm_ids=farm_ids or None, variety_ids=variety_ids or None)
                if plot.farm_id or plot.variety_id
            ],
            _repo=repo,
        ),
    )


@router.post("/talhoes", include_in_schema=False)
@router.post("/setores")
def create_plot_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    area_hectares: float = Form(...),
    farm_id: str | None = Form(None),
    planting_date: str | None = Form(None),
    plant_count: int = Form(...),
    spacing_row_meters: str | None = Form(None),
    spacing_plant_centimeters: str | None = Form(None),
    estimated_yield_sacks: str | None = Form(None),
    variety_id: str | None = Form(None),
    centroid_lat: str | None = Form(None),
    centroid_lng: str | None = Form(None),
    boundary_geojson: str | None = Form(None),
    boundary_geojson_file: UploadFile | None = File(None),
    irrigation_type: str | None = Form("none"),
    irrigation_line_count: str | None = Form(None),
    irrigation_line_length_meters: str | None = Form(None),
    drip_spacing_centimeters: str | None = Form(None),
    drip_liters_per_hour: str | None = Form(None),
    sprinkler_count: str | None = Form(None),
    sprinkler_liters_per_hour: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id)
    if not selected_farm_id:
        _flash(request, "error", "Selecione uma fazenda antes de salvar o setor.")
        return _redirect("/setores")
    farm = repo.get_farm(selected_farm_id)
    if not farm:
        _flash(request, "error", "A fazenda selecionada nao foi encontrada.")
        return _redirect("/setores")
    geometry, geometry_ok, upload_payload, upload_filename = _resolve_geojson(boundary_geojson_file, boundary_geojson)
    if boundary_geojson_file and boundary_geojson_file.filename and not geometry_ok:
        _flash(request, "error", "O arquivo GeoJSON do setor nao e valido.")
        return _redirect_with_query("/setores", selected_farm_id=selected_farm_id)
    new_plot = create_plot(
        repo,
        _build_plot_payload(
            farm=farm,
            name=name,
            area_hectares=area_hectares,
            planting_date=planting_date,
            plant_count=plant_count,
            spacing_row_meters=spacing_row_meters,
            spacing_plant_centimeters=spacing_plant_centimeters,
            estimated_yield_sacks=estimated_yield_sacks,
            variety_id=variety_id,
            centroid_lat=centroid_lat,
            centroid_lng=centroid_lng,
            boundary_geojson=geometry,
            notes=notes,
            irrigation_type=irrigation_type,
            irrigation_line_count=irrigation_line_count,
            irrigation_line_length_meters=irrigation_line_length_meters,
            drip_spacing_centimeters=drip_spacing_centimeters,
            drip_liters_per_hour=drip_liters_per_hour,
            sprinkler_count=sprinkler_count,
            sprinkler_liters_per_hour=sprinkler_liters_per_hour,
        ),
    )
    if geometry:
        try:
            farm_gj = farm.boundary_geojson if farm.boundary_geojson else None
            generate_plot_preview_image(new_plot.id, geometry, farm_gj)
        except Exception:
            logging.getLogger(__name__).exception("Falha ao gerar imagem de satelite do setor %s", new_plot.id)
    _maybe_record_plot_boundary_attachments(
        repo,
        new_plot,
        geometry,
        upload_payload,
        upload_filename,
        None,
        is_new_plot=True,
    )
    _flash(request, "success", "Setor salvo com sucesso.")
    return _redirect("/setores")


@router.post("/talhoes/{plot_id}/editar", include_in_schema=False)
@router.post("/setores/{plot_id}/editar")
def update_plot_action(
    plot_id: int,
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    area_hectares: float = Form(...),
    farm_id: int = Form(...),
    planting_date: str | None = Form(None),
    plant_count: int = Form(...),
    spacing_row_meters: str | None = Form(None),
    spacing_plant_centimeters: str | None = Form(None),
    estimated_yield_sacks: str | None = Form(None),
    variety_id: str | None = Form(None),
    centroid_lat: str | None = Form(None),
    centroid_lng: str | None = Form(None),
    boundary_geojson: str | None = Form(None),
    boundary_geojson_file: UploadFile | None = File(None),
    irrigation_type: str | None = Form("none"),
    irrigation_line_count: str | None = Form(None),
    irrigation_line_length_meters: str | None = Form(None),
    drip_spacing_centimeters: str | None = Form(None),
    drip_liters_per_hour: str | None = Form(None),
    sprinkler_count: str | None = Form(None),
    sprinkler_liters_per_hour: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot = repo.get_plot(plot_id)
    if not plot:
        _flash(request, "error", "Setor nao encontrado.")
        return _redirect("/setores")
    old_geometry = plot.boundary_geojson
    farm = repo.get_farm(farm_id)
    if not farm:
        _flash(request, "error", "A fazenda selecionada nao foi encontrada.")
        return _redirect("/setores")
    geometry, geometry_ok, upload_payload, upload_filename = _resolve_geojson(
        boundary_geojson_file, boundary_geojson, plot.boundary_geojson
    )
    if boundary_geojson_file and boundary_geojson_file.filename and not geometry_ok:
        _flash(request, "error", "O arquivo GeoJSON do setor nao e valido.")
        return _redirect_with_query("/setores", edit_id=plot_id)
    update_plot(
        repo,
        plot,
        _build_plot_payload(
            farm=farm,
            name=name,
            area_hectares=area_hectares,
            planting_date=planting_date,
            plant_count=plant_count,
            spacing_row_meters=spacing_row_meters,
            spacing_plant_centimeters=spacing_plant_centimeters,
            estimated_yield_sacks=estimated_yield_sacks,
            variety_id=variety_id,
            centroid_lat=centroid_lat,
            centroid_lng=centroid_lng,
            boundary_geojson=geometry,
            notes=notes,
            irrigation_type=irrigation_type,
            irrigation_line_count=irrigation_line_count,
            irrigation_line_length_meters=irrigation_line_length_meters,
            drip_spacing_centimeters=drip_spacing_centimeters,
            drip_liters_per_hour=drip_liters_per_hour,
            sprinkler_count=sprinkler_count,
            sprinkler_liters_per_hour=sprinkler_liters_per_hour,
        ),
    )
    try:
        farm_gj = farm.boundary_geojson if farm.boundary_geojson else None
        generate_plot_preview_image(plot_id, geometry or "", farm_gj)
    except Exception:
        logging.getLogger(__name__).exception("Falha ao atualizar imagem de satelite do setor %s", plot_id)
    finally:
        remove_plot_preview_draft(plot_id)
    _maybe_record_plot_boundary_attachments(
        repo,
        plot,
        geometry,
        upload_payload,
        upload_filename,
        old_geometry,
        is_new_plot=False,
    )
    _flash(request, "success", "Setor atualizado com sucesso.")
    return _redirect("/setores")


@router.get("/talhoes/anexos/{attachment_id}", include_in_schema=False)
@router.get("/setores/anexos/{attachment_id}")
def open_plot_attachment(
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    attachment = repo.get_plot_attachment(attachment_id)
    if not attachment or not attachment.plot:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect("/setores")
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(attachment.plot.farm_id, scope):
        _flash(request, "error", "Este anexo nao pertence ao contexto ativo.")
        return _redirect("/setores")
    return _attachment_response(attachment.filename, attachment.content_type, attachment.file_data)


@router.post("/talhoes/{plot_id}/anexos/{attachment_id}/excluir", include_in_schema=False)
@router.post("/setores/{plot_id}/anexos/{attachment_id}/excluir")
def delete_plot_attachment_action(
    plot_id: int,
    attachment_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot = repo.get_plot(plot_id)
    if not plot:
        _flash(request, "error", "Setor nao encontrado.")
        return _redirect("/setores")
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(plot.farm_id, scope):
        _flash(request, "error", "Este setor nao pertence ao contexto ativo.")
        return _redirect_with_query("/setores", edit_id=plot_id)
    attachment = repo.get_plot_attachment(attachment_id)
    if not attachment or attachment.plot_id != plot.id:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect_with_query("/setores", edit_id=plot_id)
    repo.delete(attachment)
    if repo.get_plot_attachment(attachment_id):
        _flash(request, "error", "Nao foi possivel remover o anexo agora. Tente novamente.")
        return _redirect_with_query("/setores", edit_id=plot_id)
    _flash(request, "success", "Anexo removido com sucesso.")
    return _redirect_with_query("/setores", edit_id=plot_id)


@router.post("/talhoes/{plot_id}/preview-rascunho", include_in_schema=False)
@router.post("/setores/{plot_id}/preview-rascunho")
def save_plot_preview_draft(
    plot_id: int,
    request: Request,
    csrf_token: str = Form(...),
    boundary_geojson: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot = repo.get_plot(plot_id)
    if not plot:
        return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
    normalized = normalize_geojson(boundary_geojson)
    if not normalized:
        return JSONResponse({"ok": False, "error": "invalid_geojson"}, status_code=400)
    ok = generate_plot_preview_draft(plot_id, normalized, _farm_boundary_for_plot_preview(plot))
    if not ok:
        return JSONResponse({"ok": False, "error": "generate_failed"}, status_code=400)
    revision = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:14]
    url = f"/static/generated/plot_previews/{plot_id}_draft.png"
    return JSONResponse({"ok": True, "url": url, "revision": revision})


@router.post("/talhoes/preview-geometria", include_in_schema=False)
@router.post("/setores/preview-geometria")
def preview_plot_geometry_session(
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str = Form(...),
    boundary_geojson: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    """Prévia de satélite para geometria ainda sem setor persistido (cadastro novo)."""
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    fid = _int_or_none(farm_id)
    if not fid:
        return JSONResponse({"ok": False, "error": "farm_required"}, status_code=400)
    farm = repo.get_farm(fid)
    if not farm:
        return JSONResponse({"ok": False, "error": "farm_not_found"}, status_code=404)
    normalized = normalize_geojson(boundary_geojson)
    if not normalized:
        return JSONResponse({"ok": False, "error": "invalid_geojson"}, status_code=400)
    farm_gj = farm.boundary_geojson if farm.boundary_geojson and str(farm.boundary_geojson).strip() else None
    url, revision = generate_plot_geometry_session_preview(normalized, farm_gj)
    if not url or not revision:
        return JSONResponse({"ok": False, "error": "generate_failed"}, status_code=400)
    return JSONResponse({"ok": True, "url": url, "revision": revision})


@router.post("/talhoes/{plot_id}/preview-rascunho/descartar", include_in_schema=False)
@router.post("/setores/{plot_id}/preview-rascunho/descartar")
def discard_plot_preview_draft(
    plot_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    if not repo.get_plot(plot_id):
        return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
    remove_plot_preview_draft(plot_id)
    return JSONResponse({"ok": True})


@router.post("/talhoes/{plot_id}/excluir", include_in_schema=False)
@router.post("/setores/{plot_id}/excluir")
def delete_plot_action(
    plot_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot = repo.get_plot(plot_id)
    if not plot:
        _flash(request, "error", "Setor nao encontrado.")
        return _redirect("/setores")
    remove_plot_preview_image(plot_id)
    repo.delete(plot)
    _flash(request, "success", "Setor excluido com sucesso.")
    return _redirect("/setores")


@router.get("/fazendas")
def farms_page(
    request: Request,
    background_tasks: BackgroundTasks,
    edit_id: int | None = None,
    view: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    edit_farm = repo.get_farm(edit_id) if edit_id else None
    farm_form_view_only = bool(edit_farm and view == 1)
    farms = repo.list_farms()
    farm_preview_ready: dict[int, bool] = {}
    farm_preview_fingerprint: dict[int, str] = {}
    for farm in farms:
        if not farm.boundary_geojson:
            continue
        farm_preview_fingerprint[farm.id] = hashlib.sha256(farm.boundary_geojson.encode("utf-8")).hexdigest()[:14]
        full_path = farm_preview_fs_path(farm.id)
        thumb_path = farm_preview_thumb_fs_path(farm.id)
        full_ok = full_path.is_file() and full_path.stat().st_size > 0
        thumb_ok = thumb_path.is_file() and thumb_path.stat().st_size > 0
        farm_preview_ready[farm.id] = full_ok and thumb_ok
        if not full_ok:
            background_tasks.add_task(generate_farm_preview_image, farm.id, farm.boundary_geojson)
        elif not thumb_ok:
            background_tasks.add_task(ensure_farm_preview_thumb, farm.id)
    farm_boundary_by_id: dict[str, object] = {}
    for farm in farms:
        if not farm.boundary_geojson:
            continue
        try:
            farm_boundary_by_id[str(farm.id)] = json.loads(farm.boundary_geojson)
        except json.JSONDecodeError:
            pass
    farm_boundary_geometries_json = json.dumps(farm_boundary_by_id)
    edit_farm_geometry_json = "null"
    if edit_farm and edit_farm.boundary_geojson:
        try:
            edit_farm_geometry_json = json.dumps(json.loads(edit_farm.boundary_geojson))
        except json.JSONDecodeError:
            edit_farm_geometry_json = "null"
    google_maps_web_key = (get_settings().google_maps_api_key or "").strip()
    return templates.TemplateResponse(
        "farms.html",
        _base_context(
            request,
            user,
            csrf_token,
            "farms",
            farms=farms,
            edit_farm=edit_farm,
            farm_form_view_only=farm_form_view_only,
            farm_preview_ready=farm_preview_ready,
            farm_preview_fingerprint=farm_preview_fingerprint,
            farm_boundary_geometries_json=farm_boundary_geometries_json,
            edit_farm_geometry_json=edit_farm_geometry_json,
            google_maps_web_key=google_maps_web_key,
            _repo=repo,
        ),
    )


@router.get("/fazendas/geocodificar")
def geocode_farm_location(
    q: str = "",
    user: User = Depends(get_current_user_web),
):
    """Geocodifica texto de localização (Nominatim/OSM) para centralizar o mapa do cadastro."""
    del user
    query = (q or "").strip()
    if len(query) < 2:
        return JSONResponse({"ok": False, "code": "empty"}, status_code=400)
    q_lower = query.lower()
    search_q = query
    if "brasil" not in q_lower and "brazil" not in q_lower:
        search_q = f"{query}, Brasil"
    params = urlencode(
        {
            "format": "json",
            "limit": "1",
            "q": search_q,
            "countrycodes": "br",
        }
    )
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "FazendaBelaVista/1.0 (+https://github.com/renan087/fazenda_bela_vista)",
            "Accept-Language": "pt-BR,pt;q=0.9",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        logging.getLogger(__name__).warning("Nominatim HTTP %s", exc.code)
        return JSONResponse({"ok": False, "code": "upstream"}, status_code=502)
    except Exception:
        logging.getLogger(__name__).exception("geocode nominatim")
        return JSONResponse({"ok": False, "code": "upstream"}, status_code=502)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return JSONResponse({"ok": False, "code": "upstream"}, status_code=502)
    if not data:
        return JSONResponse({"ok": False, "code": "not_found"})
    row = data[0]
    try:
        lat = float(row["lat"])
        lng = float(row["lon"])
    except (KeyError, TypeError, ValueError):
        return JSONResponse({"ok": False, "code": "not_found"})
    return JSONResponse({"ok": True, "lat": lat, "lng": lng})


@router.post("/fazendas")
def create_farm_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    location: str = Form(...),
    total_area: float = Form(...),
    boundary_geojson: str | None = Form(None),
    boundary_geojson_file: UploadFile | None = File(None),
    notes: str | None = Form(None),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    geometry, geometry_ok, _, _ = _resolve_geojson(boundary_geojson_file, boundary_geojson)
    if boundary_geojson_file and boundary_geojson_file.filename and not geometry_ok:
        _flash(request, "error", "O arquivo GeoJSON da fazenda nao e valido.")
        return _redirect("/setores") if redirect_to == "/setores" else _redirect("/fazendas")
    farm = create_farm(
        _repository(db),
        {
            "name": name,
            "location": location,
            "total_area": total_area,
            "boundary_geojson": geometry,
            "notes": notes,
        },
    )
    if geometry:
        try:
            generate_farm_preview_image(farm.id, geometry)
        except Exception:
            logging.getLogger(__name__).exception("Falha ao gerar imagem de satelite da fazenda %s", farm.id)
    _flash(request, "success", "Fazenda cadastrada com sucesso.")
    if redirect_to == "/setores":
        return _redirect_with_query("/setores", selected_farm_id=farm.id)
    return _redirect("/fazendas")


@router.post("/fazendas/{farm_id}/editar")
def update_farm_action(
    farm_id: int,
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    location: str = Form(...),
    total_area: float = Form(...),
    boundary_geojson: str | None = Form(None),
    boundary_geojson_file: UploadFile | None = File(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    farm = repo.get_farm(farm_id)
    if not farm:
        _flash(request, "error", "Fazenda nao encontrada.")
        return _redirect("/fazendas")
    geometry, geometry_ok, _, _ = _resolve_geojson(boundary_geojson_file, boundary_geojson, farm.boundary_geojson)
    if boundary_geojson_file and boundary_geojson_file.filename and not geometry_ok:
        _flash(request, "error", "O arquivo GeoJSON da fazenda nao e valido.")
        return _redirect_with_query("/fazendas", edit_id=farm_id)
    update_farm(
        repo,
        farm,
        {"name": name, "location": location, "total_area": total_area, "boundary_geojson": geometry, "notes": notes},
    )
    if geometry:
        try:
            generate_farm_preview_image(farm.id, geometry)
        except Exception:
            logging.getLogger(__name__).exception("Falha ao gerar imagem de satelite da fazenda %s", farm.id)
    else:
        remove_farm_preview_image(farm.id)
    _flash(request, "success", "Fazenda atualizada com sucesso.")
    return _redirect("/fazendas")


@router.post("/fazendas/{farm_id}/geometria")
def update_farm_geometry_only(
    farm_id: int,
    request: Request,
    csrf_token: str = Form(...),
    boundary_geojson: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    farm = repo.get_farm(farm_id)
    if not farm:
        _flash(request, "error", "Fazenda nao encontrada.")
        return _redirect("/fazendas")
    normalized = normalize_geojson(boundary_geojson)
    if not normalized:
        _flash(request, "error", "A geometria enviada nao e valida.")
        return _redirect("/fazendas")
    repo.update(farm, {"boundary_geojson": normalized})
    try:
        generate_farm_preview_image(farm.id, normalized)
    except Exception:
        logging.getLogger(__name__).exception("Falha ao gerar imagem de satelite da fazenda %s", farm.id)
    _flash(request, "success", "Geometria da fazenda atualizada.")
    return _redirect("/fazendas")


@router.post("/fazendas/{farm_id}/excluir")
def delete_farm_action(
    farm_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    farm = repo.get_farm(farm_id)
    if not farm:
        _flash(request, "error", "Fazenda nao encontrada.")
        return _redirect("/fazendas")
    if farm.plots:
        _flash(request, "error", "Nao e possivel excluir a fazenda enquanto houver setores vinculados.")
        return _redirect("/fazendas")
    remove_farm_preview_image(farm_id)
    repo.delete(farm)
    _flash(request, "success", "Fazenda excluida com sucesso.")
    return _redirect("/fazendas")


@router.get("/usuarios")
def users_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    repo = _repository(db)
    pending_super_admin_two_factor_disable = None
    pending_super_admin_two_factor_disable_confirmed = False
    pending_payload = _get_super_admin_2fa_disable_pending(request)
    if pending_payload and int(pending_payload.get("actor_user_id") or 0) == int(user.id):
        pending_target_user_id = int(pending_payload.get("target_user_id") or 0)
        pending_confirmed = bool(pending_payload.get("confirmed"))
        if not pending_confirmed and not get_active_login_code(db, user.id):
            _clear_super_admin_2fa_disable_pending(request)
        elif edit_id and pending_target_user_id == edit_id:
            pending_super_admin_two_factor_disable = pending_payload
            pending_super_admin_two_factor_disable_confirmed = pending_confirmed
    return templates.TemplateResponse(
        "users.html",
        _base_context(
            request,
            user,
            csrf_token,
            "users",
            title="Administracao de Usuarios",
            users=repo.list_users(),
            edit_user=repo.get_user(edit_id) if edit_id else None,
            format_app_datetime=format_app_datetime,
            super_admin_email=(settings.super_admin_email or settings.admin_email or "").strip().lower(),
            pending_super_admin_two_factor_disable=pending_super_admin_two_factor_disable,
            pending_super_admin_two_factor_disable_confirmed=pending_super_admin_two_factor_disable_confirmed,
            _repo=repo,
        ),
    )


@router.get("/usuarios/{user_id}/avatar")
def user_avatar_view(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    repo = _repository(db)
    target_user = repo.get_user(user_id)
    if not target_user or not target_user.avatar_data:
        return Response(status_code=404)
    return Response(
        content=target_user.avatar_data,
        media_type=target_user.avatar_content_type or "image/jpeg",
    )


@router.post("/usuarios")
def create_user_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    is_active: str | None = Form(None),
    is_admin: str | None = Form(None),
    is_two_factor_enabled: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    existing = db.query(User).filter(User.email == email.strip().lower()).first()
    if existing:
        _flash(request, "error", "Ja existe um usuario com este email.")
        return _redirect("/usuarios")
    create_user(
        repo,
        {
            "name": name,
            "email": email,
            "password": password,
            "is_active": _bool_from_form(is_active),
            "is_admin": _bool_from_form(is_admin),
            "is_two_factor_enabled": _bool_from_form(is_two_factor_enabled),
        },
    )
    _flash(request, "success", "Usuario criado com sucesso.")
    return _redirect("/usuarios")


@router.post("/usuarios/{user_id}/editar")
def update_user_action(
    user_id: int,
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    password: str | None = Form(None),
    is_active: str | None = Form(None),
    is_admin: str | None = Form(None),
    is_two_factor_enabled: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    try:
        validate_csrf(request, csrf_token)
        normalized_name = (name or "").strip()
        normalized_email = (email or "").strip().lower()
        if not normalized_name or not normalized_email:
            _flash(request, "error", "Nome e email sao obrigatorios para atualizar o usuario.")
            return _redirect_with_query("/usuarios", edit_id=user_id)

        repo = _repository(db)
        target_user = repo.get_user(user_id)
        if not target_user:
            _flash(request, "error", "Usuario nao encontrado.")
            return _redirect("/usuarios")

        existing = db.query(User).filter(User.email == normalized_email, User.id != user_id).first()
        if existing:
            _flash(request, "error", "Ja existe outro usuario com este email.")
            return _redirect_with_query("/usuarios", edit_id=user_id)

        requested_is_admin = _bool_from_form(is_admin)
        requested_is_two_factor_enabled = _bool_from_form(is_two_factor_enabled)
        # Só exige fluxo de código ao desligar 2FA que ainda está ativo no banco (evita bloquear edição quando já está desligado).
        pending_super_admin_two_factor_disable = (
            is_super_admin_email(normalized_email)
            and not requested_is_two_factor_enabled
            and bool(target_user.is_two_factor_enabled)
        )
        confirmed_super_admin_two_factor_disable = _super_admin_2fa_disable_pending_confirmed(
            request,
            target_user_id=user_id,
            actor_user_id=user.id,
        )

        updated_user = update_user(
            repo,
            target_user,
            {
                "name": normalized_name,
                "email": normalized_email,
                "password": password,
                "is_active": _bool_from_form(is_active),
                "is_admin": requested_is_admin,
                "is_two_factor_enabled": requested_is_two_factor_enabled,
                "allow_super_admin_two_factor_disable": pending_super_admin_two_factor_disable and confirmed_super_admin_two_factor_disable,
            },
        )

        if pending_super_admin_two_factor_disable and confirmed_super_admin_two_factor_disable:
            _discard_super_admin_2fa_disable_pending(request, db, user.id)
        elif _has_super_admin_2fa_disable_pending(request, target_user_id=updated_user.id, actor_user_id=user.id):
            _discard_super_admin_2fa_disable_pending(request, db, user.id)

        if (password or "").strip():
            revoke_user_trusted_browsers(db, updated_user.id)
        if not updated_user.is_two_factor_enabled:
            revoke_active_login_codes(db, updated_user.id)

        if updated_user.id == user.id:
            request.session["user_email"] = updated_user.email
            if not has_admin_access(updated_user):
                _flash(request, "success", "Usuario atualizado com sucesso.")
                return _redirect("/dashboard")

        if pending_super_admin_two_factor_disable and not confirmed_super_admin_two_factor_disable:
            _flash(
                request,
                "error",
                "Confirme antes a desabilitacao do 2FA para concluir essa alteracao.",
            )
            return _redirect_with_query("/usuarios", edit_id=user_id)

        _flash(request, "success", "Usuario atualizado com sucesso.")
        return _redirect("/usuarios")
    except Exception:
        db.rollback()
        _flash(request, "error", "Nao foi possivel atualizar o usuario agora. Revise os dados e tente novamente.")
        return _redirect_with_query("/usuarios", edit_id=user_id)


@router.post("/usuarios/{user_id}/confirmar-desabilitar-2fa")
def confirm_super_admin_two_factor_disable_action(
    user_id: int,
    request: Request,
    csrf_token: str = Form(...),
    code: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)

    if not _has_super_admin_2fa_disable_pending(request, target_user_id=user_id, actor_user_id=user.id):
        return JSONResponse({"ok": False, "message": "Nao ha confirmacao pendente para desabilitar o 2FA deste usuario."}, status_code=400)

    normalized_code = "".join(char for char in (code or "") if char.isdigit())[:6]
    if len(normalized_code) != 6:
        _flash(request, "error", "Informe um codigo numerico de 6 digitos.")
        return _redirect_with_query("/usuarios", edit_id=user_id)

    valid, message = verify_login_code(db, user.id, normalized_code)
    if not valid:
        if not get_active_login_code(db, user.id):
            _clear_super_admin_2fa_disable_pending(request)
        return JSONResponse({"ok": False, "message": message}, status_code=400)

    repo = _repository(db)
    target_user = repo.get_user(user_id)
    if not target_user:
        _discard_super_admin_2fa_disable_pending(request, db, user.id)
        return JSONResponse({"ok": False, "message": "Usuario nao encontrado."}, status_code=404)
    if not is_super_admin_email(target_user.email):
        _discard_super_admin_2fa_disable_pending(request, db, user.id)
        return JSONResponse({"ok": False, "message": "A confirmacao pendente nao e mais valida para este usuario."}, status_code=400)

    _mark_super_admin_2fa_disable_confirmed(request, target_user_id=user_id, actor_user_id=user.id)
    return JSONResponse({"ok": True, "message": "Codigo confirmado. Agora clique em Salvar para concluir a alteracao."})


@router.post("/usuarios/{user_id}/iniciar-desabilitar-2fa")
def start_super_admin_two_factor_disable_action(
    user_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)

    repo = _repository(db)
    target_user = repo.get_user(user_id)
    if not target_user or not is_super_admin_email(target_user.email):
        _discard_super_admin_2fa_disable_pending(request, db, user.id)
        return JSONResponse({"ok": False, "message": "Usuario nao encontrado para esta confirmacao."}, status_code=404)

    try:
        code = issue_login_verification_code(db, user)
        send_access_code_email(user.email, code)
    except Exception:
        db.rollback()
        _discard_super_admin_2fa_disable_pending(request, db, user.id)
        return JSONResponse(
            {"ok": False, "message": "Nao foi possivel enviar o codigo de confirmacao agora. Tente novamente."},
            status_code=500,
        )

    _mark_super_admin_2fa_disable_pending(request, target_user_id=user_id, actor_user_id=user.id)
    return JSONResponse({"ok": True, "message": "Enviamos um codigo de 6 digitos para o seu email."})


@router.post("/usuarios/{user_id}/cancelar-desabilitar-2fa")
def cancel_super_admin_two_factor_disable_action(
    user_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)

    if _has_super_admin_2fa_disable_pending(request, target_user_id=user_id, actor_user_id=user.id):
        _discard_super_admin_2fa_disable_pending(request, db, user.id)

    return JSONResponse({"ok": True})


@router.post("/usuarios/{user_id}/excluir")
def delete_user_action(
    user_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    target_user = repo.get_user(user_id)
    if not target_user:
        _flash(request, "error", "Usuario nao encontrado.")
        return _redirect("/usuarios")
    if target_user.id == user.id:
        _flash(request, "error", "Nao e permitido excluir o usuario atualmente logado.")
        return _redirect("/usuarios")
    repo.delete(target_user)
    _flash(request, "success", "Usuario excluido com sucesso.")
    return _redirect("/usuarios")


@router.get("/backups")
def backups_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    repo = _repository(db)
    page_param = request.query_params.get("page", "1")
    try:
        page = max(1, int(page_param or "1"))
    except (TypeError, ValueError):
        page = 1
    per_page = 10
    total_runs = repo.count_backup_runs()
    total_pages = max(1, (total_runs + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page
    history = [
        {"run": item, "details": _backup_details(item.details_json)}
        for item in repo.list_backup_runs(limit=per_page, offset=offset)
    ]
    return templates.TemplateResponse(
        "backups.html",
        _base_context(
            request,
            user,
            csrf_token,
            "backups",
            title="Backups do Sistema",
            backup_history=history,
            backup_db_bucket=settings.supabase_bucket_db,
            backup_files_bucket=settings.supabase_bucket_files,
            backup_page=page,
            backup_total_pages=total_pages,
            backup_total_runs=total_runs,
            backup_has_prev=page > 1,
            backup_has_next=page < total_pages,
            _repo=repo,
        ),
    )


@router.post("/backups/executar")
def execute_backup_action(
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)
    run = execute_backup(db, initiated_by=user, trigger_source="web_manual")
    if run.status == "success":
        _flash(request, "success", "Backup concluido com sucesso no Supabase Storage.")
    elif run.status == "partial":
        _flash(request, "error", "Backup executado parcialmente. Revise o historico para ver os detalhes.")
    else:
        _flash(request, "error", "Backup nao concluido. Revise o historico para ver o erro detalhado.")
    return _redirect("/backups")


@router.post("/backups/{backup_run_id}/excluir")
def delete_backup_action(
    backup_run_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    denied = _require_admin(request, user)
    if denied:
        return denied
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    run = repo.get_backup_run(backup_run_id)
    page = request.query_params.get("page")
    if not run:
        _flash(request, "error", "Backup nao encontrado.")
        return _redirect_with_query("/backups", page=page)

    try:
        warnings = delete_backup_run(db, run)
    except Exception as exc:
        logger.exception("Falha ao excluir backup. run_id=%s", backup_run_id)
        _flash(request, "error", f"Nao foi possivel excluir o backup agora. {exc}")
        return _redirect_with_query("/backups", page=page)

    if warnings:
        _flash(request, "success", f"Backup removido do historico. {' '.join(warnings)}")
    else:
        _flash(request, "success", "Backup excluido com sucesso.")
    return _redirect_with_query("/backups", page=page)


@router.get("/meu-perfil")
def profile_page(
    request: Request,
    aba: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    active_profile_tab = "security" if (aba or "").strip().lower() == "seguranca" else "profile"
    pending_password_change = None
    if active_profile_tab == "security":
        if _has_password_change_pending(request, user.id):
            pending_password_change = get_active_password_change_verification(db, user.id)
            if not pending_password_change:
                _clear_password_change_pending(request)
        else:
            stale_password_change = get_active_password_change_verification(db, user.id)
            if stale_password_change:
                _discard_password_change_pending(request, db, user.id)
    elif _has_password_change_pending(request, user.id):
        _discard_password_change_pending(request, db, user.id)
    return _render_profile_page(
        request,
        user,
        csrf_token,
        repo,
        pending_password_change=pending_password_change,
        active_profile_tab=active_profile_tab,
    )


@router.post("/meu-perfil")
async def update_profile_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    display_name: str | None = Form(None),
    gender: str | None = Form(None),
    birth_date: str | None = Form(None),
    phone: str | None = Form(None),
    email: str = Form(...),
    job_title: str | None = Form(None),
    notes: str | None = Form(None),
    avatar: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    validate_csrf(request, csrf_token)
    repo = _repository(db)

    normalized_name = _clean_text(name)
    normalized_email = (email or "").strip().lower()
    normalized_gender = (gender or "").strip()
    if normalized_gender and normalized_gender not in {option[0] for option in _profile_gender_options()}:
        normalized_gender = None

    if not normalized_name or not normalized_email:
        _flash(request, "error", "Nome completo e email sao obrigatorios.")
        return _redirect("/meu-perfil?aba=perfil")

    existing = db.query(User).filter(User.email == normalized_email, User.id != user.id).first()
    if existing:
        _flash(request, "error", "Ja existe outro usuario com este email.")
        return _redirect("/meu-perfil?aba=perfil")

    try:
        normalized_birth_date = _date_or_none(birth_date)
    except ValueError:
        _flash(request, "error", "Informe uma data de nascimento valida.")
        return _redirect("/meu-perfil?aba=perfil")

    avatar_payload, avatar_error = await _read_avatar_upload(avatar)
    if avatar_error:
        _flash(request, "error", avatar_error)
        return _redirect("/meu-perfil?aba=perfil")

    payload = {
        "name": normalized_name,
        "email": normalized_email,
        "display_name": _clean_text(display_name),
        "gender": normalized_gender or None,
        "birth_date": normalized_birth_date,
        "phone": _clean_text(phone),
        "job_title": _clean_text(job_title),
        "notes": (notes or "").strip() or None,
    }
    if avatar_payload:
        payload.update(avatar_payload)

    updated_user = repo.update(user, payload)
    request.session["user_email"] = updated_user.email
    _flash(request, "success", "Perfil atualizado com sucesso.")
    return _redirect("/meu-perfil?aba=perfil")


@router.post("/meu-perfil/avatar/remover")
def remove_profile_avatar_action(
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    repo.update(
        user,
        {
            "avatar_filename": None,
            "avatar_content_type": None,
            "avatar_data": None,
        },
    )
    _flash(request, "success", "Foto removida com sucesso.")
    return _redirect("/meu-perfil?aba=perfil")


@router.get("/meu-perfil/avatar")
def profile_avatar_view(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    current_user = _repository(db).get_user(user.id)
    if not current_user or not current_user.avatar_data:
        return Response(status_code=status.HTTP_404_NOT_FOUND)
    return Response(
        content=current_user.avatar_data,
        media_type=current_user.avatar_content_type or "image/jpeg",
        headers={"Cache-Control": "no-store"},
    )


@router.post("/meu-perfil/alterar-senha")
def change_own_password_action(
    request: Request,
    csrf_token: str = Form(...),
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    validate_csrf(request, csrf_token)
    normalized_current_password = (current_password or "").strip()
    normalized_new_password = (new_password or "").strip()
    normalized_confirm_password = (confirm_password or "").strip()

    if not verify_password(normalized_current_password, user.hashed_password):
        _flash(request, "error", "A senha atual informada nao confere.")
        return _redirect("/meu-perfil?aba=seguranca")
    if not normalized_new_password:
        _flash(request, "error", "Informe a nova senha.")
        return _redirect("/meu-perfil?aba=seguranca")
    if normalized_new_password != normalized_confirm_password:
        _flash(request, "error", "A confirmacao da nova senha nao confere.")
        return _redirect("/meu-perfil?aba=seguranca")
    if verify_password(normalized_new_password, user.hashed_password):
        _flash(request, "error", "A nova senha precisa ser diferente da senha atual.")
        return _redirect("/meu-perfil?aba=seguranca")

    try:
        code = issue_password_change_verification(db, user, get_password_hash(normalized_new_password))
        send_password_change_code_email(user.email, code, get_settings().two_factor_code_minutes)
    except RuntimeError as exc:
        _discard_password_change_pending(request, db, user.id)
        _flash(request, "error", str(exc))
        return _redirect("/meu-perfil?aba=seguranca")
    except Exception:
        db.rollback()
        _discard_password_change_pending(request, db, user.id)
        _flash(request, "error", "Nao foi possivel enviar o codigo de confirmacao agora. Tente novamente.")
        return _redirect("/meu-perfil?aba=seguranca")

    _mark_password_change_pending(request, user.id)
    _flash(request, "success", "Enviamos um codigo de confirmacao para o seu email. Informe o codigo para concluir a alteracao da senha.")
    return _redirect("/meu-perfil?aba=seguranca")


@router.post("/meu-perfil/alterar-senha/confirmar")
def confirm_own_password_change_action(
    request: Request,
    csrf_token: str = Form(...),
    code: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    validate_csrf(request, csrf_token)
    normalized_code = "".join(char for char in (code or "") if char.isdigit())[:6]
    if len(normalized_code) != 6:
        _flash(request, "error", "Informe um codigo numerico de 6 digitos.")
        return _redirect("/meu-perfil?aba=seguranca")

    valid, message, verification = verify_password_change_code(db, user.id, normalized_code)
    if not valid or not verification:
        if not get_active_password_change_verification(db, user.id):
            _clear_password_change_pending(request)
        _flash(request, "error", message)
        return _redirect("/meu-perfil?aba=seguranca")

    repo = _repository(db)
    repo.update(user, {"hashed_password": verification.new_password_hash})
    revoke_password_change_verifications(db, user.id)
    _clear_password_change_pending(request)
    revoke_user_password_reset_tokens(db, user.id)
    revoke_user_trusted_browsers(db, user.id)
    revoke_active_login_codes(db, user.id)
    _flash(request, "success", "Senha atualizada com sucesso.")
    return _redirect("/meu-perfil?aba=seguranca")


@router.post("/meu-perfil/alterar-senha/cancelar")
async def cancel_own_password_change_action(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    form = await request.form()
    validate_csrf(request, str(form.get("csrf_token") or ""))
    _discard_password_change_pending(request, db, user.id)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/insumos/comprados")
def purchased_inputs_page(
    request: Request,
    edit_id: int | None = None,
    item_type: str | None = None,
    farm_id: int | None = None,
    purchased_tab: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    effective_farm_id = farm_id or scope["active_farm_id"]
    finance_accounts = repo.list_finance_accounts(farm_id=effective_farm_id) if effective_farm_id else []
    edit_input = repo.get_purchased_input(edit_id) if edit_id else None
    selected_item_type = item_type if item_type in {"insumo_agricola", "combustivel", "all"} else None
    normalized_item_type = item_type if item_type in {"insumo_agricola", "combustivel"} else None
    if not normalized_item_type and edit_input and edit_input.input_catalog:
        normalized_item_type = edit_input.input_catalog.item_type
    if not normalized_item_type and selected_item_type != "all":
        normalized_item_type = "insumo_agricola"
    stock_context = _build_stock_context(repo, farm_id=effective_farm_id, item_type=normalized_item_type)
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    purchase_entries_pagination = _paginate_collection(request, purchase_entries, "entries_page")
    stock_outputs_pagination = _paginate_collection(request, stock_outputs, "outputs_page")
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )
    extract_rows_pagination = _paginate_collection(request, extract_rows, "extract_page")
    selected_purchased_tab = str(request.query_params.get("purchased_tab") or purchased_tab or "entries")
    if selected_purchased_tab not in {"entries", "outputs", "extract"}:
        selected_purchased_tab = "entries"
    return templates.TemplateResponse(
        "purchased_inputs.html",
        _base_context(
            request,
            user,
            csrf_token,
            "purchased_inputs",
            _repo=repo,
            title="Compras de Insumos",
            farms=repo.list_farms(),
            selected_item_type=selected_item_type or normalized_item_type or "insumo_agricola",
            selected_farm_id=effective_farm_id,
            inputs=purchase_entries_pagination["items"],
            inputs_pagination=purchase_entries_pagination,
            inputs_catalog=stock_context["catalog_inputs"],
            inputs_catalog_all=repo.list_input_catalog(),
            finance_account_options=finance_accounts,
            finance_account_default=next((item for item in finance_accounts if item.is_default), None),
            input_stock=stock_context["input_stock"],
            stock_outputs=stock_outputs_pagination["items"],
            stock_outputs_pagination=stock_outputs_pagination,
            extract_rows=extract_rows_pagination["items"],
            extract_rows_pagination=extract_rows_pagination,
            purchased_inputs_export_query=_purchased_inputs_export_query(
                farm_id=effective_farm_id,
                item_type=selected_item_type if selected_item_type in {"insumo_agricola", "combustivel"} else None,
                purchased_tab=selected_purchased_tab,
            ),
            edit_input=edit_input,
            selected_purchased_tab=selected_purchased_tab,
            purchased_tab_urls={
                "entries": _url_with_query(request, purchased_tab="entries"),
                "outputs": _url_with_query(request, purchased_tab="outputs"),
                "extract": _url_with_query(request, purchased_tab="extract"),
            },
        ),
    )


@router.post("/insumos/comprados")
async def create_purchased_input_action(
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str | None = Form(None),
    item_type: str = Form("insumo_agricola"),
    name: str = Form(...),
    quantity_purchased: float = Form(...),
    package_size: float = Form(...),
    package_unit: str = Form(...),
    unit_price: float = Form(...),
    finance_account_id: str | None = Form(None),
    purchase_date: str | None = Form(None),
    low_stock_threshold: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/comprados")
    if denied:
        return denied
    try:
        finance_account = _resolve_optional_finance_account(repo, active_farm=scope["active_farm"], account_id=finance_account_id)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_with_query("/insumos/comprados", item_type=item_type)
    try:
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_with_query("/insumos/comprados", item_type=item_type)
    item = create_purchased_input(
        repo,
        {
            "farm_id": scope["active_farm_id"],
            "item_type": item_type,
            "name": name,
            "quantity_purchased": quantity_purchased,
            "package_size": package_size,
            "package_unit": package_unit,
            "unit_price": unit_price,
            "finance_account_id": finance_account.id if finance_account else None,
            "purchase_date": purchase_date,
            "low_stock_threshold": low_stock_threshold,
            "notes": notes,
        },
    )
    try:
        saved_attachments = _save_purchased_input_attachments(repo, item, attachment_payloads)
    except Exception:
        _flash(request, "error", "O lancamento foi salvo, mas nao foi possivel gravar os anexos agora.")
        return _redirect_with_query("/insumos/comprados", edit_id=item.id, item_type=item_type)
    if saved_attachments:
        _flash(request, "success", f"Insumo comprado cadastrado com sucesso. {saved_attachments} anexo(s) salvo(s).")
        return _redirect_with_query("/insumos/comprados", item_type=item_type)
    _flash(request, "success", "Insumo comprado cadastrado com sucesso.")
    return _redirect_with_query("/insumos/comprados", item_type=item_type)


@router.post("/insumos/comprados/{input_id}/editar")
async def update_purchased_input_action(
    input_id: int,
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str | None = Form(None),
    item_type: str = Form("insumo_agricola"),
    name: str = Form(...),
    quantity_purchased: float = Form(...),
    package_size: float = Form(...),
    package_unit: str = Form(...),
    unit_price: float = Form(...),
    finance_account_id: str | None = Form(None),
    purchase_date: str | None = Form(None),
    low_stock_threshold: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/comprados")
    if denied:
        return denied
    item = repo.get_purchased_input(input_id)
    if not item:
        _flash(request, "error", "Insumo comprado nao encontrado.")
        return _redirect_for_request(request, "/insumos/comprados")
    if not _farm_matches_scope(item.farm_id, scope):
        _flash(request, "error", "Este lancamento de entrada nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/insumos/comprados")
    try:
        finance_account = _resolve_optional_finance_account(repo, active_farm=scope["active_farm"], account_id=finance_account_id)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id, item_type=item_type)
    try:
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id, item_type=item_type)
    update_purchased_input(
        repo,
        item,
        {
            "farm_id": scope["active_farm_id"],
            "item_type": item_type,
            "name": name,
            "quantity_purchased": quantity_purchased,
            "package_size": package_size,
            "package_unit": package_unit,
            "unit_price": unit_price,
            "finance_account_id": finance_account.id if finance_account else None,
            "purchase_date": purchase_date,
            "low_stock_threshold": low_stock_threshold,
            "notes": notes,
        },
    )
    try:
        saved_attachments = _save_purchased_input_attachments(repo, item, attachment_payloads)
    except Exception:
        _flash(request, "error", "As alteracoes foram salvas, mas nao foi possivel incluir os novos anexos.")
        return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id, item_type=item_type)
    if saved_attachments:
        _flash(request, "success", f"Alteracoes salvas com sucesso. {saved_attachments} novo(s) anexo(s) adicionado(s).")
        return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id, item_type=item_type)
    _flash(request, "success", "Insumo comprado atualizado com sucesso.")
    return _redirect_for_request(request, "/insumos/comprados", item_type=item_type)


@router.post("/insumos/comprados/{input_id}/excluir")
def delete_purchased_input_action(
    input_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    item = repo.get_purchased_input(input_id)
    if not item:
        _flash(request, "error", "Insumo comprado nao encontrado.")
        return _redirect_for_request(request, "/insumos/comprados")
    if item.recommendations:
        _flash(request, "error", "Nao e possivel excluir o insumo enquanto houver recomendacoes vinculadas.")
        return _redirect("/insumos/comprados")
    if item.recommendation_items or item.schedule_items or item.stock_allocations or item.stock_outputs:
        _flash(request, "error", "Nao e possivel excluir o insumo enquanto houver recomendacoes, agendamentos ou aplicacoes vinculadas.")
        return _redirect("/insumos/comprados")
    repo.delete(item)
    _flash(request, "success", "Insumo comprado excluido com sucesso.")
    return _redirect("/insumos/comprados")


@router.get("/insumos/comprados/anexos/{attachment_id}")
def open_purchased_input_attachment(
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    attachment = repo.get_purchased_input_attachment(attachment_id)
    if not attachment or not attachment.purchased_input:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect("/insumos/comprados")
    scope = _global_scope_context(request, repo)
    item = attachment.purchased_input
    if not _farm_matches_scope(item.farm_id, scope):
        _flash(request, "error", "Este anexo nao pertence ao contexto ativo.")
        return _redirect("/insumos/comprados")
    return _attachment_response(attachment.filename, attachment.content_type, attachment.file_data)


@router.post("/insumos/comprados/{input_id}/anexos/{attachment_id}/excluir")
def delete_purchased_input_attachment_action(
    input_id: int,
    attachment_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    item = repo.get_purchased_input(input_id)
    if not item:
        _flash(request, "error", "Insumo comprado nao encontrado.")
        return _redirect("/insumos/comprados")
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(item.farm_id, scope):
        _flash(request, "error", "Este lancamento de entrada nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/insumos/comprados")
    attachment = repo.get_purchased_input_attachment(attachment_id)
    if not attachment or attachment.purchased_input_id != item.id:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id)
    repo.delete(attachment)
    if repo.get_purchased_input_attachment(attachment_id):
        _flash(request, "error", "Nao foi possivel remover o anexo agora. Tente novamente.")
        return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id)
    _flash(request, "success", "Anexo removido com sucesso.")
    item_type = item.input_catalog.item_type if item.input_catalog else None
    return _redirect_for_request(request, "/insumos/comprados", edit_id=input_id, item_type=item_type)


@router.get("/insumos/comprados/exportar.xlsx")
def export_purchased_inputs_xlsx(
    request: Request,
    farm_id: str | None = None,
    item_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    normalized_item_type = item_type if item_type in {"insumo_agricola", "combustivel"} else None
    export_tab = _export_purchased_tab_param(request)
    stock_context = _build_stock_context(repo, farm_id=selected_farm_id, item_type=normalized_item_type)
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )

    workbook = Workbook()
    sheet = workbook.active
    if export_tab == "entries":
        sheet.title = "Entradas"
        _xlsx_write_purchase_entries(sheet, purchase_entries)
    elif export_tab == "outputs":
        sheet.title = "Saídas"
        _xlsx_write_stock_outputs(sheet, stock_outputs)
    else:
        sheet.title = "Extrato"
        _xlsx_write_extract_rows(sheet, extract_rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="gestao_compras.xlsx"'},
    )


@router.get("/insumos/comprados/exportar.pdf")
def export_purchased_inputs_pdf(
    request: Request,
    farm_id: str | None = None,
    item_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    export_tab = _export_purchased_tab_param(request)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_farm = repo.get_farm(selected_farm_id) if selected_farm_id else None
    normalized_item_type = item_type if item_type in {"insumo_agricola", "combustivel"} else None
    stock_context = _build_stock_context(repo, farm_id=selected_farm_id, item_type=normalized_item_type)
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )
    extract_totals = _stock_report_totals(extract_rows)

    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    farm_name = selected_farm.name if selected_farm else "Fazenda Bela Vista"
    item_type_label = {
        "insumo_agricola": "Insumos agrícolas",
        "combustivel": "Combustíveis",
    }.get(normalized_item_type or "", "Todos os itens")
    entries_total = sum(float(item.total_cost or 0) for item in purchase_entries)
    outputs_total = sum(float(output.total_cost or 0) for output in stock_outputs)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle("PurchasedPdfFarmHeader", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, alignment=TA_RIGHT, textColor=colors.HexColor("#1e293b"))
    meta_label_style = ParagraphStyle("PurchasedPdfMetaLabel", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#446a36"), spaceAfter=2)
    meta_value_style = ParagraphStyle("PurchasedPdfMetaValue", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13, textColor=colors.HexColor("#334155"))
    cell_style = ParagraphStyle("PurchasedPdfCell", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.2, leading=10.4, textColor=colors.HexColor("#0f172a"))
    cell_muted_style = ParagraphStyle("PurchasedPdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    cell_numeric_style = ParagraphStyle("PurchasedPdfCellNumeric", parent=cell_style, alignment=TA_RIGHT)
    summary_value_style = ParagraphStyle("PurchasedPdfSummaryValue", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#0f172a"))

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    tab_label = {"entries": "Entradas", "outputs": "Saídas", "extract": "Extrato"}.get(export_tab, "Entradas")
    if export_tab == "extract":
        summary_table = Table([
            [
                [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                [Paragraph("ESCOPO", meta_label_style), Paragraph(item_type_label, meta_value_style)],
                [Paragraph("LISTAGEM", meta_label_style), Paragraph(tab_label, meta_value_style)],
            ],
            [
                [Paragraph("LANÇAMENTOS", meta_label_style), Paragraph(str(extract_totals["movements_count"]), summary_value_style)],
                [Paragraph("TOTAL MOVIMENTADO", meta_label_style), Paragraph(_format_currency(extract_totals["grand_total"]), summary_value_style)],
                [Paragraph("", meta_value_style), Paragraph("", meta_value_style)],
            ],
        ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    elif export_tab == "entries":
        summary_table = Table([
            [
                [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                [Paragraph("ESCOPO", meta_label_style), Paragraph(item_type_label, meta_value_style)],
                [Paragraph("LISTAGEM", meta_label_style), Paragraph(tab_label, meta_value_style)],
            ],
            [
                [Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(purchase_entries)), summary_value_style)],
                [Paragraph("TOTAL FINANCEIRO", meta_label_style), Paragraph(_format_currency(entries_total), summary_value_style)],
                [Paragraph("", meta_value_style), Paragraph("", meta_value_style)],
            ],
        ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    else:
        summary_table = Table([
            [
                [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                [Paragraph("ESCOPO", meta_label_style), Paragraph(item_type_label, meta_value_style)],
                [Paragraph("LISTAGEM", meta_label_style), Paragraph(tab_label, meta_value_style)],
            ],
            [
                [Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(stock_outputs)), summary_value_style)],
                [Paragraph("TOTAL FINANCEIRO", meta_label_style), Paragraph(_format_currency(outputs_total), summary_value_style)],
                [Paragraph("", meta_value_style), Paragraph("", meta_value_style)],
            ],
        ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")

    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 14)]

    if export_tab == "extract":
        elements.extend(
            _pdf_flowables_extract_detail_table(
                doc,
                extract_rows,
                extract_totals,
                cell_style=cell_style,
                cell_muted_style=cell_muted_style,
                cell_numeric_style=cell_numeric_style,
                meta_value_style=meta_value_style,
                summary_value_style=summary_value_style,
            )
        )
    else:
        table_data = [["Mov.", "Data", "Insumo", "Tipo", "Origem / Fazenda", "Quantidade", "Valor", "Observações"]]
        report_rows = []
        if export_tab == "entries":
            for item in purchase_entries:
                report_rows.append({
                    "kind": "Entrada",
                    "date": item.purchase_date,
                    "name": item.input_catalog.name if item.input_catalog else item.name,
                    "type": _catalog_item_type_label_pt(item.input_catalog),
                    "origin": item.farm.name if item.farm else "Sem fazenda vinculada",
                    "quantity": f"{_format_decimal_br(item.total_quantity, 2)} {item.package_unit}",
                    "value": float(item.total_cost or 0),
                    "notes": item.notes or "-",
                    "sort_key": (item.purchase_date or today_in_app_timezone(), 0, item.id),
                })
        else:
            for output in stock_outputs:
                report_rows.append({
                    "kind": "Saída",
                    "date": output.movement_date,
                    "name": output.input_catalog.name if output.input_catalog else "Insumo removido",
                    "type": _catalog_item_type_label_pt(output.input_catalog),
                    "origin": f"{output.origin} • {output.farm.name if output.farm else 'Sem fazenda'}{(' / ' + output.plot.name) if output.plot else ''}",
                    "quantity": f"{_format_decimal_br(output.quantity, 2)} {output.unit}",
                    "value": float(output.total_cost or 0),
                    "notes": output.notes or "-",
                    "sort_key": (output.movement_date or today_in_app_timezone(), 1, output.id),
                })
        report_rows.sort(key=lambda row: row["sort_key"], reverse=True)

        for row in report_rows:
            movement_color = "#166534" if row["kind"] == "Entrada" else "#be123c"
            table_data.append([
                Paragraph(f'<font color="{movement_color}"><b>{row["kind"]}</b></font>', cell_style),
                Paragraph(row["date"].strftime("%d/%m/%Y") if row["date"] else "-", cell_style),
                Paragraph(row["name"], cell_style),
                Paragraph(row["type"], cell_muted_style),
                Paragraph(row["origin"], cell_muted_style),
                Paragraph(row["quantity"], cell_numeric_style),
                Paragraph(_format_currency(row["value"]), cell_numeric_style),
                Paragraph(row["notes"][:90], cell_muted_style),
            ])

        column_weights = [7, 8, 18, 12, 20, 11, 12, 20]
        weight_total = sum(column_weights)
        table_col_widths = [doc.width * (weight / weight_total) for weight in column_weights[:-1]]
        table_col_widths.append(doc.width - sum(table_col_widths))
        table = Table(table_data, repeatRows=1, colWidths=table_col_widths, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#36552a")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("ALIGN", (5, 1), (6, -1), "RIGHT"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))

        if export_tab == "entries":
            footer_line = Paragraph(f"<b>Total (entradas):</b> {_format_currency(entries_total)}", summary_value_style)
        else:
            footer_line = Paragraph(f"<b>Total (saídas):</b> {_format_currency(outputs_total)}", summary_value_style)
        footer_summary = Table([[footer_line]], colWidths=[doc.width], hAlign="LEFT")
        footer_summary.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef6ee")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cfe1d0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(footer_summary)

    generated_at_label = generated_at.strftime("%d/%m/%Y %H:%M")

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="gestao_compras.pdf"'},
    )


@router.get("/insumos/estoque")
def stock_page(
    request: Request,
    edit_output_id: int | None = None,
    farm_id: str | None = None,
    input_id: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str = "all",
    item_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    selected_farm_id = _int_or_none(farm_id) or scope["active_farm_id"]
    selected_input_id = _int_or_none(input_id)
    start = _date_or_none(start_date)
    end = _date_or_none(end_date)
    normalized_item_type = item_type if item_type in {"insumo_agricola", "combustivel"} else None
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    stock_context = _build_stock_context(
        repo,
        farm_id=selected_farm_id,
        input_id=selected_input_id,
        start_date=start,
        end_date=end,
        movement_type=movement_type,
        item_type=normalized_item_type,
    )
    raw_stock_tab = request.query_params.get("stock_tab")
    if raw_stock_tab in {"entries", "outputs", "extract"}:
        selected_stock_tab = raw_stock_tab
    else:
        selected_stock_tab = "entries"
        if movement_type == "entrada":
            selected_stock_tab = "entries"
        elif movement_type == "saida":
            selected_stock_tab = "outputs"
    if selected_stock_tab not in {"entries", "outputs", "extract"}:
        selected_stock_tab = "entries"
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )
    purchase_entries_pagination = _paginate_collection(request, purchase_entries, "entries_page")
    stock_outputs_pagination = _paginate_collection(request, stock_outputs, "outputs_page")
    extract_rows_pagination = _paginate_collection(request, extract_rows, "extract_page")
    edit_output = repo.get_stock_output(edit_output_id) if edit_output_id else None
    edit_output_mode = None
    edit_fertilization_item = None
    edit_fertilization_record = None
    edit_inputs_catalog = repo.list_input_catalog()
    if edit_output:
        if edit_output.reference_type == "manual_stock_output":
            edit_output_mode = "manual"
        elif edit_output.reference_type in {"fertilization_item", "fertilization_record"}:
            edit_fertilization_item, edit_fertilization_record = _resolve_fertilization_output_context(repo, db, edit_output)
            if edit_fertilization_record:
                edit_output_mode = "fertilization"
                edit_inputs_catalog = repo.list_input_catalog(item_type="insumo_agricola")
            else:
                edit_output = None
        else:
            edit_output = None
    stock_filters_active = _stock_page_filters_active(
        request,
        active_farm_id=scope["active_farm_id"],
        selected_input_id=selected_input_id,
        movement_type=movement_type,
        normalized_item_type=normalized_item_type,
        edit_output_id=edit_output_id,
    )
    return templates.TemplateResponse(
        "stock.html",
        _base_context(
            request,
            user,
            csrf_token,
            "stock",
            _repo=repo,
            title="Estoque de Insumos",
            farms=repo.list_farms(),
            plots=repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids),
            selected_farm_id=selected_farm_id,
            selected_input_id=selected_input_id,
            selected_start_date=start_date,
            selected_end_date=end_date,
            selected_movement_type=movement_type,
            selected_item_type=normalized_item_type or "",
            stock_filters_active=stock_filters_active,
            stock_export_query=_stock_export_query(
                farm_id=selected_farm_id,
                input_id=selected_input_id,
                start_date=start,
                end_date=end,
                movement_type=movement_type,
                item_type=normalized_item_type,
                stock_tab=selected_stock_tab,
            ),
            inputs_catalog=stock_context["catalog_inputs"],
            input_stock=stock_context["input_stock"],
            stock_catalog_rows=stock_context["stock_catalog_rows"],
            purchase_entries=purchase_entries_pagination["items"],
            purchase_entries_pagination=purchase_entries_pagination,
            stock_outputs=stock_outputs_pagination["items"],
            stock_outputs_pagination=stock_outputs_pagination,
            extract_rows=extract_rows_pagination["items"],
            extract_rows_pagination=extract_rows_pagination,
            selected_stock_tab=selected_stock_tab,
            stock_tab_urls={
                "entries": _url_with_query(request, stock_tab="entries"),
                "outputs": _url_with_query(request, stock_tab="outputs"),
                "extract": _url_with_query(request, stock_tab="extract"),
            },
            edit_output=edit_output,
            edit_output_mode=edit_output_mode,
            edit_fertilization_item=edit_fertilization_item,
            edit_fertilization_record=edit_fertilization_record,
            edit_inputs_catalog=edit_inputs_catalog,
        ),
    )


@router.post("/insumos/estoque/saida-manual")
def create_manual_stock_output_action(
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str | None = Form(None),
    plot_id: str | None = Form(None),
    input_id: int = Form(...),
    movement_date: str | None = Form(None),
    quantity: float = Form(...),
    unit: str = Form(...),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, scope, denied = _resolve_optional_plot_in_scope(
        request,
        repo,
        _int_or_none(plot_id),
        "/insumos/estoque",
    )
    if denied:
        return denied
    try:
        create_manual_stock_output(
            repo,
            {
                "farm_id": scope["active_farm_id"],
                "plot_id": plot.id if plot else None,
                "input_id": input_id,
                "season_id": scope["active_season_id"],
                "movement_date": movement_date,
                "quantity": quantity,
                "unit": unit,
                "notes": notes,
            },
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect("/insumos/estoque")
    _flash(request, "success", "Saida manual registrada com sucesso.")
    return _redirect("/insumos/estoque")


@router.get("/insumos/estoque/saidas/{output_id}/editar")
def edit_stock_output_entry(
    output_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    output = repo.get_stock_output(output_id)
    if not output:
        _flash(request, "error", "Lancamento de saida nao encontrado.")
        return _redirect("/insumos/estoque")
    if output.reference_type in {"manual_stock_output", "fertilization_item", "fertilization_record"}:
        return _redirect_with_query("/insumos/estoque", edit_output_id=output_id)
    _flash(request, "error", f"Este lancamento ainda nao possui edicao integrada para o modulo {output.origin}.")
    return _redirect("/insumos/estoque")


@router.post("/insumos/estoque/saidas/{output_id}/editar")
async def update_stock_output_entry_action(
    output_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    csrf_token = str(form.get("csrf_token") or "")
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    output = repo.get_stock_output(output_id)
    if not output:
        _flash(request, "error", "Lancamento de saida nao encontrado.")
        return _redirect("/insumos/estoque")
    try:
        if output.reference_type == "manual_stock_output":
            scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/estoque")
            if denied:
                return denied
            if not _farm_matches_scope(output.farm_id, scope) or (output.plot_id and not _plot_matches_scope(output.plot, scope)):
                _flash(request, "error", "Este lancamento de saida nao pertence ao contexto ativo.")
                return _redirect("/insumos/estoque")
            plot, _, invalid_plot = _resolve_optional_plot_in_scope(
                request,
                repo,
                _int_or_none(form.get("plot_id")),
                "/insumos/estoque",
            )
            if invalid_plot:
                return invalid_plot
            update_manual_stock_output(
                repo,
                output,
                {
                    "farm_id": scope["active_farm_id"],
                    "plot_id": plot.id if plot else None,
                    "movement_date": str(form.get("movement_date") or ""),
                    "quantity": float(form.get("quantity") or 0),
                    "unit": str(form.get("unit") or ""),
                    "notes": str(form.get("notes") or "") or None,
                },
            )
        elif output.reference_type in {"fertilization_item", "fertilization_record"}:
            fertilization_item, fertilization = _resolve_fertilization_output_context(repo, db, output)
            if not fertilization_item or not fertilization:
                raise ValueError("Nao foi possivel localizar o item de fertilizacao vinculado.")
            plot, scope, invalid_plot = _resolve_plot_in_scope(
                request,
                repo,
                int(fertilization.plot_id or 0),
                "/insumos/estoque",
            )
            if invalid_plot:
                return invalid_plot
            input_catalog = repo.get_input_catalog(int(fertilization_item.input_id or 0))
            if not input_catalog or input_catalog.item_type != "insumo_agricola" or not input_catalog.is_active:
                raise ValueError("Selecione apenas insumos agrícolas válidos para a fertilização.")
            quantity = float(form.get("quantity") or 0)
            if quantity <= 0:
                raise ValueError("Informe uma quantidade valida para a saida.")
            unit = str(form.get("unit") or input_catalog.default_unit or fertilization_item.unit or "kg").strip()
            if not unit:
                raise ValueError("Informe uma unidade valida para a saida.")

            items = []
            for item in fertilization.items:
                if item.id == fertilization_item.id:
                    items.append(
                        {
                            "input_id": input_catalog.id,
                            "purchased_input_id": None,
                            "name": input_catalog.name,
                            "unit": unit,
                            "quantity": quantity,
                        }
                    )
                else:
                    items.append(
                        {
                            "input_id": item.input_id,
                            "purchased_input_id": None,
                            "name": item.name,
                            "unit": item.unit,
                            "quantity": float(item.total_quantity or 0),
                        }
                    )
            update_fertilization(
                repo,
                fertilization,
                {
                    "plot_id": plot.id,
                    "application_date": str(form.get("movement_date") or fertilization.application_date.isoformat()),
                    "season_id": scope["active_season_id"],
                    "application_method": fertilization.application_method,
                    "duration_minutes": fertilization.duration_minutes,
                    "notes": str(form.get("notes") or "") or None,
                    "items": items,
                },
            )
        else:
            raise ValueError(f"Este lancamento ainda nao possui edicao integrada para o modulo {output.origin}.")
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_with_query("/insumos/estoque", edit_output_id=output_id)
    _flash(request, "success", "Lancamento atualizado com sucesso.")
    return _redirect("/insumos/estoque")


@router.post("/insumos/estoque/saidas/{output_id}/excluir")
def delete_stock_output_entry_action(
    output_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    output = repo.get_stock_output(output_id)
    if not output:
        _flash(request, "error", "Lancamento de saida nao encontrado.")
        return _redirect("/insumos/estoque")
    try:
        delete_manual_stock_output(repo, output)
    except ValueError:
        _flash(request, "error", f"Este lancamento esta vinculado ao modulo {output.origin}. Exclua por la.")
        return _redirect("/insumos/estoque")
    _flash(request, "success", "Saida manual excluida com sucesso.")
    return _redirect("/insumos/estoque")


@router.get("/insumos/estoque/exportar.xlsx")
def export_stock_extract_xlsx(
    request: Request,
    farm_id: str | None = None,
    input_id: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str = "all",
    item_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_input_id = _int_or_none(input_id)
    normalized_item_type = item_type if item_type in {"insumo_agricola", "combustivel"} else None
    export_tab = _export_stock_tab_param(request, movement_type)
    stock_context = _build_stock_context(
        repo,
        farm_id=selected_farm_id,
        input_id=selected_input_id,
        start_date=_date_or_none(start_date),
        end_date=_date_or_none(end_date),
        movement_type=movement_type,
        item_type=normalized_item_type,
    )
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )
    workbook = Workbook()
    sheet = workbook.active
    if export_tab == "entries":
        sheet.title = "Entradas"
        _xlsx_write_purchase_entries(sheet, purchase_entries)
    elif export_tab == "outputs":
        sheet.title = "Saidas"
        _xlsx_write_stock_outputs(sheet, stock_outputs)
    else:
        sheet.title = "Extrato de estoque"
        _xlsx_write_extract_rows(sheet, extract_rows)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="extrato_estoque.xlsx"'},
    )


@router.get("/insumos/estoque/exportar.pdf")
def export_stock_extract_pdf(
    request: Request,
    farm_id: str | None = None,
    input_id: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str = "all",
    item_type: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_input_id = _int_or_none(input_id)
    normalized_item_type = item_type if item_type in {"insumo_agricola", "combustivel"} else None
    export_tab = _export_stock_tab_param(request, movement_type)
    stock_context = _build_stock_context(
        repo,
        farm_id=selected_farm_id,
        input_id=selected_input_id,
        start_date=_date_or_none(start_date),
        end_date=_date_or_none(end_date),
        movement_type=movement_type,
        item_type=normalized_item_type,
    )
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )
    entries_total_sum = sum(float(item.total_cost or 0) for item in purchase_entries)
    outputs_total_sum = sum(float(output.total_cost or 0) for output in stock_outputs)
    extract_totals = _stock_report_totals(extract_rows)
    selected_farm = repo.get_farm(selected_farm_id) if selected_farm_id else None
    generated_at = app_now()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle(
        "StockPdfFarmHeader",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#1e293b"),
    )
    meta_label_style = ParagraphStyle(
        "StockPdfMetaLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#446a36"),
        spaceAfter=2,
    )
    meta_value_style = ParagraphStyle(
        "StockPdfMetaValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#334155"),
    )
    cell_style = ParagraphStyle(
        "StockPdfCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.3,
        leading=10.5,
        textColor=colors.HexColor("#0f172a"),
    )
    cell_muted_style = ParagraphStyle(
        "StockPdfCellMuted",
        parent=cell_style,
        textColor=colors.HexColor("#475569"),
    )
    cell_numeric_style = ParagraphStyle(
        "StockPdfCellNumeric",
        parent=cell_style,
        alignment=TA_RIGHT,
    )
    summary_value_style = ParagraphStyle(
        "StockPdfSummaryValue",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#0f172a"),
    )

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    farm_name = selected_farm.name if selected_farm else "Fazenda Bela Vista"
    movement_label = {
        "entrada": "Somente entradas",
        "saida": "Somente saídas",
        "all": "Entradas, saídas e extrato consolidado",
    }.get(movement_type, "Entradas, saídas e extrato consolidado")
    item_type_label = {
        "insumo_agricola": "Insumos agrícolas",
        "combustivel": "Combustíveis",
    }.get(normalized_item_type or "", "Todos os consumíveis")
    period_label = "Período completo"
    if start_date or end_date:
        period_label = f"{start_date or 'Início'} até {end_date or 'Hoje'}"
    listagem_label = {"entries": "Entradas", "outputs": "Saídas", "extract": "Extrato"}.get(export_tab, "Extrato")

    header_table = Table(
        [[
            logo_flowable,
            Paragraph(farm_name, farm_header_style),
        ]],
        colWidths=[76, doc.width - 76],
        hAlign="LEFT",
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    if export_tab == "extract":
        summary_table = Table(
            [
                [
                    [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                    [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
                    [Paragraph("ESCOPO", meta_label_style), Paragraph(item_type_label, meta_value_style)],
                ],
                [
                    [Paragraph("LISTAGEM", meta_label_style), Paragraph(listagem_label, meta_value_style)],
                    [Paragraph("MOVIMENTAÇÕES", meta_label_style), Paragraph(movement_label, meta_value_style)],
                    [Paragraph("INSUMO", meta_label_style), Paragraph("Filtrado" if selected_input_id else "Todos", meta_value_style)],
                ],
                [
                    [Paragraph("LANÇAMENTOS", meta_label_style), Paragraph(str(extract_totals["movements_count"]), summary_value_style)],
                    [Paragraph("TOTAL MOVIMENTADO", meta_label_style), Paragraph(_format_currency(extract_totals["grand_total"]), summary_value_style)],
                    [Paragraph("", meta_value_style), Paragraph("", meta_value_style)],
                ],
            ],
            colWidths=[doc.width / 3] * 3,
            hAlign="LEFT",
        )
    elif export_tab == "entries":
        summary_table = Table(
            [
                [
                    [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                    [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
                    [Paragraph("ESCOPO", meta_label_style), Paragraph(item_type_label, meta_value_style)],
                ],
                [
                    [Paragraph("LISTAGEM", meta_label_style), Paragraph(listagem_label, meta_value_style)],
                    [Paragraph("MOVIMENTAÇÕES", meta_label_style), Paragraph(movement_label, meta_value_style)],
                    [Paragraph("INSUMO", meta_label_style), Paragraph("Filtrado" if selected_input_id else "Todos", meta_value_style)],
                ],
                [
                    [Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(purchase_entries)), summary_value_style)],
                    [Paragraph("TOTAL FINANCEIRO", meta_label_style), Paragraph(_format_currency(entries_total_sum), summary_value_style)],
                    [Paragraph("", meta_value_style), Paragraph("", meta_value_style)],
                ],
            ],
            colWidths=[doc.width / 3] * 3,
            hAlign="LEFT",
        )
    else:
        summary_table = Table(
            [
                [
                    [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                    [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
                    [Paragraph("ESCOPO", meta_label_style), Paragraph(item_type_label, meta_value_style)],
                ],
                [
                    [Paragraph("LISTAGEM", meta_label_style), Paragraph(listagem_label, meta_value_style)],
                    [Paragraph("MOVIMENTAÇÕES", meta_label_style), Paragraph(movement_label, meta_value_style)],
                    [Paragraph("INSUMO", meta_label_style), Paragraph("Filtrado" if selected_input_id else "Todos", meta_value_style)],
                ],
                [
                    [Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(stock_outputs)), summary_value_style)],
                    [Paragraph("TOTAL FINANCEIRO", meta_label_style), Paragraph(_format_currency(outputs_total_sum), summary_value_style)],
                    [Paragraph("", meta_value_style), Paragraph("", meta_value_style)],
                ],
            ],
            colWidths=[doc.width / 3] * 3,
            hAlign="LEFT",
        )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 14)]
    if export_tab == "extract":
        elements.extend(
            _pdf_flowables_extract_detail_table(
                doc,
                extract_rows,
                extract_totals,
                cell_style=cell_style,
                cell_muted_style=cell_muted_style,
                cell_numeric_style=cell_numeric_style,
                meta_value_style=meta_value_style,
                summary_value_style=summary_value_style,
            )
        )
    else:
        table_data = [["Mov.", "Data", "Insumo", "Tipo", "Origem / Fazenda", "Quantidade", "Valor", "Observações"]]
        report_rows = []
        if export_tab == "entries":
            for item in purchase_entries:
                report_rows.append({
                    "kind": "Entrada",
                    "date": item.purchase_date,
                    "name": item.input_catalog.name if item.input_catalog else item.name,
                    "type": _catalog_item_type_label_pt(item.input_catalog),
                    "origin": item.farm.name if item.farm else "Sem fazenda vinculada",
                    "quantity": f"{_format_decimal_br(item.total_quantity, 2)} {item.package_unit}",
                    "value": float(item.total_cost or 0),
                    "notes": item.notes or "-",
                    "sort_key": (item.purchase_date or today_in_app_timezone(), 0, item.id),
                })
        else:
            for output in stock_outputs:
                report_rows.append({
                    "kind": "Saída",
                    "date": output.movement_date,
                    "name": output.input_catalog.name if output.input_catalog else "Insumo removido",
                    "type": _catalog_item_type_label_pt(output.input_catalog),
                    "origin": f"{output.origin} • {output.farm.name if output.farm else 'Sem fazenda'}{(' / ' + output.plot.name) if output.plot else ''}",
                    "quantity": f"{_format_decimal_br(output.quantity, 2)} {output.unit}",
                    "value": float(output.total_cost or 0),
                    "notes": output.notes or "-",
                    "sort_key": (output.movement_date or today_in_app_timezone(), 1, output.id),
                })
        report_rows.sort(key=lambda row: row["sort_key"], reverse=True)
        for row in report_rows:
            movement_color = "#166534" if row["kind"] == "Entrada" else "#be123c"
            table_data.append([
                Paragraph(f'<font color="{movement_color}"><b>{row["kind"]}</b></font>', cell_style),
                Paragraph(row["date"].strftime("%d/%m/%Y") if row["date"] else "-", cell_style),
                Paragraph(row["name"], cell_style),
                Paragraph(row["type"], cell_muted_style),
                Paragraph(row["origin"], cell_muted_style),
                Paragraph(row["quantity"], cell_numeric_style),
                Paragraph(_format_currency(row["value"]), cell_numeric_style),
                Paragraph(row["notes"][:90], cell_muted_style),
            ])
        column_weights = [7, 8, 18, 12, 20, 11, 12, 20]
        weight_total = sum(column_weights)
        table_col_widths = [doc.width * (weight / weight_total) for weight in column_weights[:-1]]
        table_col_widths.append(doc.width - sum(table_col_widths))
        table = Table(table_data, repeatRows=1, colWidths=table_col_widths, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#36552a")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("ALIGN", (5, 1), (6, -1), "RIGHT"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))
        if export_tab == "entries":
            footer_line = Paragraph(f"<b>Total (entradas):</b> {_format_currency(entries_total_sum)}", summary_value_style)
        else:
            footer_line = Paragraph(f"<b>Total (saídas):</b> {_format_currency(outputs_total_sum)}", summary_value_style)
        footer_summary = Table([[footer_line]], colWidths=[doc.width], hAlign="LEFT")
        footer_summary.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef6ee")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cfe1d0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(footer_summary)

    generated_by = user.display_name or user.name or user.email
    generated_at_label = generated_at.strftime("%d/%m/%Y %H:%M")

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="extrato_estoque.pdf"'},
    )


@router.get("/insumos/patrimonio/exportar.xlsx")
def export_equipment_assets_xlsx(
    request: Request,
    farm_id: str | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_status = status if status in {"ativo", "em_manutencao", "baixado"} else None
    assets = _sort_collection_desc(
        _filter_equipment_assets_by_status(repo.list_equipment_assets(farm_id=selected_farm_id), selected_status),
        lambda item: item.acquisition_date,
        lambda item: item.id,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Patrimônio"
    headers = [
        "Nome",
        "Categoria",
        "Fazenda",
        "Fabricante",
        "Modelo",
        "Ano",
        "Identificação",
        "Data de aquisição",
        "Valor de aquisição",
        "Status",
        "Observações",
    ]
    sheet.append(headers)
    for asset in assets:
        sheet.append(
            [
                asset.name or "",
                asset.category or "",
                asset.farm.name if asset.farm else "",
                asset.manufacturer or "",
                asset.brand_model or "",
                asset.manufacture_year or "",
                asset.asset_code or "",
                asset.acquisition_date.isoformat() if asset.acquisition_date else "",
                float(asset.acquisition_value or 0),
                asset.status.replace("_", " ").title() if asset.status else "",
                asset.notes or "",
            ]
        )
    for index, width in enumerate([28, 18, 24, 20, 20, 10, 18, 15, 18, 18, 38], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="patrimonio_equipamentos.xlsx"'},
    )


@router.get("/insumos/patrimonio/exportar.pdf")
def export_equipment_assets_pdf(
    request: Request,
    farm_id: str | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_farm = repo.get_farm(selected_farm_id) if selected_farm_id else None
    selected_status = status if status in {"ativo", "em_manutencao", "baixado"} else None
    assets = _sort_collection_desc(
        _filter_equipment_assets_by_status(repo.list_equipment_assets(farm_id=selected_farm_id), selected_status),
        lambda item: item.acquisition_date,
        lambda item: item.id,
    )
    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    total_assets = len(assets)
    total_acquisition = round(sum(float(asset.acquisition_value or 0) for asset in assets), 2)
    active_assets = sum(1 for asset in assets if asset.status == "ativo")
    maintenance_assets = sum(1 for asset in assets if asset.status == "em_manutencao")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle(
        "AssetPdfFarmHeader",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#1e293b"),
    )
    meta_label_style = ParagraphStyle(
        "AssetPdfMetaLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#446a36"),
        spaceAfter=2,
    )
    meta_value_style = ParagraphStyle(
        "AssetPdfMetaValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#334155"),
    )
    cell_style = ParagraphStyle(
        "AssetPdfCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.3,
        leading=10.5,
        textColor=colors.HexColor("#0f172a"),
    )
    cell_muted_style = ParagraphStyle(
        "AssetPdfCellMuted",
        parent=cell_style,
        textColor=colors.HexColor("#475569"),
    )
    cell_numeric_style = ParagraphStyle(
        "AssetPdfCellNumeric",
        parent=cell_style,
        alignment=TA_RIGHT,
    )
    summary_value_style = ParagraphStyle(
        "AssetPdfSummaryValue",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#0f172a"),
    )

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    farm_name = selected_farm.name if selected_farm else "Fazenda Bela Vista"
    scope_label = selected_farm.name if selected_farm else "Todas as fazendas"

    header_table = Table(
        [[logo_flowable, Paragraph(farm_name, farm_header_style)]],
        colWidths=[76, doc.width - 76],
        hAlign="LEFT",
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    summary_table = Table(
        [
            [
                [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
                [Paragraph("ESCOPO", meta_label_style), Paragraph(scope_label, meta_value_style)],
                [Paragraph("TOTAL DE BENS", meta_label_style), Paragraph(str(total_assets), summary_value_style)],
            ],
            [
                [Paragraph("ATIVOS", meta_label_style), Paragraph(str(active_assets), summary_value_style)],
                [Paragraph("EM MANUTENÇÃO", meta_label_style), Paragraph(str(maintenance_assets), summary_value_style)],
                [Paragraph("VALOR TOTAL", meta_label_style), Paragraph(_format_currency(total_acquisition), summary_value_style)],
            ],
        ],
        colWidths=[doc.width / 3] * 3,
        hAlign="LEFT",
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 14)]
    data = [[
        "Nome",
        "Categoria / Fazenda",
        "Fabricante / Modelo",
        "Ano / ID",
        "Aquisição",
        "Valor",
        "Status",
        "Observações",
    ]]
    for asset in assets:
        status_color = {
            "ativo": "#166534",
            "em_manutencao": "#b45309",
            "baixado": "#be123c",
        }.get(asset.status or "", "#334155")
        category_farm = asset.category or "-"
        if asset.farm:
            category_farm = f"{category_farm} • {asset.farm.name}"
        maker_model = asset.manufacturer or "-"
        if asset.brand_model:
            maker_model = f"{maker_model} • {asset.brand_model}" if maker_model != "-" else asset.brand_model
        year_code = asset.manufacture_year or "-"
        if asset.asset_code:
            year_code = f"{year_code} • {asset.asset_code}" if year_code != "-" else asset.asset_code

        data.append([
            Paragraph(asset.name or "-", cell_style),
            Paragraph(category_farm, cell_muted_style),
            Paragraph(maker_model, cell_muted_style),
            Paragraph(year_code, cell_muted_style),
            Paragraph(asset.acquisition_date.strftime("%d/%m/%Y") if asset.acquisition_date else "-", cell_style),
            Paragraph(_format_currency(asset.acquisition_value or 0), cell_numeric_style),
            Paragraph(f'<font color="{status_color}"><b>{(asset.status or "-").replace("_", " ").title()}</b></font>', cell_style),
            Paragraph((asset.notes or "-")[:90], cell_muted_style),
        ])
    data.append([
        "",
        "",
        "",
        Paragraph("<b>Total geral</b>", cell_style),
        "",
        Paragraph(f"<b>{_format_currency(total_acquisition)}</b>", cell_numeric_style),
        "",
        "",
    ])

    column_weights = [16, 17, 17, 12, 9, 11, 8, 18]
    weight_total = sum(column_weights)
    table_col_widths = [doc.width * (weight / weight_total) for weight in column_weights[:-1]]
    table_col_widths.append(doc.width - sum(table_col_widths))
    table = Table(data, repeatRows=1, colWidths=table_col_widths, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#36552a")),
                ("LINEBELOW", (0, 1), (-1, -2), 0.35, colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f8fafc")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.2),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("ALIGN", (5, 1), (5, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 12))

    footer_col_widths = [
        sum(table_col_widths[:3]),
        sum(table_col_widths[3:5]),
        sum(table_col_widths[5:]),
    ]
    footer_summary = Table(
        [[
            Paragraph(f"<b>Total de bens:</b> {total_assets}", meta_value_style),
            Paragraph(f"<b>Ativos:</b> {active_assets} • <b>Em manutenção:</b> {maintenance_assets}", meta_value_style),
            Paragraph(f"<b>Valor total:</b> {_format_currency(total_acquisition)}", summary_value_style),
        ]],
        colWidths=footer_col_widths,
        hAlign="LEFT",
    )
    footer_summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef6ee")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cfe1d0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    elements.append(footer_summary)

    generated_at_label = generated_at.strftime("%d/%m/%Y %H:%M")

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="patrimonio_equipamentos.pdf"'},
    )


@router.get("/insumos/patrimonio")
def equipment_assets_page(
    request: Request,
    edit_id: int | None = None,
    farm_id: int | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    effective_farm_id = farm_id or _active_farm_id(request)
    finance_accounts = repo.list_finance_accounts(farm_id=effective_farm_id) if effective_farm_id else []
    selected_status = status if status in {"ativo", "em_manutencao", "baixado"} else None
    assets = _sort_collection_desc(
        _filter_equipment_assets_by_status(repo.list_equipment_assets(farm_id=effective_farm_id), selected_status),
        lambda item: item.acquisition_date,
        lambda item: item.id,
    )
    assets_pagination = _paginate_collection(request, assets, "assets_page", per_page=5)
    return templates.TemplateResponse(
        "equipment_assets.html",
        _base_context(
            request,
            user,
            csrf_token,
            "assets",
            _repo=repo,
            title="Patrimonio e Equipamentos",
            farms=repo.list_farms(),
            selected_farm_id=effective_farm_id,
            selected_asset_status=selected_status or "",
            assets_export_query=_assets_export_query(farm_id=effective_farm_id, status=selected_status),
            asset_category_options=EQUIPMENT_ASSET_CATEGORY_OPTIONS,
            finance_account_options=finance_accounts,
            finance_account_default=next((item for item in finance_accounts if item.is_default), None),
            current_year=today_in_app_timezone().year,
            assets=assets_pagination["items"],
            assets_pagination=assets_pagination,
            edit_asset=repo.get_equipment_asset(edit_id) if edit_id else None,
        ),
    )


@router.get("/insumos/suprimentos")
def supplies_page(
    request: Request,
    edit_id: int | None = None,
    edit_output_id: int | None = None,
    farm_id: int | None = None,
    input_id: int | None = None,
    category: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str = "all",
    supplies_tab: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    effective_farm_id = farm_id or scope["active_farm_id"]
    finance_accounts = repo.list_finance_accounts(farm_id=effective_farm_id) if effective_farm_id else []
    selected_input_id = input_id
    start = _date_or_none(start_date)
    end = _date_or_none(end_date)
    stock_context = _build_stock_context(
        repo,
        farm_id=effective_farm_id,
        input_id=selected_input_id,
        start_date=start,
        end_date=end,
        movement_type=movement_type,
        item_type="suprimento",
    )
    purchase_entries = _sort_collection_desc(
        stock_context["purchase_entries"],
        lambda item: item.purchase_date,
        lambda item: item.id,
    )
    stock_outputs = _sort_collection_desc(
        stock_context["stock_outputs"],
        lambda item: item.movement_date,
        lambda item: item.id,
    )
    extract_rows = _sort_collection_desc(
        stock_context["extract_rows"],
        lambda row: row.get("date"),
        lambda row: row.get("reference"),
    )
    selected_category = (category or "").strip()
    if selected_category:
        purchase_entries = [item for item in purchase_entries if (item.input_catalog.category if item.input_catalog and item.input_catalog.category else "Geral") == selected_category]
        stock_outputs = [item for item in stock_outputs if (item.input_catalog.category if item.input_catalog and item.input_catalog.category else "Geral") == selected_category]
        allowed_input_ids = {
            item.id
            for item in stock_context["catalog_inputs"]
            if ((item.category if getattr(item, "category", None) else "Geral") == selected_category)
        }
        extract_rows = [row for row in extract_rows if row.get("input_id") in allowed_input_ids]
        stock_context["catalog_inputs"] = [item for item in stock_context["catalog_inputs"] if item.id in allowed_input_ids]
        stock_context["stock_catalog_rows"] = [row for row in stock_context["stock_catalog_rows"] if row.get("id") in allowed_input_ids]
    purchase_entries_pagination = _paginate_collection(request, purchase_entries, "entries_page")
    stock_outputs_pagination = _paginate_collection(request, stock_outputs, "outputs_page")
    extract_rows_pagination = _paginate_collection(request, extract_rows, "extract_page")
    edit_supply = repo.get_purchased_input(edit_id) if edit_id else None
    edit_output = repo.get_stock_output(edit_output_id) if edit_output_id else None
    if edit_output and (
        edit_output.reference_type != "manual_stock_output"
        or not edit_output.input_catalog
        or edit_output.input_catalog.item_type != "suprimento"
    ):
        edit_output = None
    selected_supplies_tab = supplies_tab if supplies_tab in {"entries", "outputs", "extract"} else "entries"
    if edit_output:
        selected_supplies_tab = "outputs"
    supplies_edit_urls = {
        item.id: _url_with_query(request, edit_id=item.id)
        for item in purchase_entries_pagination["items"]
    }
    supplies_after_edit_close_url = _url_with_query(request, edit_id=None)
    output_edit_urls = {
        item.id: _url_with_query(request, edit_output_id=item.id)
        for item in stock_outputs_pagination["items"]
    }
    supplies_after_output_edit_close_url = _url_with_query(request, edit_output_id=None)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    return templates.TemplateResponse(
        "supplies.html",
        _base_context(
            request,
            user,
            csrf_token,
            "supplies",
            _repo=repo,
            title="Suprimentos",
            farms=repo.list_farms(),
            plots=repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids),
            selected_farm_id=effective_farm_id,
            selected_input_id=selected_input_id,
            selected_category=selected_category,
            selected_start_date=start_date,
            selected_end_date=end_date,
            selected_movement_type=movement_type,
            selected_supplies_tab=selected_supplies_tab,
            supplies_tab_urls={
                "entries": _url_with_query(request, supplies_tab="entries"),
                "outputs": _url_with_query(request, supplies_tab="outputs"),
                "extract": _url_with_query(request, supplies_tab="extract"),
            },
            supplies_export_query=_stock_export_query(
                farm_id=effective_farm_id,
                input_id=selected_input_id,
                start_date=start,
                end_date=end,
                movement_type=movement_type,
                item_type="suprimento",
                stock_tab=selected_supplies_tab,
            ),
            supply_category_options=SUPPLY_CATEGORY_OPTIONS,
            finance_account_options=finance_accounts,
            finance_account_default=next((item for item in finance_accounts if item.is_default), None),
            supplies=purchase_entries_pagination["items"],
            supplies_pagination=purchase_entries_pagination,
            purchase_entries=purchase_entries_pagination["items"],
            purchase_entries_pagination=purchase_entries_pagination,
            stock_outputs=stock_outputs_pagination["items"],
            stock_outputs_pagination=stock_outputs_pagination,
            extract_rows=extract_rows_pagination["items"],
            extract_rows_pagination=extract_rows_pagination,
            inputs_catalog=stock_context["catalog_inputs"],
            input_stock=stock_context["input_stock"],
            stock_catalog_rows=stock_context["stock_catalog_rows"],
            edit_supply=edit_supply,
            edit_output=edit_output,
            supplies_edit_urls=supplies_edit_urls,
            supplies_after_edit_close_url=supplies_after_edit_close_url,
            output_edit_urls=output_edit_urls,
            supplies_after_output_edit_close_url=supplies_after_output_edit_close_url,
            current_supplies_url=_url_with_query(request),
        ),
    )


@router.post("/insumos/suprimentos")
async def create_supply_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    quantity_purchased: float = Form(...),
    package_size: float = Form(...),
    package_unit: str = Form(...),
    category: str | None = Form(None),
    unit_price: float = Form(...),
    finance_account_id: str | None = Form(None),
    purchase_date: str | None = Form(None),
    low_stock_threshold: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/suprimentos")
    if denied:
        return denied
    target_base = f"/insumos/suprimentos?{urlencode({'farm_id': scope['active_farm_id']})}" if scope["active_farm_id"] else "/insumos/suprimentos"
    try:
        finance_account = _resolve_optional_finance_account(repo, active_farm=scope["active_farm"], account_id=finance_account_id)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(target_base)
    try:
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(target_base)
    item = create_purchased_input(
        repo,
        {
            "farm_id": scope["active_farm_id"],
            "item_type": "suprimento",
            "name": name,
            "quantity_purchased": quantity_purchased,
            "package_size": package_size,
            "package_unit": package_unit,
            "category": category,
            "unit_price": unit_price,
            "finance_account_id": finance_account.id if finance_account else None,
            "purchase_date": purchase_date,
            "low_stock_threshold": low_stock_threshold,
            "notes": notes,
        },
    )
    try:
        saved_attachments = _save_purchased_input_attachments(repo, item, attachment_payloads)
    except Exception:
        _flash(request, "error", "O suprimento foi salvo, mas nao foi possivel gravar os anexos agora.")
        return _redirect_with_query("/insumos/suprimentos", farm_id=scope["active_farm_id"], edit_id=item.id)
    if saved_attachments:
        _flash(request, "success", f"Suprimento cadastrado com sucesso. {saved_attachments} anexo(s) salvo(s).")
        return _redirect_with_query("/insumos/suprimentos", farm_id=scope["active_farm_id"], edit_id=item.id)
    _flash(request, "success", "Suprimento cadastrado com sucesso.")
    return _redirect(target_base)


@router.post("/insumos/suprimentos/{input_id}/editar")
async def update_supply_action(
    input_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    name: str = Form(...),
    quantity_purchased: float = Form(...),
    package_size: float = Form(...),
    package_unit: str = Form(...),
    category: str | None = Form(None),
    unit_price: float = Form(...),
    finance_account_id: str | None = Form(None),
    purchase_date: str | None = Form(None),
    low_stock_threshold: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/insumos/suprimentos"
    scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/suprimentos")
    if denied:
        return denied
    item = repo.get_purchased_input(input_id)
    if not item:
        _flash(request, "error", "Suprimento nao encontrado.")
        return _redirect(target_url)
    if not _farm_matches_scope(item.farm_id, scope):
        _flash(request, "error", "Este lancamento nao pertence ao contexto ativo.")
        return _redirect(target_url)
    try:
        finance_account = _resolve_optional_finance_account(repo, active_farm=scope["active_farm"], account_id=finance_account_id)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={input_id}")
    try:
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={input_id}")
    update_purchased_input(
        repo,
        item,
        {
            "farm_id": scope["active_farm_id"],
            "item_type": "suprimento",
            "name": name,
            "quantity_purchased": quantity_purchased,
            "package_size": package_size,
            "package_unit": package_unit,
            "category": category,
            "unit_price": unit_price,
            "finance_account_id": finance_account.id if finance_account else None,
            "purchase_date": purchase_date,
            "low_stock_threshold": low_stock_threshold,
            "notes": notes,
        },
    )
    try:
        saved_attachments = _save_purchased_input_attachments(repo, item, attachment_payloads)
    except Exception:
        _flash(request, "error", "As alteracoes foram salvas, mas nao foi possivel incluir os novos anexos.")
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={input_id}")
    if saved_attachments:
        _flash(request, "success", f"Alteracoes salvas com sucesso. {saved_attachments} novo(s) anexo(s) adicionado(s).")
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={input_id}")
    _flash(request, "success", "Suprimento atualizado com sucesso.")
    return _redirect(target_url)


@router.post("/insumos/suprimentos/{input_id}/excluir")
def delete_supply_action(
    input_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/insumos/suprimentos"
    item = repo.get_purchased_input(input_id)
    if not item:
        _flash(request, "error", "Suprimento nao encontrado.")
        return _redirect(target_url)
    repo.delete(item)
    _flash(request, "success", "Suprimento excluido com sucesso.")
    return _redirect(target_url)


@router.get("/insumos/suprimentos/anexos/{attachment_id}")
def open_supply_attachment(
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    attachment = repo.get_purchased_input_attachment(attachment_id)
    if not attachment or not attachment.purchased_input:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect("/insumos/suprimentos")
    scope = _global_scope_context(request, repo)
    item = attachment.purchased_input
    if not _farm_matches_scope(item.farm_id, scope):
        _flash(request, "error", "Este anexo nao pertence ao contexto ativo.")
        return _redirect("/insumos/suprimentos")
    return _attachment_response(attachment.filename, attachment.content_type, attachment.file_data)


@router.post("/insumos/suprimentos/{input_id}/anexos/{attachment_id}/excluir")
def delete_supply_attachment_action(
    input_id: int,
    attachment_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/insumos/suprimentos"
    item = repo.get_purchased_input(input_id)
    if not item:
        _flash(request, "error", "Suprimento nao encontrado.")
        return _redirect(target_url)
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(item.farm_id, scope):
        _flash(request, "error", "Este lancamento nao pertence ao contexto ativo.")
        return _redirect(target_url)
    attachment = repo.get_purchased_input_attachment(attachment_id)
    if not attachment or attachment.purchased_input_id != item.id:
        _flash(request, "error", "Anexo nao encontrado.")
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={input_id}")
    repo.delete(attachment)
    _flash(request, "success", "Anexo removido com sucesso.")
    separator = "&" if "?" in target_url else "?"
    return _redirect(f"{target_url}{separator}edit_id={input_id}")


@router.post("/insumos/suprimentos/saida-manual")
def create_supply_output_action(
    request: Request,
    csrf_token: str = Form(...),
    plot_id: str | None = Form(None),
    input_id: int = Form(...),
    movement_date: str | None = Form(None),
    quantity: float = Form(...),
    unit: str = Form(...),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, scope, denied = _resolve_optional_plot_in_scope(
        request,
        repo,
        _int_or_none(plot_id),
        "/insumos/suprimentos",
    )
    if denied:
        return denied
    try:
        create_manual_stock_output(
            repo,
            {
                "farm_id": scope["active_farm_id"],
                "plot_id": plot.id if plot else None,
                "input_id": input_id,
                "season_id": scope["active_season_id"],
                "movement_date": movement_date,
                "quantity": quantity,
                "unit": unit,
                "notes": notes,
            },
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_with_query("/insumos/suprimentos", supplies_tab="outputs")
    _flash(request, "success", "Saida manual registrada com sucesso.")
    return _redirect_with_query("/insumos/suprimentos", supplies_tab="outputs")


@router.get("/insumos/suprimentos/saidas/{output_id}/editar")
def edit_supply_output_entry(
    output_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    output = repo.get_stock_output(output_id)
    if not output:
        _flash(request, "error", "Lancamento de saida nao encontrado.")
        return _redirect("/insumos/suprimentos")
    if output.reference_type != "manual_stock_output" or not output.input_catalog or output.input_catalog.item_type != "suprimento":
        _flash(request, "error", "Este lancamento nao possui edicao integrada em Suprimentos.")
        return _redirect("/insumos/suprimentos")
    return _redirect_with_query("/insumos/suprimentos", supplies_tab="outputs", edit_output_id=output_id)


@router.post("/insumos/suprimentos/saidas/{output_id}/editar")
async def update_supply_output_entry_action(
    output_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    csrf_token = str(form.get("csrf_token") or "")
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    output = repo.get_stock_output(output_id)
    redirect_to = str(form.get("redirect_to") or "")
    fallback_url = "/insumos/suprimentos?supplies_tab=outputs"
    target_url = redirect_to if redirect_to.startswith("/") else fallback_url
    if not output:
        _flash(request, "error", "Lancamento de saida nao encontrado.")
        return _redirect(target_url)
    if output.reference_type != "manual_stock_output" or not output.input_catalog or output.input_catalog.item_type != "suprimento":
        _flash(request, "error", "Este lancamento nao pode ser editado por aqui.")
        return _redirect(target_url)
    try:
        scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/suprimentos")
        if denied:
            return denied
        if not _farm_matches_scope(output.farm_id, scope) or (output.plot_id and not _plot_matches_scope(output.plot, scope)):
            _flash(request, "error", "Este lancamento de saida nao pertence ao contexto ativo.")
            return _redirect(target_url)
        plot, _, invalid_plot = _resolve_optional_plot_in_scope(
            request,
            repo,
            _int_or_none(form.get("plot_id")),
            "/insumos/suprimentos",
        )
        if invalid_plot:
            return invalid_plot
        update_manual_stock_output(
            repo,
            output,
            {
                "farm_id": scope["active_farm_id"],
                "plot_id": plot.id if plot else None,
                "movement_date": str(form.get("movement_date") or ""),
                "quantity": float(form.get("quantity") or 0),
                "unit": str(form.get("unit") or ""),
                "notes": str(form.get("notes") or "") or None,
            },
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_output_id={output_id}")
    _flash(request, "success", "Saida atualizada com sucesso.")
    return _redirect(target_url)


@router.post("/insumos/suprimentos/saidas/{output_id}/excluir")
def delete_supply_output_entry_action(
    output_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    fallback_url = "/insumos/suprimentos?supplies_tab=outputs"
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else fallback_url
    output = repo.get_stock_output(output_id)
    if not output:
        _flash(request, "error", "Lancamento de saida nao encontrado.")
        return _redirect(target_url)
    if output.reference_type != "manual_stock_output" or not output.input_catalog or output.input_catalog.item_type != "suprimento":
        _flash(request, "error", "Este lancamento esta vinculado a outro modulo e nao pode ser excluido por aqui.")
        return _redirect(target_url)
    try:
        delete_manual_stock_output(repo, output)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(target_url)
    _flash(request, "success", "Saida manual excluida com sucesso.")
    return _redirect(target_url)


@router.get("/insumos/suprimentos/exportar.xlsx")
def export_supplies_extract_xlsx(
    request: Request,
    farm_id: str | None = None,
    input_id: str | None = None,
    category: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str = "all",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_input_id = _int_or_none(input_id)
    export_tab = _export_stock_tab_param(request, movement_type)
    stock_context = _build_stock_context(
        repo,
        farm_id=selected_farm_id,
        input_id=selected_input_id,
        start_date=_date_or_none(start_date),
        end_date=_date_or_none(end_date),
        movement_type=movement_type,
        item_type="suprimento",
    )
    purchase_entries = _sort_collection_desc(stock_context["purchase_entries"], lambda item: item.purchase_date, lambda item: item.id)
    stock_outputs = _sort_collection_desc(stock_context["stock_outputs"], lambda item: item.movement_date, lambda item: item.id)
    extract_rows = _sort_collection_desc(stock_context["extract_rows"], lambda row: row.get("date"), lambda row: row.get("reference"))
    selected_category = (category or "").strip()
    if selected_category:
        purchase_entries = [item for item in purchase_entries if (item.input_catalog.category if item.input_catalog and item.input_catalog.category else "Geral") == selected_category]
        stock_outputs = [item for item in stock_outputs if (item.input_catalog.category if item.input_catalog and item.input_catalog.category else "Geral") == selected_category]
        allowed_input_ids = {item.id for item in stock_context["catalog_inputs"] if ((item.category if getattr(item, "category", None) else "Geral") == selected_category)}
        extract_rows = [row for row in extract_rows if row.get("input_id") in allowed_input_ids]
    workbook = Workbook()
    sheet = workbook.active
    if export_tab == "entries":
        sheet.title = "Entradas"
        _xlsx_write_purchase_entries(sheet, purchase_entries)
    elif export_tab == "outputs":
        sheet.title = "Saidas"
        _xlsx_write_stock_outputs(sheet, stock_outputs)
    else:
        sheet.title = "Extrato de suprimentos"
        _xlsx_write_extract_rows(sheet, extract_rows)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="extrato_suprimentos.xlsx"'},
    )


@router.get("/insumos/suprimentos/exportar.pdf")
def export_supplies_extract_pdf(
    request: Request,
    farm_id: str | None = None,
    input_id: str | None = None,
    category: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str = "all",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    selected_farm_id = _int_or_none(farm_id) or _active_farm_id(request)
    selected_input_id = _int_or_none(input_id)
    export_tab = _export_stock_tab_param(request, movement_type)
    stock_context = _build_stock_context(
        repo,
        farm_id=selected_farm_id,
        input_id=selected_input_id,
        start_date=_date_or_none(start_date),
        end_date=_date_or_none(end_date),
        movement_type=movement_type,
        item_type="suprimento",
    )
    purchase_entries = _sort_collection_desc(stock_context["purchase_entries"], lambda item: item.purchase_date, lambda item: item.id)
    stock_outputs = _sort_collection_desc(stock_context["stock_outputs"], lambda item: item.movement_date, lambda item: item.id)
    extract_rows = _sort_collection_desc(stock_context["extract_rows"], lambda row: row.get("date"), lambda row: row.get("reference"))
    selected_category = (category or "").strip()
    if selected_category:
        purchase_entries = [item for item in purchase_entries if (item.input_catalog.category if item.input_catalog and item.input_catalog.category else "Geral") == selected_category]
        stock_outputs = [item for item in stock_outputs if (item.input_catalog.category if item.input_catalog and item.input_catalog.category else "Geral") == selected_category]
        allowed_input_ids = {item.id for item in stock_context["catalog_inputs"] if ((item.category if getattr(item, "category", None) else "Geral") == selected_category)}
        extract_rows = [row for row in extract_rows if row.get("input_id") in allowed_input_ids]
    entries_total_sum = sum(float(item.total_cost or 0) for item in purchase_entries)
    outputs_total_sum = sum(float(output.total_cost or 0) for output in stock_outputs)
    extract_totals = _stock_report_totals(extract_rows)
    selected_farm = repo.get_farm(selected_farm_id) if selected_farm_id else None
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle("SupplyPdfFarmHeader", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, alignment=TA_RIGHT, textColor=colors.HexColor("#1e293b"))
    meta_label_style = ParagraphStyle("SupplyPdfMetaLabel", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#446a36"), spaceAfter=2)
    meta_value_style = ParagraphStyle("SupplyPdfMetaValue", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13, textColor=colors.HexColor("#334155"))
    cell_style = ParagraphStyle("SupplyPdfCell", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.3, leading=10.5, textColor=colors.HexColor("#0f172a"))
    cell_muted_style = ParagraphStyle("SupplyPdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    cell_numeric_style = ParagraphStyle("SupplyPdfCellNumeric", parent=cell_style, alignment=TA_RIGHT)
    summary_value_style = ParagraphStyle("SupplyPdfSummaryValue", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#0f172a"))
    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    farm_name = selected_farm.name if selected_farm else "Fazenda Bela Vista"
    movement_label = {"entrada": "Somente entradas", "saida": "Somente saídas", "all": "Entradas, saídas e extrato consolidado"}.get(movement_type, "Entradas, saídas e extrato consolidado")
    period_label = "Período completo"
    if start_date or end_date:
        period_label = f"{start_date or 'Início'} até {end_date or 'Hoje'}"
    listagem_label = {"entries": "Entradas", "outputs": "Saídas", "extract": "Extrato"}.get(export_tab, "Extrato")
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
    if export_tab == "extract":
        summary_table = Table([
            [[Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)], [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)], [Paragraph("ESCOPO", meta_label_style), Paragraph("Suprimentos", meta_value_style)]],
            [[Paragraph("LISTAGEM", meta_label_style), Paragraph(listagem_label, meta_value_style)], [Paragraph("MOVIMENTAÇÕES", meta_label_style), Paragraph(movement_label, meta_value_style)], [Paragraph("CATEGORIA", meta_label_style), Paragraph(selected_category or "Todas", meta_value_style)]],
            [[Paragraph("LANÇAMENTOS", meta_label_style), Paragraph(str(extract_totals["movements_count"]), summary_value_style)], [Paragraph("TOTAL MOVIMENTADO", meta_label_style), Paragraph(_format_currency(extract_totals["grand_total"]), summary_value_style)], [Paragraph("", meta_value_style), Paragraph("", meta_value_style)]],
        ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    elif export_tab == "entries":
        summary_table = Table([
            [[Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)], [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)], [Paragraph("ESCOPO", meta_label_style), Paragraph("Suprimentos", meta_value_style)]],
            [[Paragraph("LISTAGEM", meta_label_style), Paragraph(listagem_label, meta_value_style)], [Paragraph("MOVIMENTAÇÕES", meta_label_style), Paragraph(movement_label, meta_value_style)], [Paragraph("CATEGORIA", meta_label_style), Paragraph(selected_category or "Todas", meta_value_style)]],
            [[Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(purchase_entries)), summary_value_style)], [Paragraph("TOTAL FINANCEIRO", meta_label_style), Paragraph(_format_currency(entries_total_sum), summary_value_style)], [Paragraph("", meta_value_style), Paragraph("", meta_value_style)]],
        ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    else:
        summary_table = Table([
            [[Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)], [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)], [Paragraph("ESCOPO", meta_label_style), Paragraph("Suprimentos", meta_value_style)]],
            [[Paragraph("LISTAGEM", meta_label_style), Paragraph(listagem_label, meta_value_style)], [Paragraph("MOVIMENTAÇÕES", meta_label_style), Paragraph(movement_label, meta_value_style)], [Paragraph("CATEGORIA", meta_label_style), Paragraph(selected_category or "Todas", meta_value_style)]],
            [[Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(stock_outputs)), summary_value_style)], [Paragraph("TOTAL FINANCEIRO", meta_label_style), Paragraph(_format_currency(outputs_total_sum), summary_value_style)], [Paragraph("", meta_value_style), Paragraph("", meta_value_style)]],
        ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    summary_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")), ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")), ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")), ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 14)]
    if export_tab == "extract":
        elements.extend(_pdf_flowables_extract_detail_table(doc, extract_rows, extract_totals, cell_style=cell_style, cell_muted_style=cell_muted_style, cell_numeric_style=cell_numeric_style, meta_value_style=meta_value_style, summary_value_style=summary_value_style))
    else:
        table_data = [["Mov.", "Data", "Origem / Fazenda", "Produto", "Quantidade", "Valor", "Observações"]]
        report_rows = []
        if export_tab == "entries":
            for item in purchase_entries:
                report_rows.append({
                    "kind": "Entrada",
                    "date": item.purchase_date,
                    "name": item.input_catalog.name if item.input_catalog else item.name,
                    "origin": item.farm.name if item.farm else "Sem fazenda vinculada",
                    "quantity": f"{_format_quantity_br(item.total_quantity, item.package_unit)} {item.package_unit}",
                    "value": float(item.total_cost or 0),
                    "notes": item.notes or "-",
                    "sort_key": (item.purchase_date or today_in_app_timezone(), 0, item.id),
                })
        else:
            for output in stock_outputs:
                report_rows.append({
                    "kind": "Saída",
                    "date": output.movement_date,
                    "name": output.input_catalog.name if output.input_catalog else "Suprimento removido",
                    "origin": f"{output.origin} • {output.farm.name if output.farm else 'Sem fazenda'}{(' / ' + output.plot.name) if output.plot else ''}",
                    "quantity": f"{_format_quantity_br(output.quantity, output.unit)} {output.unit}",
                    "value": float(output.total_cost or 0),
                    "notes": output.notes or "-",
                    "sort_key": (output.movement_date or today_in_app_timezone(), 1, output.id),
                })
        report_rows.sort(key=lambda row: row["sort_key"], reverse=True)
        for row in report_rows:
            movement_color = "#166534" if row["kind"] == "Entrada" else "#be123c"
            table_data.append([
                Paragraph(f'<font color="{movement_color}"><b>{row["kind"]}</b></font>', cell_style),
                Paragraph(row["date"].strftime("%d/%m/%Y") if row["date"] else "-", cell_style),
                Paragraph(row["origin"], cell_muted_style),
                Paragraph(row["name"], cell_style),
                Paragraph(row["quantity"], cell_numeric_style),
                Paragraph(_format_currency(row["value"]), cell_numeric_style),
                Paragraph(row["notes"][:90], cell_muted_style),
            ])
        column_weights = [8, 9, 24, 22, 13, 12, 18]
        weight_total = sum(column_weights)
        table_col_widths = [doc.width * (weight / weight_total) for weight in column_weights[:-1]]
        table_col_widths.append(doc.width - sum(table_col_widths))
        table = Table(table_data, repeatRows=1, colWidths=table_col_widths, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.2),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
            ("TOPPADDING", (0, 0), (-1, 0), 9),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ]))
        elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": 'attachment; filename="extrato_suprimentos.pdf"'})


@router.post("/insumos/patrimonio")
async def create_equipment_asset_action(
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str | None = Form(None),
    name: str = Form(...),
    category: str = Form(...),
    manufacturer: str | None = Form(None),
    manufacture_year: str | None = Form(None),
    brand_model: str | None = Form(None),
    asset_code: str | None = Form(None),
    acquisition_date: str | None = Form(None),
    acquisition_value: str | None = Form(None),
    finance_account_id: str | None = Form(None),
    status_value: str = Form("ativo", alias="status"),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/patrimonio")
    if denied:
        return denied
    try:
        finance_account = _resolve_optional_finance_account(repo, active_farm=scope["active_farm"], account_id=finance_account_id)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect("/insumos/patrimonio")
    normalized_category = _clean_text(category)
    if normalized_category not in EQUIPMENT_ASSET_CATEGORY_OPTIONS:
        _flash(request, "error", "Selecione uma categoria valida para o patrimonio.")
        return _redirect("/insumos/patrimonio")
    try:
        normalized_manufacture_year = _parse_equipment_asset_manufacture_year(manufacture_year)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect("/insumos/patrimonio")
    try:
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect("/insumos/patrimonio")
    asset = create_equipment_asset(
        repo,
        {
            "farm_id": scope["active_farm_id"],
            "name": name,
            "category": normalized_category,
            "manufacturer": _clean_text(manufacturer),
            "manufacture_year": normalized_manufacture_year,
            "brand_model": _clean_text(brand_model),
            "asset_code": _clean_text(asset_code),
            "acquisition_date": acquisition_date,
            "acquisition_value": _float_or_none(acquisition_value),
            "finance_account_id": finance_account.id if finance_account else None,
            "status": status_value,
            "notes": notes,
        },
    )
    try:
        saved_attachments = _save_equipment_asset_attachments(repo, asset, attachment_payloads)
    except Exception:
        _flash(request, "error", "O patrimonio foi salvo, mas nao foi possivel gravar os anexos agora.")
        return _redirect_with_query("/insumos/patrimonio", edit_id=asset.id)
    if saved_attachments:
        _flash(request, "success", f"Patrimonio cadastrado com sucesso. {saved_attachments} anexo(s) salvo(s).")
        return _redirect_with_query("/insumos/patrimonio", edit_id=asset.id)
    _flash(request, "success", "Patrimonio cadastrado com sucesso.")
    return _redirect("/insumos/patrimonio")


@router.post("/insumos/patrimonio/{asset_id}/editar")
async def update_equipment_asset_action(
    asset_id: int,
    request: Request,
    csrf_token: str = Form(...),
    farm_id: str | None = Form(None),
    name: str = Form(...),
    category: str = Form(...),
    manufacturer: str | None = Form(None),
    manufacture_year: str | None = Form(None),
    brand_model: str | None = Form(None),
    asset_code: str | None = Form(None),
    acquisition_date: str | None = Form(None),
    acquisition_value: str | None = Form(None),
    finance_account_id: str | None = Form(None),
    status_value: str = Form("ativo", alias="status"),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/insumos/patrimonio")
    if denied:
        return denied
    asset = repo.get_equipment_asset(asset_id)
    if not asset:
        _flash(request, "error", "Patrimonio nao encontrado.")
        return _redirect_for_request(request, "/insumos/patrimonio")
    if not _farm_matches_scope(asset.farm_id, scope):
        _flash(request, "error", "Este patrimonio nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/insumos/patrimonio")
    try:
        finance_account = _resolve_optional_finance_account(repo, active_farm=scope["active_farm"], account_id=finance_account_id)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    normalized_category = _clean_text(category)
    if normalized_category not in EQUIPMENT_ASSET_CATEGORY_OPTIONS and normalized_category != asset.category:
        _flash(request, "error", "Selecione uma categoria valida para o patrimonio.")
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    try:
        normalized_manufacture_year = _parse_equipment_asset_manufacture_year(manufacture_year)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    try:
        attachment_payloads = _read_attachments(await _request_attachments(request))
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    update_equipment_asset(
        repo,
        asset,
        {
            "farm_id": scope["active_farm_id"],
            "name": name,
            "category": normalized_category,
            "manufacturer": _clean_text(manufacturer),
            "manufacture_year": normalized_manufacture_year,
            "brand_model": _clean_text(brand_model),
            "asset_code": _clean_text(asset_code),
            "acquisition_date": acquisition_date,
            "acquisition_value": _float_or_none(acquisition_value),
            "finance_account_id": finance_account.id if finance_account else None,
            "status": status_value,
            "notes": notes,
        },
    )
    try:
        saved_attachments = _save_equipment_asset_attachments(repo, asset, attachment_payloads)
    except Exception:
        _flash(request, "error", "As alteracoes foram salvas, mas nao foi possivel incluir os novos anexos.")
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    if saved_attachments:
        _flash(request, "success", f"Alteracoes salvas com sucesso. {saved_attachments} novo(s) anexo(s) adicionado(s).")
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    _flash(request, "success", "Patrimonio atualizado com sucesso.")
    return _redirect_for_request(request, "/insumos/patrimonio")


@router.post("/insumos/patrimonio/{asset_id}/excluir")
def delete_equipment_asset_action(
    asset_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    asset = repo.get_equipment_asset(asset_id)
    if not asset:
        _flash(request, "error", "Patrimonio nao encontrado.")
        return _redirect_for_request(request, "/insumos/patrimonio")
    repo.delete(asset)
    _flash(request, "success", "Patrimonio excluido com sucesso.")
    return _redirect("/insumos/patrimonio")


@router.get("/insumos/patrimonio/anexos/{attachment_id}")
def open_equipment_asset_attachment(
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    attachment = repo.get_equipment_asset_attachment(attachment_id)
    if not attachment or not attachment.equipment_asset:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect("/insumos/patrimonio")
    scope = _global_scope_context(request, repo)
    asset = attachment.equipment_asset
    if not _farm_matches_scope(asset.farm_id, scope):
        _flash(request, "error", "Este anexo nao pertence ao contexto ativo.")
        return _redirect("/insumos/patrimonio")
    return _attachment_response(attachment.filename, attachment.content_type, attachment.file_data)


@router.post("/insumos/patrimonio/{asset_id}/anexos/{attachment_id}/excluir")
def delete_equipment_asset_attachment_action(
    asset_id: int,
    attachment_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    asset = repo.get_equipment_asset(asset_id)
    if not asset:
        _flash(request, "error", "Patrimonio nao encontrado.")
        return _redirect("/insumos/patrimonio")
    scope = _global_scope_context(request, repo)
    if not _farm_matches_scope(asset.farm_id, scope):
        _flash(request, "error", "Este patrimonio nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/insumos/patrimonio")
    attachment = repo.get_equipment_asset_attachment(attachment_id)
    if not attachment or attachment.equipment_asset_id != asset.id:
        _flash(request, "error", "Anexo nao encontrado.")
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    repo.delete(attachment)
    if repo.get_equipment_asset_attachment(attachment_id):
        _flash(request, "error", "Nao foi possivel remover o anexo agora. Tente novamente.")
        return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)
    _flash(request, "success", "Anexo removido com sucesso.")
    return _redirect_for_request(request, "/insumos/patrimonio", edit_id=asset_id)


@router.get("/insumos/recomendacao")
def input_recommendations_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    catalog_inputs = repo.list_input_catalog(item_type="insumo_agricola")
    purchase_entries = repo.list_purchased_inputs()
    input_stock = {
        item.id: {
            "available": round(
                sum(float(entry.available_quantity or 0) for entry in purchase_entries if entry.input_id == item.id),
                2,
            ),
            "unit": item.default_unit,
        }
        for item in catalog_inputs
    }
    recommendations = [
        recommendation
        for recommendation in repo.list_input_recommendations()
        if (
            (not scope["active_farm_id"] or recommendation.farm_id == scope["active_farm_id"])
            and (not plot_ids or recommendation.plot_id is None or recommendation.plot_id in plot_ids)
        )
    ]
    recommendations = _sort_collection_desc(
        recommendations,
        lambda item: item.id,
    )
    recommendations_pagination = _paginate_collection(request, recommendations, "recommendations_page")
    return templates.TemplateResponse(
        "input_recommendations.html",
        _base_context(
            request,
            user,
            csrf_token,
            "input_recommendations",
            _repo=repo,
            title="Recomendacao de Insumos",
            farms=repo.list_farms(),
            plots=plots,
            inputs_catalog=catalog_inputs,
            input_stock=input_stock,
            recommendations=recommendations_pagination["items"],
            recommendations_pagination=recommendations_pagination,
            edit_recommendation=repo.get_input_recommendation(edit_id) if edit_id else None,
        ),
    )


@router.post("/insumos/recomendacao")
async def create_input_recommendation_action(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    validate_csrf(request, str(form.get("csrf_token") or ""))
    application_name = str(form.get("application_name") or "").strip()
    items = _parse_recommendation_items(form)
    if not application_name or not items:
        _flash(request, "error", "Informe a aplicacao e adicione ao menos um insumo.")
        return _redirect("/insumos/recomendacao")
    repo = _repository(db)
    plot, scope, denied = _resolve_optional_plot_in_scope(
        request,
        repo,
        _int_or_none(form.get("plot_id")),
        "/insumos/recomendacao",
    )
    if denied:
        return denied
    notes = str(form.get("notes") or "") or None
    create_input_recommendation(
        repo,
        {
            "farm_id": scope["active_farm_id"],
            "plot_id": plot.id if plot else None,
            "application_name": application_name,
            "items": items,
            "notes": notes,
        },
    )
    _flash(request, "success", "Recomendacao cadastrada com sucesso.")
    return _redirect("/insumos/recomendacao")


@router.post("/insumos/recomendacao/{recommendation_id}/editar")
async def update_input_recommendation_action(
    recommendation_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    validate_csrf(request, str(form.get("csrf_token") or ""))
    repo = _repository(db)
    plot, scope, denied = _resolve_optional_plot_in_scope(
        request,
        repo,
        _int_or_none(form.get("plot_id")),
        "/insumos/recomendacao",
    )
    if denied:
        return denied
    recommendation = repo.get_input_recommendation(recommendation_id)
    if not recommendation:
        _flash(request, "error", "Recomendacao nao encontrada.")
        return _redirect_for_request(request, "/insumos/recomendacao")
    if not _farm_matches_scope(recommendation.farm_id, scope):
        _flash(request, "error", "Esta recomendacao nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/insumos/recomendacao")
    if recommendation.plot_id and not _plot_matches_scope(recommendation.plot, scope):
        _flash(request, "error", "Esta recomendacao nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/insumos/recomendacao")
    application_name = str(form.get("application_name") or "").strip()
    items = _parse_recommendation_items(form)
    if not application_name or not items:
        _flash(request, "error", "Informe a aplicacao e adicione ao menos um insumo.")
        return _redirect_for_request(request, "/insumos/recomendacao", edit_id=recommendation_id)
    update_input_recommendation(
        repo,
        recommendation,
        {
            "farm_id": scope["active_farm_id"],
            "plot_id": plot.id if plot else None,
            "application_name": application_name,
            "items": items,
            "notes": str(form.get("notes") or "") or None,
        },
    )
    _flash(request, "success", "Recomendacao atualizada com sucesso.")
    return _redirect_for_request(request, "/insumos/recomendacao")


@router.post("/insumos/recomendacao/{recommendation_id}/excluir")
def delete_input_recommendation_action(
    recommendation_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    recommendation = repo.get_input_recommendation(recommendation_id)
    if not recommendation:
        _flash(request, "error", "Recomendacao nao encontrada.")
        return _redirect("/insumos/recomendacao")
    repo.delete(recommendation)
    _flash(request, "success", "Recomendacao excluida com sucesso.")
    return _redirect("/insumos/recomendacao")


@router.get("/variedades")
def varieties_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    return templates.TemplateResponse(
        "varieties.html",
        _base_context(
            request,
            user,
            csrf_token,
            "varieties",
            _repo=repo,
            varieties=repo.list_varieties(),
            edit_variety=repo.get_variety(edit_id) if edit_id else None,
        ),
    )


@router.post("/variedades")
def create_variety_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    species: str = Form(...),
    maturation_cycle: str = Form(...),
    flavor_profile: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    create_variety(
        _repository(db),
        {
            "name": name,
            "species": species,
            "maturation_cycle": maturation_cycle,
            "flavor_profile": flavor_profile,
            "notes": notes,
        },
    )
    _flash(request, "success", "Variedade cadastrada com sucesso.")
    return _redirect("/variedades")


@router.post("/variedades/{variety_id}/editar")
def update_variety_action(
    variety_id: int,
    request: Request,
    csrf_token: str = Form(...),
    name: str = Form(...),
    species: str = Form(...),
    maturation_cycle: str = Form(...),
    flavor_profile: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    variety = repo.get_variety(variety_id)
    if not variety:
        _flash(request, "error", "Variedade nao encontrada.")
        return _redirect("/variedades")
    update_variety(
        repo,
        variety,
        {
            "name": name,
            "species": species,
            "maturation_cycle": maturation_cycle,
            "flavor_profile": flavor_profile,
            "notes": notes,
        },
    )
    _flash(request, "success", "Variedade atualizada com sucesso.")
    return _redirect("/variedades")


@router.post("/variedades/{variety_id}/excluir")
def delete_variety_action(
    variety_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    variety = repo.get_variety(variety_id)
    if not variety:
        _flash(request, "error", "Variedade nao encontrada.")
        return _redirect("/variedades")
    if variety.plots:
        _flash(request, "error", "Nao e possivel excluir a variedade enquanto houver setores vinculados.")
        return _redirect("/variedades")
    repo.delete(variety)
    _flash(request, "success", "Variedade excluida com sucesso.")
    return _redirect("/variedades")


@router.get("/safras")
def crop_seasons_page(
    request: Request,
    edit_id: int | None = None,
    view_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    effective_farm_id = scope["active_farm_id"]
    crop_seasons = repo.list_crop_seasons(farm_id=effective_farm_id)
    fertilizations = repo.list_fertilizations()
    stock_outputs = repo.list_stock_outputs(farm_id=effective_farm_id) if effective_farm_id else repo.list_stock_outputs()
    edit_season = repo.get_crop_season(edit_id) if edit_id else None
    if edit_season and not _farm_matches_scope(edit_season.farm_id, scope):
        _flash(request, "error", "Esta safra nao pertence ao contexto ativo.")
        edit_season = None
    view_season = None
    if not edit_season and view_id:
        view_season = repo.get_crop_season(view_id)
        if view_season and not _farm_matches_scope(view_season.farm_id, scope):
            _flash(request, "error", "Esta safra nao pertence ao contexto ativo.")
            view_season = None
    season_costs: dict[int, dict] = {}
    for season in crop_seasons:
        season_outputs = [output for output in stock_outputs if output.season_id == season.id]
        season_fertilizations = [record for record in fertilizations if record.season_id == season.id]
        input_cost = round(
            sum(
                float(output.total_cost or 0)
                for output in season_outputs
                if output.input_catalog and output.input_catalog.item_type != "combustivel"
            ),
            2,
        )
        fuel_cost = round(
            sum(
                float(output.total_cost or 0)
                for output in season_outputs
                if output.input_catalog and output.input_catalog.item_type == "combustivel"
            ),
            2,
        )
        application_cost = round(sum(float(record.cost or 0) for record in season_fertilizations), 2)
        consumed_inputs_cost = round(input_cost + fuel_cost, 2)
        operational_total = round(consumed_inputs_cost + application_cost, 2)
        cultivated_area = float(season.cultivated_area or 0)
        season_costs[season.id] = {
            "input_cost": input_cost,
            "fuel_cost": fuel_cost,
            "application_cost": application_cost,
            "consumed_inputs_cost": consumed_inputs_cost,
            "operational_total": operational_total,
            "cost_per_hectare": round(operational_total / cultivated_area, 2) if cultivated_area else 0,
            "farm_cost": operational_total,
            "culture_cost": operational_total,
            "variety_cost": operational_total if season.variety_id else 0,
        }
    return templates.TemplateResponse(
        "seasons.html",
        _base_context(
            request,
            user,
            csrf_token,
            "seasons",
            _repo=repo,
            title="Safras",
            farms=repo.list_farms(),
            varieties=repo.list_varieties(),
            crop_seasons=crop_seasons,
            season_costs=season_costs,
            edit_season=edit_season,
            view_season=view_season,
        ),
    )


@router.post("/safras")
def create_crop_season_action(
    request: Request,
    csrf_token: str = Form(...),
    name: str | None = Form(None),
    start_date: str = Form(...),
    end_date: str = Form(...),
    culture: str = Form(...),
    variety_id: str | None = Form(None),
    cultivated_area: str | None = Form(None),
    area_unit: str = Form("ha"),
    notes: str | None = Form(None),
    status_value: str = Form(..., alias="status"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/safras", require_season=False)
    if denied:
        return denied
    try:
        parsed_start_date = date.fromisoformat(start_date)
        parsed_end_date = date.fromisoformat(end_date)
    except ValueError:
        _flash(request, "error", "Informe datas validas para a safra.")
        return _redirect("/safras")
    try:
        normalized_area = str(cultivated_area or "").strip().replace(",", ".")
        parsed_cultivated_area = float(normalized_area)
    except ValueError:
        _flash(request, "error", "Informe uma area cultivada valida.")
        return _redirect("/safras")
    if parsed_cultivated_area <= 0:
        _flash(request, "error", "A area cultivada deve ser maior que zero.")
        return _redirect("/safras")
    if parsed_end_date < parsed_start_date:
        _flash(request, "error", "A data final da safra nao pode ser anterior a data inicial.")
        return _redirect("/safras")
    try:
        create_crop_season(
            repo,
            {
                "farm_id": scope["active_farm_id"],
                "name": name,
                "start_date": start_date,
                "end_date": end_date,
                "culture": culture,
                "variety_id": _int_or_none(variety_id),
                "cultivated_area": parsed_cultivated_area,
                "area_unit": area_unit,
                "notes": notes,
                "status": status_value,
            },
        )
    except Exception:
        logger.exception("Falha ao criar safra", extra={"farm_id": scope["active_farm_id"], "start_date": start_date, "end_date": end_date})
        _flash(request, "error", "Nao foi possivel salvar a safra agora. Revise os dados e tente novamente.")
        return _redirect("/safras")
    _flash(request, "success", "Safra cadastrada com sucesso.")
    return _redirect("/safras")


@router.post("/safras/{season_id}/editar")
def update_crop_season_action(
    season_id: int,
    request: Request,
    csrf_token: str = Form(...),
    name: str | None = Form(None),
    start_date: str = Form(...),
    end_date: str = Form(...),
    culture: str = Form(...),
    variety_id: str | None = Form(None),
    cultivated_area: str | None = Form(None),
    area_unit: str = Form("ha"),
    notes: str | None = Form(None),
    status_value: str = Form(..., alias="status"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/safras", require_season=False)
    if denied:
        return denied
    crop_season = repo.get_crop_season(season_id)
    if not crop_season:
        _flash(request, "error", "Safra nao encontrada.")
        return _redirect("/safras")
    if not _farm_matches_scope(crop_season.farm_id, scope):
        _flash(request, "error", "Esta safra nao pertence ao contexto ativo.")
        return _redirect("/safras")
    try:
        parsed_start_date = date.fromisoformat(start_date)
        parsed_end_date = date.fromisoformat(end_date)
    except ValueError:
        _flash(request, "error", "Informe datas validas para a safra.")
        return _redirect("/safras")
    try:
        normalized_area = str(cultivated_area or "").strip().replace(",", ".")
        parsed_cultivated_area = float(normalized_area)
    except ValueError:
        _flash(request, "error", "Informe uma area cultivada valida.")
        return _redirect("/safras")
    if parsed_cultivated_area <= 0:
        _flash(request, "error", "A area cultivada deve ser maior que zero.")
        return _redirect("/safras")
    if parsed_end_date < parsed_start_date:
        _flash(request, "error", "A data final da safra nao pode ser anterior a data inicial.")
        return _redirect("/safras")
    try:
        update_crop_season(
            repo,
            crop_season,
            {
                "farm_id": scope["active_farm_id"],
                "name": name,
                "start_date": start_date,
                "end_date": end_date,
                "culture": culture,
                "variety_id": _int_or_none(variety_id),
                "cultivated_area": parsed_cultivated_area,
                "area_unit": area_unit,
                "notes": notes,
                "status": status_value,
            },
        )
    except Exception:
        logger.exception("Falha ao atualizar safra", extra={"season_id": season_id, "farm_id": scope["active_farm_id"]})
        _flash(request, "error", "Nao foi possivel atualizar a safra agora. Revise os dados e tente novamente.")
        return _redirect("/safras")
    _flash(request, "success", "Safra atualizada com sucesso.")
    return _redirect("/safras")


@router.post("/safras/{season_id}/excluir")
def delete_crop_season_action(
    season_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    crop_season = repo.get_crop_season(season_id)
    if not crop_season:
        _flash(request, "error", "Safra nao encontrada.")
        return _redirect("/safras")
    repo.delete(crop_season)
    _flash(request, "success", "Safra excluida com sucesso.")
    return _redirect("/safras")


def _irrigation_history_dataset(
    request: Request,
    repo: FarmRepository,
    *,
    flash_invalid: bool,
) -> dict:
    """Mesmos filtros da listagem em /irrigacao (sem paginação)."""
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date, filter_start_str, filter_end_str = _schedule_filter_date_bounds(
        request, scope["active_season"], flash_invalid=flash_invalid
    )
    selected_irrigation_range = (
        _fertilization_filter_range_preset(
            request.query_params.get("schedule_range"),
            filter_start_str,
            filter_end_str,
        )
        if _period_filter_explicit_in_query(request)
        else ""
    )
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    raw_plot = (request.query_params.get("plot_id") or "").strip()
    raw_plot_lower = raw_plot.lower()
    # Todos / sem filtro por setor: vazio ou valores explícitos (URL, export, API)
    if not raw_plot or raw_plot_lower in {"all", "todos", "todas"}:
        plot_id_filter = None
    else:
        plot_id_filter = _int_or_none(raw_plot)
        if plot_id_filter is not None and plot_id_filter not in plot_ids:
            plot_id_filter = None
    search_q = (request.query_params.get("search") or "").strip() or None

    irrigations_in_period = [
        irrigation
        for irrigation in repo.list_irrigations()
        if irrigation.plot_id in plot_ids and _within_scope(irrigation.irrigation_date, start_date, end_date)
    ]
    plot_ids_with_irrigation = {item.plot_id for item in irrigations_in_period}
    plots_for_irrigation_filter = sorted(
        (p for p in plots if p.id in plot_ids_with_irrigation),
        key=lambda p: (p.name or "").lower(),
    )
    if plot_id_filter is not None and plot_id_filter not in plot_ids_with_irrigation:
        orphan = next((p for p in plots if p.id == plot_id_filter), None)
        if orphan:
            plots_for_irrigation_filter = [orphan, *plots_for_irrigation_filter]

    irrigations = list(irrigations_in_period)
    if plot_id_filter is not None:
        irrigations = [item for item in irrigations if item.plot_id == plot_id_filter]
    if search_q:
        query_norm = _normalize_search_value(search_q)
        irrigations = [
            item
            for item in irrigations
            if query_norm
            in _normalize_search_value(
                f"{item.plot.name if item.plot else ''} "
                f"{item.irrigation_date or ''} "
                f"{item.volume_liters or ''} "
                f"{item.duration_minutes or ''} "
                f"{item.notes or ''}"
            )
        ]
    irrigations = _sort_collection_desc(
        irrigations,
        lambda item: item.irrigation_date,
        lambda item: item.id,
    )

    return {
        "scope": scope,
        "plots": plots,
        "plots_for_irrigation_filter": plots_for_irrigation_filter,
        "filter_start_str": filter_start_str,
        "filter_end_str": filter_end_str,
        "selected_irrigation_range": selected_irrigation_range,
        "raw_plot": raw_plot,
        "plot_id_filter": plot_id_filter,
        "search_q": search_q,
        "irrigations_filtered": irrigations,
    }


@router.get("/irrigacao")
def irrigation_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    data = _irrigation_history_dataset(request, repo, flash_invalid=True)
    scope = data["scope"]
    plots = data["plots"]
    plots_for_irrigation_filter = data["plots_for_irrigation_filter"]
    filter_start_str = data["filter_start_str"]
    filter_end_str = data["filter_end_str"]
    selected_irrigation_range = data["selected_irrigation_range"]
    raw_plot = data["raw_plot"]
    plot_id_filter = data["plot_id_filter"]
    search_q = data["search_q"]
    irrigations = data["irrigations_filtered"]

    irrigation_filter_clear_url = _url_with_query(
        request,
        start_date=None,
        end_date=None,
        schedule_range=None,
        plot_id=None,
        irrigations_page=None,
        search=None,
    )
    irrigation_after_edit_close_url = _url_with_query(request, edit_id=None)

    irrigations_pagination = _paginate_collection(request, irrigations, "irrigations_page")
    irrigation_edit_urls = {
        item.id: _url_with_query(request, edit_id=item.id) for item in irrigations_pagination["items"]
    }
    irrigation_filters_active = bool(
        _period_filter_explicit_in_query(request) or (plot_id_filter is not None) or bool(search_q)
    )
    raw_plot_lower = raw_plot.lower()
    irrigation_plot_filter_hint = raw_plot == ""
    selected_irrigation_plot_all = bool(raw_plot) and raw_plot_lower in {"all", "todos", "todas"}
    return templates.TemplateResponse(
        "irrigation.html",
        _base_context(
            request,
            user,
            csrf_token,
            "irrigation",
            _repo=repo,
            plots=plots,
            plots_for_irrigation_filter=plots_for_irrigation_filter,
            irrigations=irrigations_pagination["items"],
            irrigations_pagination=irrigations_pagination,
            irrigation_filter_start_date=filter_start_str or None,
            irrigation_filter_end_date=filter_end_str or None,
            irrigation_filter_clear_url=irrigation_filter_clear_url,
            irrigation_after_edit_close_url=irrigation_after_edit_close_url,
            selected_irrigation_range=selected_irrigation_range,
            selected_irrigation_plot_id=plot_id_filter,
            irrigation_plot_filter_hint=irrigation_plot_filter_hint,
            selected_irrigation_plot_all=selected_irrigation_plot_all,
            irrigation_filters_active=irrigation_filters_active,
            irrigation_edit_urls=irrigation_edit_urls,
            edit_irrigation=repo.get_irrigation(edit_id) if edit_id else None,
            plot_irrigation_configs=[
                {
                    "id": plot.id,
                    "name": plot.name,
                    "type": plot.irrigation_type,
                    "line_count": plot.irrigation_line_count,
                    "line_length": float(plot.irrigation_line_length_meters) if plot.irrigation_line_length_meters is not None else None,
                    "drip_spacing_cm": round(float(plot.drip_spacing_meters) * 100, 2) if plot.drip_spacing_meters is not None else None,
                    "drip_lph": float(plot.drip_liters_per_hour) if plot.drip_liters_per_hour is not None else None,
                    "sprinkler_count": plot.sprinkler_count,
                    "sprinkler_lph": float(plot.sprinkler_liters_per_hour) if plot.sprinkler_liters_per_hour is not None else None,
                }
                for plot in plots
            ],
        ),
    )


@router.get("/irrigacao/exportar.xlsx")
def export_irrigation_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    data = _irrigation_history_dataset(request, repo, flash_invalid=False)
    irrigations = data["irrigations_filtered"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Irrigacao"
    sheet.append(["Setor", "Data", "Volume (L)", "Duração (min)", "Observações"])
    for item in irrigations:
        sheet.append([
            item.plot.name if item.plot else "Setor removido",
            item.irrigation_date.isoformat() if item.irrigation_date else "",
            _format_decimal_br(float(item.volume_liters or 0), 2),
            int(item.duration_minutes or 0),
            item.notes or "",
        ])
    for index, width in enumerate([28, 14, 16, 16, 48], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="irrigacao.xlsx"'},
    )


@router.get("/irrigacao/exportar.pdf")
def export_irrigation_pdf(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    data = _irrigation_history_dataset(request, repo, flash_invalid=False)
    irrigations = data["irrigations_filtered"]
    scope = data["scope"]
    raw_start = data["filter_start_str"]
    raw_end = data["filter_end_str"]
    plot_id_filter = data["plot_id_filter"]
    plots = data["plots"]
    search_q = data["search_q"]

    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    farm_name = scope["active_farm"].name if scope.get("active_farm") else "Fazenda Bela Vista"
    season_label = scope["active_season"].name if scope.get("active_season") else "Safra ativa"
    period_label = "Safra ativa"
    if raw_start or raw_end:
        period_label = f"{_format_iso_date_br(raw_start)} a {_format_iso_date_br(raw_end)}"
    plot_filter_label = "Todos os setores"
    if plot_id_filter:
        sel_plot = next((p for p in plots if p.id == plot_id_filter), None)
        plot_filter_label = sel_plot.name if sel_plot else f"Setor #{plot_id_filter}"
    search_label = (search_q or "").strip() or "—"
    total_volume_liters = sum(float(item.volume_liters or 0) for item in irrigations)
    total_volume_m3 = total_volume_liters / 1000.0
    total_duration = sum(int(item.duration_minutes or 0) for item in irrigations)
    duration_total_label = _format_duration_hours_minutes(total_duration)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle(
        "IrrigationPdfFarmHeader",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#1e293b"),
    )
    meta_label_style = ParagraphStyle(
        "IrrigationPdfMetaLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#446a36"),
        spaceAfter=2,
    )
    meta_value_style = ParagraphStyle(
        "IrrigationPdfMetaValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#334155"),
    )
    cell_style = ParagraphStyle(
        "IrrigationPdfCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.2,
        leading=10.4,
        textColor=colors.HexColor("#0f172a"),
    )
    cell_muted_style = ParagraphStyle("IrrigationPdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    summary_value_style = ParagraphStyle(
        "IrrigationPdfSummaryValue",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#0f172a"),
    )

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    summary_table = Table([
        [
            [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
            [Paragraph("SAFRA", meta_label_style), Paragraph(season_label, meta_value_style)],
            [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
        ],
        [
            [Paragraph("SETOR (FILTRO)", meta_label_style), Paragraph(plot_filter_label, meta_value_style)],
            [Paragraph("BUSCA", meta_label_style), Paragraph(search_label, meta_value_style)],
            [Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(irrigations)), summary_value_style)],
        ],
    ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    title_style = ParagraphStyle(
        "IrrigationPdfTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=6,
    )
    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 12), Paragraph("Histórico de irrigação", title_style), Spacer(1, 8)]
    data_rows = [["Setor", "Data", "Volume (L)", "Duração (min)", "Observações"]]
    for item in irrigations:
        data_rows.append([
            Paragraph(item.plot.name if item.plot else "Setor removido", cell_style),
            Paragraph(item.irrigation_date.strftime("%d/%m/%Y") if item.irrigation_date else "-", cell_style),
            Paragraph(_format_decimal_br(float(item.volume_liters or 0), 2), cell_style),
            Paragraph(str(int(item.duration_minutes or 0)), cell_style),
            Paragraph(item.notes or "-", cell_muted_style),
        ])

    table = Table(data_rows, colWidths=[120, 72, 88, 88, doc.width - 368], repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 14))

    footer_summary = Table([
        ["Registros", str(len(irrigations))],
        ["Volume total (m³)", _format_decimal_br(total_volume_m3, 3)],
        ["Duração total", duration_total_label],
    ], colWidths=[doc.width * 0.28, doc.width * 0.18], hAlign="LEFT")
    footer_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(footer_summary)

    generated_at_label = format_app_datetime(generated_at)

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="irrigacao.pdf"'},
    )


@router.post("/irrigacao")
def create_irrigation_action(
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    irrigation_date: str = Form(...),
    volume_liters: str | None = Form(None),
    duration_minutes: int = Form(...),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, _, denied = _resolve_plot_in_scope(request, repo, plot_id, "/irrigacao")
    if denied:
        return denied
    calculated_volume = calculate_irrigation_volume(plot, duration_minutes)
    manual_volume = _float_or_none(volume_liters)
    if calculated_volume is None and manual_volume is None:
        _flash(request, "error", "Informe o volume manual em litros ou cadastre os dados de irrigacao no setor.")
        return _redirect("/irrigacao")
    create_irrigation(
        repo,
        {
            "plot_id": plot_id,
            "irrigation_date": irrigation_date,
            "volume_liters": calculated_volume if calculated_volume is not None else manual_volume,
            "duration_minutes": duration_minutes,
            "notes": notes,
        },
    )
    _flash(request, "success", "Irrigacao registrada com sucesso.")
    return _redirect("/irrigacao")


@router.post("/irrigacao/{record_id}/editar")
def update_irrigation_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    plot_id: int = Form(...),
    irrigation_date: str = Form(...),
    volume_liters: str | None = Form(None),
    duration_minutes: int = Form(...),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/irrigacao"
    repo = _repository(db)
    irrigation = repo.get_irrigation(record_id)
    if not irrigation:
        _flash(request, "error", "Registro de irrigacao nao encontrado.")
        return _redirect(target_url)
    if (irrigation.origin or "") == "fertilizacao":
        _flash(
            request,
            "info",
            "Este lancamento foi gerado via Fertilizacao. A alteracao deve ser feita no modulo de origem (Agendamentos/Fertilizacao).",
        )
        return _redirect(target_url)
    plot, scope, denied = _resolve_plot_in_scope(request, repo, plot_id, "/irrigacao")
    if denied:
        return denied
    if not _plot_matches_scope(irrigation.plot, scope):
        _flash(request, "error", "Este registro de irrigacao nao pertence ao contexto ativo.")
        return _redirect(target_url)
    calculated_volume = calculate_irrigation_volume(plot, duration_minutes)
    manual_volume = _float_or_none(volume_liters)
    if calculated_volume is None and manual_volume is None:
        _flash(request, "error", "Informe o volume manual em litros ou cadastre os dados de irrigacao no setor.")
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={record_id}")
    update_irrigation(
        repo,
        irrigation,
        {
            "plot_id": plot.id,
            "irrigation_date": irrigation_date,
            "volume_liters": calculated_volume if calculated_volume is not None else manual_volume,
            "duration_minutes": duration_minutes,
            "notes": notes,
        },
    )
    _flash(request, "success", "Irrigacao atualizada com sucesso.")
    return _redirect(target_url)


@router.post("/irrigacao/{record_id}/excluir")
def delete_irrigation_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/irrigacao"
    repo = _repository(db)
    irrigation = repo.get_irrigation(record_id)
    if not irrigation:
        _flash(request, "error", "Registro de irrigacao nao encontrado.")
        return _redirect(target_url)
    if (irrigation.origin or "") == "fertilizacao":
        _flash(
            request,
            "info",
            "Este lancamento foi gerado via Fertilizacao. A exclusao deve ser feita no modulo de origem (Agendamentos/Fertilizacao).",
        )
        return _redirect(target_url)
    repo.delete(irrigation)
    _flash(request, "success", "Irrigacao excluida com sucesso.")
    return _redirect(target_url)


def _rainfall_history_dataset(
    request: Request,
    repo: FarmRepository,
    *,
    flash_invalid: bool,
) -> dict:
    """Listagem de pluviometria: fazenda ativa, período (safra + filtro) e busca."""
    scope = _global_scope_context(request, repo)
    farm_scope_id = _active_farm_id(request)
    start_date, end_date, filter_start_str, filter_end_str = _schedule_filter_date_bounds(
        request, scope["active_season"], flash_invalid=flash_invalid
    )
    selected_rainfall_range = (
        _fertilization_filter_range_preset(
            request.query_params.get("schedule_range"),
            filter_start_str,
            filter_end_str,
        )
        if _period_filter_explicit_in_query(request)
        else ""
    )
    search_q = (request.query_params.get("search") or "").strip() or None
    rainfalls = repo.list_rainfalls(
        farm_id=farm_scope_id,
        start_date=start_date,
        end_date=end_date,
    )
    if search_q:
        query_norm = _normalize_search_value(search_q)
        rainfalls = [
            item
            for item in rainfalls
            if query_norm
            in _normalize_search_value(
                f"{item.farm.name if item.farm else ''} "
                f"{item.rainfall_date or ''} "
                f"{item.millimeters or ''} "
                f"{item.source or ''} "
                f"{item.notes or ''}"
            )
        ]
    rainfalls = _sort_collection_desc(
        rainfalls,
        lambda item: item.rainfall_date,
        lambda item: item.id,
    )
    return {
        "scope": scope,
        "farm_scope_id": farm_scope_id,
        "filter_start_str": filter_start_str,
        "filter_end_str": filter_end_str,
        "selected_rainfall_range": selected_rainfall_range,
        "search_q": search_q,
        "rainfalls_filtered": rainfalls,
    }


@router.get("/pluviometria")
def rainfall_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    data = _rainfall_history_dataset(request, repo, flash_invalid=True)
    scope = data["scope"]
    filter_start_str = data["filter_start_str"]
    filter_end_str = data["filter_end_str"]
    selected_rainfall_range = data["selected_rainfall_range"]
    search_q = data["search_q"]
    rainfalls = data["rainfalls_filtered"]

    rainfall_filter_clear_url = _url_with_query(
        request,
        start_date=None,
        end_date=None,
        schedule_range=None,
        rainfalls_page=None,
        search=None,
        edit_id=None,
    )
    rainfall_after_edit_close_url = _url_with_query(request, edit_id=None)

    rainfalls_pagination = _paginate_collection(request, rainfalls, "rainfalls_page")
    rainfall_edit_urls = {
        item.id: _url_with_query(request, edit_id=item.id) for item in rainfalls_pagination["items"]
    }
    rainfall_filters_active = bool(
        _period_filter_explicit_in_query(request) or bool(search_q)
    )
    return templates.TemplateResponse(
        "rainfall.html",
        _base_context(
            request,
            user,
            csrf_token,
            "rainfall",
            _repo=repo,
            rainfalls=rainfalls_pagination["items"],
            rainfalls_pagination=rainfalls_pagination,
            rainfall_filter_start_date=filter_start_str or None,
            rainfall_filter_end_date=filter_end_str or None,
            rainfall_filter_clear_url=rainfall_filter_clear_url,
            rainfall_after_edit_close_url=rainfall_after_edit_close_url,
            selected_rainfall_range=selected_rainfall_range,
            rainfall_filters_active=rainfall_filters_active,
            rainfall_edit_urls=rainfall_edit_urls,
            edit_rainfall=repo.get_rainfall(edit_id) if edit_id else None,
        ),
    )


@router.post("/pluviometria")
def create_rainfall_action(
    request: Request,
    csrf_token: str = Form(...),
    rainfall_date: str = Form(...),
    millimeters: float = Form(...),
    source: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/pluviometria")
    if denied:
        return denied
    create_rainfall(
        repo,
        {
            "farm_id": scope["active_farm_id"],
            "rainfall_date": rainfall_date,
            "millimeters": millimeters,
            "source": source,
            "notes": notes,
        },
    )
    _flash(request, "success", "Pluviometria registrada com sucesso.")
    return _redirect("/pluviometria")


@router.post("/pluviometria/{record_id}/editar")
def update_rainfall_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    rainfall_date: str = Form(...),
    millimeters: float = Form(...),
    source: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/pluviometria")
    if denied:
        return denied
    rainfall = repo.get_rainfall(record_id)
    if not rainfall:
        _flash(request, "error", "Registro de pluviometria nao encontrado.")
        return _redirect_for_request(request, "/pluviometria")
    if not _farm_matches_scope(rainfall.farm_id, scope):
        _flash(request, "error", "Este registro de pluviometria nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/pluviometria")
    update_rainfall(
        repo,
        rainfall,
        {
            "farm_id": scope["active_farm_id"],
            "rainfall_date": rainfall_date,
            "millimeters": millimeters,
            "source": source,
            "notes": notes,
        },
    )
    _flash(request, "success", "Pluviometria atualizada com sucesso.")
    return _redirect_for_request(request, "/pluviometria")


@router.post("/pluviometria/{record_id}/excluir")
def delete_rainfall_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    rainfall = repo.get_rainfall(record_id)
    if not rainfall:
        _flash(request, "error", "Registro de pluviometria nao encontrado.")
        return _redirect("/pluviometria")
    repo.delete(rainfall)
    _flash(request, "success", "Pluviometria excluida com sucesso.")
    return _redirect("/pluviometria")


@router.get("/pluviometria/exportar.xlsx")
def export_rainfall_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    data = _rainfall_history_dataset(request, repo, flash_invalid=False)
    rainfalls = data["rainfalls_filtered"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pluviometria"
    sheet.append(["Fazenda", "Data", "Chuva (mm)", "Origem", "Observações"])
    for item in rainfalls:
        sheet.append([
            item.farm.name if item.farm else "Fazenda removida",
            item.rainfall_date.isoformat() if item.rainfall_date else "",
            _format_decimal_br(float(item.millimeters or 0), 2),
            item.source or "",
            item.notes or "",
        ])
    for index, width in enumerate([28, 14, 14, 28, 48], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="pluviometria.xlsx"'},
    )


@router.get("/pluviometria/exportar.pdf")
def export_rainfall_pdf(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    data = _rainfall_history_dataset(request, repo, flash_invalid=False)
    rainfalls = data["rainfalls_filtered"]
    scope = data["scope"]
    raw_start = data["filter_start_str"]
    raw_end = data["filter_end_str"]
    farm_scope_id = data["farm_scope_id"]
    search_q = data["search_q"]

    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    farm_name = scope["active_farm"].name if scope.get("active_farm") else "Fazenda Bela Vista"
    season_label = scope["active_season"].name if scope.get("active_season") else "Safra ativa"
    period_label = "Safra ativa"
    if raw_start or raw_end:
        period_label = f"{_format_iso_date_br(raw_start)} a {_format_iso_date_br(raw_end)}"
    scope_farm_label = (
        farm_name
        if farm_scope_id
        else "Todas as fazendas (sem fazenda ativa no contexto)"
    )
    search_label = (search_q or "").strip() or "—"
    total_mm = sum(float(item.millimeters or 0) for item in rainfalls)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle(
        "RainfallPdfFarmHeader",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#1e293b"),
    )
    meta_label_style = ParagraphStyle(
        "RainfallPdfMetaLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#446a36"),
        spaceAfter=2,
    )
    meta_value_style = ParagraphStyle(
        "RainfallPdfMetaValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#334155"),
    )
    cell_style = ParagraphStyle(
        "RainfallPdfCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.2,
        leading=10.4,
        textColor=colors.HexColor("#0f172a"),
    )
    cell_muted_style = ParagraphStyle("RainfallPdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    summary_value_style = ParagraphStyle(
        "RainfallPdfSummaryValue",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#0f172a"),
    )

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    summary_table = Table([
        [
            [Paragraph("FAZENDA (FILTRO)", meta_label_style), Paragraph(scope_farm_label, meta_value_style)],
            [Paragraph("SAFRA", meta_label_style), Paragraph(season_label, meta_value_style)],
            [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
        ],
        [
            [Paragraph("BUSCA", meta_label_style), Paragraph(search_label, meta_value_style)],
            [Paragraph("REGISTROS", meta_label_style), Paragraph(str(len(rainfalls)), summary_value_style)],
            [Paragraph("CHUVA TOTAL (mm)", meta_label_style), Paragraph(_format_decimal_br(total_mm, 2), summary_value_style)],
        ],
    ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    title_style = ParagraphStyle(
        "RainfallPdfTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=6,
    )
    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 12), Paragraph("Histórico de pluviometria", title_style), Spacer(1, 8)]
    data_rows = [["Fazenda", "Data", "Chuva (mm)", "Origem", "Observações"]]
    for item in rainfalls:
        data_rows.append([
            Paragraph(item.farm.name if item.farm else "Fazenda removida", cell_style),
            Paragraph(item.rainfall_date.strftime("%d/%m/%Y") if item.rainfall_date else "-", cell_style),
            Paragraph(_format_decimal_br(float(item.millimeters or 0), 2), cell_style),
            Paragraph(item.source or "-", cell_style),
            Paragraph(item.notes or "-", cell_muted_style),
        ])

    table = Table(data_rows, colWidths=[110, 72, 80, 100, doc.width - 362], repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 14))

    footer_summary = Table([
        ["Registros", str(len(rainfalls))],
        ["Chuva total (mm)", _format_decimal_br(total_mm, 2)],
    ], colWidths=[doc.width * 0.28, doc.width * 0.18], hAlign="LEFT")
    footer_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(footer_summary)

    generated_at_label = format_app_datetime(generated_at)

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="pluviometria.pdf"'},
    )


@router.get("/fertilizacao")
def fertilization_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date, filter_start_str, filter_end_str = _schedule_filter_date_bounds(
        request, scope["active_season"], flash_invalid=True
    )
    selected_fertilization_range = (
        _fertilization_filter_range_preset(
            request.query_params.get("schedule_range"),
            filter_start_str,
            filter_end_str,
        )
        if _period_filter_explicit_in_query(request)
        else ""
    )
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    edit_fertilization = repo.get_fertilization(edit_id) if edit_id else None
    recommendation_groups: dict[str, list[dict]] = {}
    consolidated_inputs = repo.list_input_catalog(item_type="insumo_agricola")
    purchased_inputs = repo.list_purchased_inputs()
    avail_by_catalog = defaultdict(float)
    for entry in purchased_inputs:
        avail_by_catalog[entry.input_id] += float(entry.available_quantity or 0)
    input_stock = {
        item.id: {
            "available": round(avail_by_catalog.get(item.id, 0), 2),
            "unit": item.default_unit,
        }
        for item in consolidated_inputs
    }
    fertilization_filter_clear_url = _url_with_query(
        request,
        start_date=None,
        end_date=None,
        schedule_range=None,
        fertilizations_page=None,
    )
    search_q = (request.query_params.get("search") or "").strip() or None
    fertilizations = _filter_fertilization_records(repo, plot_ids, start_date, end_date, search=search_q)
    fertilizations_pagination = _paginate_collection(request, fertilizations, "fertilizations_page")
    schedules = repo.list_fertilization_schedules_for_scope(plot_ids, start_date, end_date)
    recs = repo.list_input_recommendations(
        farm_id=scope["active_farm_id"] if scope["active_farm_id"] else None
    )
    recommendations = [
        recommendation
        for recommendation in recs
        if (not plot_ids or recommendation.plot_id is None or recommendation.plot_id in plot_ids)
    ]
    for recommendation in recommendations:
        bucket = recommendation_groups.setdefault(recommendation.application_name, [])
        for item in recommendation.items:
            bucket.append(
                {
                    "input_id": item.input_id,
                    "name": item.input_catalog.name if item.input_catalog else (item.purchased_input.name if item.purchased_input else "Insumo removido"),
                    "unit": item.unit,
                    "quantity": float(item.quantity or 0),
                    "available": input_stock.get(item.input_id or 0, {}).get("available", 0),
                }
            )
    return templates.TemplateResponse(
        "fertilization.html",
        _base_context(
            request,
            user,
            csrf_token,
            "fertilization",
            _repo=repo,
            plots=plots,
            inputs_catalog=consolidated_inputs,
            input_stock=input_stock,
            fertilizations=fertilizations_pagination["items"],
            fertilizations_pagination=fertilizations_pagination,
            fertilization_filter_start_date=filter_start_str or None,
            fertilization_filter_end_date=filter_end_str or None,
            fertilization_filter_clear_url=fertilization_filter_clear_url,
            selected_fertilization_range=selected_fertilization_range,
            schedules=schedules,
            recommendation_groups=recommendation_groups,
            edit_fertilization=edit_fertilization,
            edit_fertilization_items=_legacy_fertilization_items(
                edit_fertilization,
                float(edit_fertilization.plot.area_hectares) if edit_fertilization and edit_fertilization.plot else None,
            ),
        ),
    )


@router.post("/fertilizacao")
async def create_fertilization_action(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    csrf_token = str(form.get("csrf_token") or "")
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/fertilizacao")
    if denied:
        return denied
    selected_plot_ids: list[int] = []
    for raw_plot_id in form.getlist("plot_id"):
        parsed_plot_id = _int_or_none(raw_plot_id)
        if parsed_plot_id and parsed_plot_id not in selected_plot_ids:
            selected_plot_ids.append(parsed_plot_id)
    if not selected_plot_ids:
        _flash(request, "error", "Selecione ao menos um setor.")
        return _redirect("/fertilizacao")
    selected_plots = []
    for plot_id in selected_plot_ids:
        plot, _, denied = _resolve_plot_in_scope(request, repo, plot_id, "/fertilizacao")
        if denied:
            return denied
        selected_plots.append(plot)
    items = _parse_fertilization_items(form)
    if not items:
        _flash(request, "error", "Adicione ao menos um insumo na atividade.")
        return _redirect("/fertilizacao")
    try:
        for plot in selected_plots:
            create_fertilization(
                repo,
                {
                    "plot_id": plot.id,
                    "application_date": str(form.get("application_date") or ""),
                    "season_id": scope["active_season_id"],
                    "duration_minutes": form.get("duration_minutes"),
                    "application_method": form.get("application_method"),
                    "notes": str(form.get("notes") or "") or None,
                    "items": items,
                },
            )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect("/fertilizacao")
    _flash(
        request,
        "success",
        "Fertilizacao registrada com sucesso."
        if len(selected_plots) == 1
        else f"Fertilizacao registrada com sucesso para {len(selected_plots)} setores.",
    )
    return _redirect("/fertilizacao")


@router.post("/fertilizacao/{record_id}/editar")
async def update_fertilization_action(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    csrf_token = str(form.get("csrf_token") or "")
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    fertilization = repo.get_fertilization(record_id)
    if not fertilization:
        _flash(request, "error", "Registro de fertilizacao nao encontrado.")
        return _redirect_for_request(request, "/fertilizacao")
    plot, scope, denied = _resolve_plot_in_scope(request, repo, int(form.get("plot_id") or 0), "/fertilizacao")
    if denied:
        return denied
    if not _plot_matches_scope(fertilization.plot, scope):
        _flash(request, "error", "Este registro de fertilizacao nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/fertilizacao")
    items = _parse_fertilization_items(form)
    if not items:
        _flash(request, "error", "Adicione ao menos um insumo na atividade.")
        return _redirect_for_request(request, "/fertilizacao", edit_id=record_id)
    try:
        update_fertilization(
            repo,
            fertilization,
            {
                "plot_id": plot.id,
                "application_date": str(form.get("application_date") or ""),
                "season_id": scope["active_season_id"],
                "duration_minutes": form.get("duration_minutes"),
                "application_method": form.get("application_method"),
                "notes": str(form.get("notes") or "") or None,
                "items": items,
            },
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect_for_request(request, "/fertilizacao", edit_id=record_id)
    _flash(request, "success", "Fertilizacao atualizada com sucesso.")
    return _redirect_for_request(request, "/fertilizacao")


@router.post("/fertilizacao/{record_id}/excluir")
def delete_fertilization_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    fertilization = repo.get_fertilization(record_id)
    if not fertilization:
        _flash(request, "error", "Registro de fertilizacao nao encontrado.")
        return _redirect("/fertilizacao")
    delete_fertilization(repo, fertilization)
    _flash(request, "success", "Fertilizacao excluida com sucesso.")
    return _redirect("/fertilizacao")


@router.get("/fertilizacao/exportar.xlsx")
def export_fertilization_xlsx(
    request: Request,
    search: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date, _, _ = _schedule_filter_date_bounds(request, scope["active_season"], flash_invalid=False)
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    fertilizations = _filter_fertilization_records(repo, plot_ids, start_date, end_date, search=search)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fertilizacao"
    sheet.append(["Data", "Setor", "Modo", "Produto", "Custo", "Insumos Aplicados", "Observações"])
    for item in fertilizations:
        sheet.append([
            item.application_date.isoformat() if item.application_date else "",
            item.plot.name if item.plot else "Setor removido",
            fertilization_application_method_label(getattr(item, "application_method", None)),
            item.product or "",
            _format_currency(item.cost),
            " | ".join(
                f"{detail.name} ({_format_decimal_br(detail.total_quantity, 2)} {detail.unit})"
                for detail in item.items
            ) or (item.dose or ""),
            item.notes or "",
        ])
    for index, width in enumerate([14, 28, 18, 26, 16, 54, 40], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="fertilizacao.xlsx"'},
    )


@router.get("/fertilizacao/exportar.pdf")
def export_fertilization_pdf(
    request: Request,
    search: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date, raw_start, raw_end = _schedule_filter_date_bounds(request, scope["active_season"], flash_invalid=False)
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    fertilizations = _filter_fertilization_records(repo, plot_ids, start_date, end_date, search=search)

    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    farm_name = scope["active_farm"].name if scope.get("active_farm") else "Fazenda Bela Vista"
    season_label = scope["active_season"].name if scope.get("active_season") else "Safra ativa"
    period_label = "Safra ativa"
    if raw_start or raw_end:
        period_label = f"{(raw_start or '--').replace('-', '/')} a {(raw_end or '--').replace('-', '/')}"
    total_cost = sum(float(item.cost or 0) for item in fertilizations)
    total_items = sum(len(item.items or []) for item in fertilizations)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle("FertilizationPdfFarmHeader", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, alignment=TA_RIGHT, textColor=colors.HexColor("#1e293b"))
    meta_label_style = ParagraphStyle("FertilizationPdfMetaLabel", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#446a36"), spaceAfter=2)
    meta_value_style = ParagraphStyle("FertilizationPdfMetaValue", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13, textColor=colors.HexColor("#334155"))
    cell_style = ParagraphStyle("FertilizationPdfCell", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.2, leading=10.4, textColor=colors.HexColor("#0f172a"))
    cell_muted_style = ParagraphStyle("FertilizationPdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    summary_value_style = ParagraphStyle("FertilizationPdfSummaryValue", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#0f172a"))

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    summary_table = Table([
        [
            [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
            [Paragraph("SAFRA", meta_label_style), Paragraph(season_label, meta_value_style)],
            [Paragraph("PERÍODO", meta_label_style), Paragraph(period_label, meta_value_style)],
        ],
        [
            [Paragraph("LANÇAMENTOS", meta_label_style), Paragraph(str(len(fertilizations)), summary_value_style)],
            [Paragraph("INSUMOS APLICADOS", meta_label_style), Paragraph(str(total_items), summary_value_style)],
            [Paragraph("CUSTO TOTAL", meta_label_style), Paragraph(_format_currency(total_cost), summary_value_style)],
        ],
    ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 14)]
    data = [["Data", "Setor", "Modo", "Produto", "Custo", "Insumos Aplicados", "Observações"]]
    for item in fertilizations:
        items_label = " • ".join(
            f"{detail.name} ({_format_decimal_br(detail.total_quantity, 2)} {detail.unit})"
            for detail in item.items
        ) or (item.dose or "-")
        data.append([
            Paragraph(item.application_date.strftime("%d/%m/%Y") if item.application_date else "-", cell_style),
            Paragraph(item.plot.name if item.plot else "Setor removido", cell_style),
            Paragraph(fertilization_application_method_label(getattr(item, "application_method", None)), cell_style),
            Paragraph(item.product or "-", cell_style),
            Paragraph(_format_currency(item.cost), cell_style),
            Paragraph(items_label, cell_muted_style),
            Paragraph(item.notes or "-", cell_muted_style),
        ])

    table = Table(data, colWidths=[56, 100, 72, 96, 70, 210, doc.width - 604], repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 14))

    footer_summary = Table([
        ["Total de lançamentos", str(len(fertilizations))],
        ["Total de insumos aplicados", str(total_items)],
        ["Custo total", _format_currency(total_cost)],
    ], colWidths=[doc.width * 0.28, doc.width * 0.18], hAlign="LEFT")
    footer_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(footer_summary)

    generated_at_label = format_app_datetime(generated_at)

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="fertilizacao.pdf"'},
    )


@router.get("/fertilizacao/agendamentos")
def fertilization_schedules_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    selected_schedule_tab = str(request.query_params.get("schedule_tab") or "active")
    if selected_schedule_tab not in {"active", "completed"}:
        selected_schedule_tab = "active"
    start_date, end_date, filter_start_str, filter_end_str = _schedule_filter_date_bounds(
        request, scope["active_season"], flash_invalid=True
    )
    selected_schedule_range = (
        _schedule_tab_filter_range_preset(
            request.query_params.get("schedule_range"),
            filter_start_str,
            filter_end_str,
            schedule_tab=selected_schedule_tab,
        )
        if _period_filter_explicit_in_query(request)
        else ""
    )
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    edit_schedule = repo.get_fertilization_schedule(edit_id) if edit_id else None
    schedules = repo.list_fertilization_schedules_for_scope(plot_ids, start_date, end_date)
    schedules.sort(key=lambda schedule: (schedule.scheduled_date, schedule.id), reverse=True)
    schedule_filter_clear_url = _url_with_query(
        request,
        start_date=None,
        end_date=None,
        schedule_range=None,
        active_page=None,
        completed_page=None,
        schedule_tab=selected_schedule_tab,
    )
    active_schedules = [schedule for schedule in schedules if schedule.status != "completed"]
    completed_schedules = [schedule for schedule in schedules if schedule.status == "completed"]
    active_schedules_pagination = _paginate_collection(request, active_schedules, "active_page")
    completed_schedules_pagination = _paginate_collection(request, completed_schedules, "completed_page")
    schedule_edit_urls = {
        schedule.id: _url_with_query(request, edit_id=schedule.id)
        for schedule in [*active_schedules_pagination["items"], *completed_schedules_pagination["items"]]
    }
    schedule_after_edit_close_url = _url_with_query(request, edit_id=None)
    purchased_inputs = repo.list_purchased_inputs()
    schedule_validations = {}
    for sch in active_schedules_pagination["items"]:
        schedule_validations[sch.id] = validate_schedule_stock(
            repo, sch, purchased_inputs_cache=purchased_inputs
        )
    for sch in completed_schedules_pagination["items"]:
        schedule_validations[sch.id] = {"ok": True, "shortages": []}
    recs = repo.list_input_recommendations(
        farm_id=scope["active_farm_id"] if scope["active_farm_id"] else None
    )
    recommendations = [
        recommendation
        for recommendation in recs
        if (not plot_ids or recommendation.plot_id is None or recommendation.plot_id in plot_ids)
    ]
    consolidated_inputs = repo.list_input_catalog(item_type="insumo_agricola")
    avail_by_catalog = defaultdict(float)
    for entry in purchased_inputs:
        avail_by_catalog[entry.input_id] += float(entry.available_quantity or 0)
    input_stock = {
        item.id: {
            "available": round(avail_by_catalog.get(item.id, 0), 2),
            "unit": item.default_unit,
        }
        for item in consolidated_inputs
    }
    edit_schedule_items = (
        [
            {
                "input_id": item.input_id,
                "unit": item.unit,
                "quantity": float(item.quantity or 0),
            }
            for item in edit_schedule.items
        ]
        if edit_schedule and edit_schedule.items
        else [{"input_id": "", "unit": "kg", "quantity": ""}]
    )
    schedule_recommendations = [
        {
            "id": recommendation.id,
            "application_name": recommendation.application_name,
            "plot_name": recommendation.plot.name if recommendation.plot else None,
            "farm_name": recommendation.farm.name if recommendation.farm else None,
            "items": [
                {
                    "input_id": item.input_id,
                    "unit": item.unit,
                    "quantity": float(item.quantity or 0),
                }
                for item in recommendation.items
                if item.input_id
            ],
        }
        for recommendation in recommendations
        if recommendation.items
    ]
    return templates.TemplateResponse(
        "fertilization_schedule.html",
        _base_context(
            request,
            user,
            csrf_token,
            "fertilization_schedules",
            _repo=repo,
            title="Agendamento de Fertilizacao",
            plots=plots,
            inputs_catalog=consolidated_inputs,
            input_stock=input_stock,
            active_schedules=active_schedules_pagination["items"],
            active_schedules_pagination=active_schedules_pagination,
            completed_schedules=completed_schedules_pagination["items"],
            completed_schedules_pagination=completed_schedules_pagination,
            selected_schedule_tab=selected_schedule_tab,
            schedule_tab_urls={
                "active": _fertilization_schedule_tab_url(request, "active"),
                "completed": _fertilization_schedule_tab_url(request, "completed"),
            },
            schedule_filter_start_date=filter_start_str or None,
            schedule_filter_end_date=filter_end_str or None,
            selected_schedule_range=selected_schedule_range,
            schedule_filter_clear_url=schedule_filter_clear_url,
            schedule_validations=schedule_validations,
            schedule_edit_urls=schedule_edit_urls,
            schedule_after_edit_close_url=schedule_after_edit_close_url,
            edit_schedule=edit_schedule,
            edit_schedule_items=edit_schedule_items,
            schedule_recommendations=schedule_recommendations,
            today=today_in_app_timezone(),
        ),
    )


@router.get("/fertilizacao/agendamentos/exportar.xlsx")
def export_fertilization_schedules_xlsx(
    request: Request,
    schedule_tab: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date, _, _ = _schedule_filter_date_bounds(request, scope["active_season"], flash_invalid=False)
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    schedules = repo.list_fertilization_schedules_for_scope(plot_ids, start_date, end_date)
    schedules.sort(key=lambda schedule: (schedule.scheduled_date, schedule.id), reverse=True)
    selected_schedule_tab = schedule_tab if schedule_tab in {"active", "completed"} else "active"
    search_query = (search or "").strip().lower()
    if search_query:
        schedules = [
            schedule
            for schedule in schedules
            if search_query in (
                f"{schedule.plot.name if schedule.plot else ''} "
                f"{schedule.scheduled_date or ''} "
                f"{schedule.status or ''} "
                f"{schedule.notes or ''} "
                f"{fertilization_application_method_label(getattr(schedule, 'application_method', None))} "
                + " ".join(item.input_catalog.name if item.input_catalog else item.name for item in schedule.items)
            ).lower()
        ]
    filtered_schedules = [schedule for schedule in schedules if schedule.status == "completed"] if selected_schedule_tab == "completed" else [schedule for schedule in schedules if schedule.status != "completed"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Concluídos" if selected_schedule_tab == "completed" else "Ativos"
    sheet.append(["Data", "Setor", "Modo", "Status", "Itens Programados", "Observações"])
    for schedule in filtered_schedules:
        sheet.append([
            schedule.scheduled_date.isoformat() if schedule.scheduled_date else "",
            schedule.plot.name if schedule.plot else "Setor removido",
            fertilization_application_method_label(getattr(schedule, "application_method", None)),
            "Concluído" if schedule.status == "completed" else "Agendado",
            " | ".join(
                f"{item.input_catalog.name if item.input_catalog else item.name} ({_format_decimal_br(item.quantity, 2)} {item.unit})"
                for item in schedule.items
            ),
            schedule.notes or "",
        ])
    for index, width in enumerate([14, 28, 18, 16, 54, 40], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="planejamento_agendamentos.xlsx"'},
    )


@router.get("/fertilizacao/agendamentos/exportar.pdf")
def export_fertilization_schedules_pdf(
    request: Request,
    schedule_tab: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date, _, _ = _schedule_filter_date_bounds(request, scope["active_season"], flash_invalid=False)
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    schedules = repo.list_fertilization_schedules_for_scope(plot_ids, start_date, end_date)
    schedules.sort(key=lambda schedule: (schedule.scheduled_date, schedule.id), reverse=True)
    selected_schedule_tab = schedule_tab if schedule_tab in {"active", "completed"} else "active"
    search_query = (search or "").strip().lower()
    if search_query:
        schedules = [
            schedule
            for schedule in schedules
            if search_query in (
                f"{schedule.plot.name if schedule.plot else ''} "
                f"{schedule.scheduled_date or ''} "
                f"{schedule.status or ''} "
                f"{schedule.notes or ''} "
                f"{fertilization_application_method_label(getattr(schedule, 'application_method', None))} "
                + " ".join(item.input_catalog.name if item.input_catalog else item.name for item in schedule.items)
            ).lower()
        ]
    schedules = [schedule for schedule in schedules if schedule.status == "completed"] if selected_schedule_tab == "completed" else [schedule for schedule in schedules if schedule.status != "completed"]

    generated_at = app_now()
    generated_by = user.display_name or user.name or user.email
    farm_name = scope["active_farm"].name if scope.get("active_farm") else "Fazenda Bela Vista"
    active_count = sum(1 for schedule in schedules if schedule.status != "completed")
    completed_count = sum(1 for schedule in schedules if schedule.status == "completed")
    total_items = sum(len(schedule.items or []) for schedule in schedules)
    season_label = scope["active_season"].name if scope.get("active_season") else "Safra ativa"
    scope_label = "Concluídos" if selected_schedule_tab == "completed" else "Ativos"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=32, bottomMargin=34)
    styles = getSampleStyleSheet()
    farm_header_style = ParagraphStyle("SchedulePdfFarmHeader", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, alignment=TA_RIGHT, textColor=colors.HexColor("#1e293b"))
    meta_label_style = ParagraphStyle("SchedulePdfMetaLabel", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#446a36"), spaceAfter=2)
    meta_value_style = ParagraphStyle("SchedulePdfMetaValue", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=13, textColor=colors.HexColor("#334155"))
    cell_style = ParagraphStyle("SchedulePdfCell", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.2, leading=10.4, textColor=colors.HexColor("#0f172a"))
    cell_muted_style = ParagraphStyle("SchedulePdfCellMuted", parent=cell_style, textColor=colors.HexColor("#475569"))
    summary_value_style = ParagraphStyle("SchedulePdfSummaryValue", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#0f172a"))

    logo_path = Path("app/static/images/logo.png")
    logo_flowable = Image(str(logo_path), width=92.8, height=73.6) if logo_path.exists() else Spacer(92.8, 73.6)
    header_table = Table([[logo_flowable, Paragraph(farm_name, farm_header_style)]], colWidths=[76, doc.width - 76], hAlign="LEFT")
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    summary_table = Table([
        [
            [Paragraph("FAZENDA", meta_label_style), Paragraph(farm_name, meta_value_style)],
            [Paragraph("SAFRA", meta_label_style), Paragraph(season_label, meta_value_style)],
            [Paragraph("LISTA", meta_label_style), Paragraph(scope_label, meta_value_style)],
        ],
        [
            [Paragraph("CONCLUÍDOS", meta_label_style), Paragraph(str(completed_count), summary_value_style)],
            [Paragraph("TOTAL DE AGENDAMENTOS", meta_label_style), Paragraph(str(len(schedules)), summary_value_style)],
            [Paragraph("ITENS PROGRAMADOS", meta_label_style), Paragraph(str(total_items), summary_value_style)],
        ],
    ], colWidths=[doc.width / 3] * 3, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#dbe5dd")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    elements = [header_table, Spacer(1, 16), summary_table, Spacer(1, 14)]
    column_weights = [9, 16, 10, 11, 30, 24]
    weight_total = sum(column_weights)
    table_col_widths = [doc.width * (weight / weight_total) for weight in column_weights[:-1]]
    table_col_widths.append(doc.width - sum(table_col_widths))
    notes_col_width = table_col_widths[-1] - 14
    notes_max_height = cell_muted_style.leading * 4 + 2

    data = [["Data", "Setor", "Modo", "Situação", "Itens Programados", "Observações"]]
    for schedule in schedules:
        items_label = " • ".join(
            f"{item.input_catalog.name if item.input_catalog else item.name} ({_format_decimal_br(item.quantity, 2)} {item.unit})"
            for item in schedule.items
        ) or "-"
        status_label = "Concluído" if schedule.status == "completed" else "Agendado"
        status_color = "#166534" if schedule.status == "completed" else "#0369a1"
        data.append([
            Paragraph(schedule.scheduled_date.strftime("%d/%m/%Y") if schedule.scheduled_date else "-", cell_style),
            Paragraph(schedule.plot.name if schedule.plot else "Setor removido", cell_style),
            Paragraph(fertilization_application_method_label(getattr(schedule, "application_method", None)), cell_style),
            Paragraph(f'<font color="{status_color}"><b>{status_label}</b></font>', cell_style),
            Paragraph(items_label, cell_muted_style),
            KeepInFrame(notes_col_width, notes_max_height, [Paragraph(schedule.notes or "-", cell_muted_style)], mode="truncate"),
        ])

    table = Table(data, repeatRows=1, colWidths=table_col_widths, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#446a36")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#36552a")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.2),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    footer_summary = Table([[
        Paragraph(f"<b>Ativos:</b> {active_count}", meta_value_style),
        Paragraph(f"<b>Concluídos:</b> {completed_count}", meta_value_style),
        Paragraph(f"<b>Total de agendamentos:</b> {len(schedules)} • <b>Itens:</b> {total_items}", summary_value_style),
    ]], colWidths=[doc.width * 0.25, doc.width * 0.25, doc.width * 0.50], hAlign="LEFT")
    footer_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef6ee")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cfe1d0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(footer_summary)

    generated_at_label = generated_at.strftime("%d/%m/%Y %H:%M")

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.line(doc.leftMargin, 22, landscape(A4)[0] - doc.rightMargin, 22)
        canvas.setFont("Helvetica", 8.2)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(doc.leftMargin, 10, f"Gerado por: {generated_by}")
        canvas.drawRightString(
            landscape(A4)[0] - doc.rightMargin,
            10,
            f"Emitido em {generated_at_label} • Página {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="planejamento_agendamentos.pdf"'},
    )


@router.post("/fertilizacao/agendamentos")
async def create_fertilization_schedule_action(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    validate_csrf(request, str(form.get("csrf_token") or ""))
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    allowed_plots = {
        plot.id: plot
        for plot in repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    }
    selected_plot_ids = []
    for raw_plot_id in form.getlist("plot_id"):
        parsed_plot_id = _int_or_none(raw_plot_id)
        if parsed_plot_id and parsed_plot_id not in selected_plot_ids:
            selected_plot_ids.append(parsed_plot_id)
    selected_plots = [allowed_plots[plot_id] for plot_id in selected_plot_ids if plot_id in allowed_plots]
    items = _parse_recommendation_items(form)
    if not selected_plots or not items:
        _flash(request, "error", "Selecione ao menos um setor e adicione ao menos um insumo.")
        return _redirect("/fertilizacao/agendamentos")
    try:
        for plot in selected_plots:
            create_fertilization_schedule(
                repo,
                {
                    "plot_id": plot.id,
                    "scheduled_date": str(form.get("scheduled_date") or ""),
                    "season_id": scope["active_season_id"],
                    "status": str(form.get("status") or "scheduled"),
                    "duration_minutes": form.get("duration_minutes"),
                    "application_method": form.get("application_method"),
                    "notes": str(form.get("notes") or "") or None,
                    "items": items,
                },
            )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect("/fertilizacao/agendamentos")
    if len(selected_plots) == 1:
        _flash(request, "success", "Agendamento salvo com sucesso.")
    else:
        _flash(request, "success", f"Agendamentos salvos com sucesso para {len(selected_plots)} setores.")
    return _redirect("/fertilizacao/agendamentos")


@router.post("/fertilizacao/agendamentos/{schedule_id}/editar")
async def update_fertilization_schedule_action(
    schedule_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    form = await request.form()
    validate_csrf(request, str(form.get("csrf_token") or ""))
    redirect_to = str(form.get("redirect_to") or "")
    target_url = redirect_to if redirect_to.startswith("/") else "/fertilizacao/agendamentos"
    repo = _repository(db)
    plot, scope, denied = _resolve_plot_in_scope(
        request,
        repo,
        int(form.get("plot_id") or 0),
        "/fertilizacao/agendamentos",
    )
    if denied:
        return denied
    schedule = repo.get_fertilization_schedule(schedule_id)
    if not schedule:
        _flash(request, "error", "Agendamento nao encontrado.")
        return _redirect(target_url)
    if not _plot_matches_scope(schedule.plot, scope):
        _flash(request, "error", "Este agendamento nao pertence ao contexto ativo.")
        return _redirect(target_url)
    items = _parse_recommendation_items(form)
    if not items:
        _flash(request, "error", "Adicione ao menos um insumo ao agendamento.")
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={schedule_id}")
    try:
        update_fertilization_schedule(
            repo,
            schedule,
            {
                "plot_id": plot.id,
                "scheduled_date": str(form.get("scheduled_date") or ""),
                "season_id": scope["active_season_id"],
                "status": str(form.get("status") or schedule.status),
                "duration_minutes": form.get("duration_minutes"),
                "application_method": form.get("application_method"),
                "notes": str(form.get("notes") or "") or None,
                "items": items,
            },
        )
    except ValueError as exc:
        _flash(request, "error", str(exc))
        separator = "&" if "?" in target_url else "?"
        return _redirect(f"{target_url}{separator}edit_id={schedule_id}")
    _flash(request, "success", "Agendamento atualizado com sucesso.")
    return _redirect(target_url)


@router.post("/fertilizacao/agendamentos/{schedule_id}/concluir")
def conclude_fertilization_schedule_action(
    schedule_id: int,
    request: Request,
    csrf_token: str = Form(...),
    application_date: str | None = Form(None),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/fertilizacao/agendamentos"
    repo = _repository(db)
    schedule = repo.get_fertilization_schedule(schedule_id)
    if not schedule:
        _flash(request, "error", "Agendamento nao encontrado.")
        return _redirect(target_url)
    if schedule.status == "completed" or schedule.fertilization_record_id:
        _flash(request, "info", "Este agendamento ja foi concluido.")
        return _redirect(target_url)
    purchased_inputs = repo.list_purchased_inputs()
    validation = validate_schedule_stock(repo, schedule, purchased_inputs_cache=purchased_inputs)
    if not validation["ok"]:
        first = validation["shortages"][0]
        _flash(request, "error", f"Estoque insuficiente. Necessario comprar {first['missing']} {first['unit']} de {first['name']}.")
        return _redirect(target_url)
    try:
        conclude_fertilization_schedule(repo, schedule, application_date)
    except ValueError as exc:
        _flash(request, "error", str(exc))
        return _redirect(target_url)
    _flash(request, "success", "Agendamento concluido e aplicacao registrada.")
    return _redirect(target_url)


@router.post("/fertilizacao/agendamentos/{schedule_id}/excluir")
def delete_fertilization_schedule_action(
    schedule_id: int,
    request: Request,
    csrf_token: str = Form(...),
    redirect_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    target_url = redirect_to if redirect_to and redirect_to.startswith("/") else "/fertilizacao/agendamentos"
    repo = _repository(db)
    schedule = repo.get_fertilization_schedule(schedule_id)
    if not schedule:
        _flash(request, "error", "Agendamento nao encontrado.")
        return _redirect(target_url)
    delete_fertilization_schedule(repo, schedule)
    _flash(request, "success", "Agendamento excluido com sucesso.")
    return _redirect(target_url)


@router.get("/producao")
def production_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date = _scoped_dates(scope["active_season"])
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    harvests = [
        harvest
        for harvest in repo.list_harvests()
        if harvest.plot_id in plot_ids and _within_scope(harvest.harvest_date, start_date, end_date)
    ]
    harvests = _sort_collection_desc(
        harvests,
        lambda item: item.harvest_date,
        lambda item: item.id,
    )
    harvests_pagination = _paginate_collection(request, harvests, "harvests_page")
    return templates.TemplateResponse(
        "production.html",
        _base_context(
            request,
            user,
            csrf_token,
            "production",
            _repo=repo,
            plots=plots,
            harvests=harvests_pagination["items"],
            harvests_pagination=harvests_pagination,
            edit_harvest=repo.get_harvest(edit_id) if edit_id else None,
        ),
    )


@router.post("/producao")
def create_harvest_action(
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    harvest_date: str = Form(...),
    sacks_produced: float = Form(...),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, _, denied = _resolve_plot_in_scope(request, repo, plot_id, "/producao")
    if denied:
        return denied
    area = float(plot.area_hectares) if plot else 0
    create_harvest(
        repo,
        {
            "plot_id": plot.id,
            "harvest_date": harvest_date,
            "sacks_produced": sacks_produced,
            "notes": notes,
        },
        area,
    )
    _flash(request, "success", "Colheita registrada com sucesso.")
    return _redirect("/producao")


@router.post("/producao/{record_id}/editar")
def update_harvest_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    harvest_date: str = Form(...),
    sacks_produced: float = Form(...),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, scope, denied = _resolve_plot_in_scope(request, repo, plot_id, "/producao")
    if denied:
        return denied
    harvest = repo.get_harvest(record_id)
    if not harvest:
        _flash(request, "error", "Registro de producao nao encontrado.")
        return _redirect_for_request(request, "/producao")
    if not _plot_matches_scope(harvest.plot, scope):
        _flash(request, "error", "Este registro de producao nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/producao")
    area = float(plot.area_hectares) if plot else 0
    update_harvest(
        repo,
        harvest,
        {
            "plot_id": plot.id,
            "harvest_date": harvest_date,
            "sacks_produced": sacks_produced,
            "notes": notes,
        },
        area,
    )
    _flash(request, "success", "Colheita atualizada com sucesso.")
    return _redirect_for_request(request, "/producao")


@router.post("/producao/{record_id}/excluir")
def delete_harvest_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    harvest = repo.get_harvest(record_id)
    if not harvest:
        _flash(request, "error", "Registro de producao nao encontrado.")
        return _redirect("/producao")
    repo.delete(harvest)
    _flash(request, "success", "Colheita excluida com sucesso.")
    return _redirect("/producao")


@router.get("/pragas")
def pests_page(
    request: Request,
    edit_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date = _scoped_dates(scope["active_season"])
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    incidents = [
        incident
        for incident in repo.list_pest_incidents()
        if incident.plot_id in plot_ids and _within_scope(incident.occurrence_date, start_date, end_date)
    ]
    incidents = _sort_collection_desc(
        incidents,
        lambda item: item.occurrence_date,
        lambda item: item.id,
    )
    incidents_pagination = _paginate_collection(request, incidents, "incidents_page")
    return templates.TemplateResponse(
        "pests.html",
        _base_context(
            request,
            user,
            csrf_token,
            "pests",
            _repo=repo,
            plots=plots,
            incidents=incidents_pagination["items"],
            incidents_pagination=incidents_pagination,
            edit_incident=repo.get_pest_incident(edit_id) if edit_id else None,
        ),
    )


@router.post("/pragas")
def create_pest_action(
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    occurrence_date: str = Form(...),
    category: str = Form(...),
    name: str = Form(...),
    severity: int = Form(...),
    treatment: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, _, denied = _resolve_plot_in_scope(request, repo, plot_id, "/pragas")
    if denied:
        return denied
    create_pest_incident(
        repo,
        {
            "plot_id": plot.id,
            "occurrence_date": occurrence_date,
            "category": category,
            "name": name,
            "severity": severity,
            "treatment": treatment,
            "notes": notes,
        },
    )
    _flash(request, "success", "Ocorrencia registrada com sucesso.")
    return _redirect("/pragas")


@router.post("/pragas/{record_id}/editar")
def update_pest_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    occurrence_date: str = Form(...),
    category: str = Form(...),
    name: str = Form(...),
    severity: int = Form(...),
    treatment: str | None = Form(None),
    notes: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, scope, denied = _resolve_plot_in_scope(request, repo, plot_id, "/pragas")
    if denied:
        return denied
    incident = repo.get_pest_incident(record_id)
    if not incident:
        _flash(request, "error", "Ocorrencia nao encontrada.")
        return _redirect_for_request(request, "/pragas")
    if not _plot_matches_scope(incident.plot, scope):
        _flash(request, "error", "Esta ocorrencia nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/pragas")
    update_pest_incident(
        repo,
        incident,
        {
            "plot_id": plot.id,
            "occurrence_date": occurrence_date,
            "category": category,
            "name": name,
            "severity": severity,
            "treatment": treatment,
            "notes": notes,
        },
    )
    _flash(request, "success", "Ocorrencia atualizada com sucesso.")
    return _redirect_for_request(request, "/pragas")


@router.post("/pragas/{record_id}/excluir")
def delete_pest_action(
    record_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    incident = repo.get_pest_incident(record_id)
    if not incident:
        _flash(request, "error", "Ocorrencia nao encontrada.")
        return _redirect("/pragas")
    repo.delete(incident)
    _flash(request, "success", "Ocorrencia excluida com sucesso.")
    return _redirect("/pragas")


@router.get("/perfil-agronomico")
def agronomic_profiles_page(
    request: Request,
    edit_farm_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    active_farm_id = scope["active_farm_id"]
    selected_farm = scope["active_farm"]
    auto_open_profile_modal = bool(
        edit_farm_id is not None and active_farm_id is not None and edit_farm_id == active_farm_id
    )
    return templates.TemplateResponse(
        "agronomic_profiles.html",
        _base_context(
            request,
            user,
            csrf_token,
            "agronomic_profiles",
            _repo=repo,
            title="Perfil Agronomico",
            farms=repo.list_farms(),
            profiles=[
                profile
                for profile in repo.list_agronomic_profiles()
                if not active_farm_id or profile.farm_id == active_farm_id
            ],
            edit_farm=selected_farm,
            edit_profile=repo.get_agronomic_profile_by_farm(active_farm_id) if active_farm_id else None,
            auto_open_profile_modal=auto_open_profile_modal,
        ),
    )


@router.post("/perfil-agronomico")
def save_agronomic_profile_action(
    request: Request,
    csrf_token: str = Form(...),
    culture: str = Form(...),
    region: str = Form(...),
    climate: str | None = Form(None),
    soil_type: str | None = Form(None),
    irrigation_system: str | None = Form(None),
    plant_spacing: str | None = Form(None),
    drip_spacing: str | None = Form(None),
    fertilizers_used: str | None = Form(None),
    crop_stage: str | None = Form(None),
    common_pests: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/perfil-agronomico")
    if denied:
        return denied
    farm_id = scope["active_farm_id"]
    existing = repo.get_agronomic_profile_by_farm(farm_id)
    payload = {
        "farm_id": farm_id,
        "culture": culture,
        "region": region,
        "climate": climate,
        "soil_type": soil_type,
        "irrigation_system": irrigation_system,
        "plant_spacing": plant_spacing,
        "drip_spacing": drip_spacing,
        "fertilizers_used": fertilizers_used,
        "crop_stage": crop_stage,
        "common_pests": common_pests,
    }
    if existing:
        update_agronomic_profile(repo, existing, payload)
        _flash(request, "success", "Perfil agronomico atualizado com sucesso.")
    else:
        create_agronomic_profile(repo, payload)
        _flash(request, "success", "Perfil agronomico salvo com sucesso.")
    return _redirect("/perfil-agronomico")


@router.post("/perfil-agronomico/{farm_id}/excluir")
def delete_agronomic_profile_action(
    farm_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    scope, denied = _launch_scope_or_redirect(request, repo, "/perfil-agronomico")
    if denied:
        return denied
    if farm_id != scope["active_farm_id"]:
        _flash(request, "error", "Este perfil agronomico nao pertence ao contexto ativo.")
        return _redirect("/perfil-agronomico")
    profile = repo.get_agronomic_profile_by_farm(farm_id)
    if not profile:
        _flash(request, "error", "Perfil agronomico nao encontrado.")
        return _redirect("/perfil-agronomico")
    repo.delete(profile)
    _flash(request, "success", "Perfil agronomico removido com sucesso.")
    return _redirect("/perfil-agronomico")


@router.get("/analise-solo")
def soil_analysis_page(
    request: Request,
    farm_id: int | None = None,
    plot_id: int | None = None,
    edit_id: int | None = None,
    compare_plot_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    effective_farm_id = farm_id or scope["active_farm_id"]
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date = _scoped_dates(scope["active_season"])
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    effective_plot_id = plot_id if plot_id else None
    analyses = [
        analysis
        for analysis in repo.list_soil_analyses(farm_id=effective_farm_id, plot_id=effective_plot_id)
        if (not plot_ids or analysis.plot_id in plot_ids) and _within_scope(analysis.analysis_date, start_date, end_date)
    ]
    analyses = _sort_collection_desc(
        analyses,
        lambda item: item.analysis_date,
        lambda item: item.id,
    )
    analyses_pagination = _paginate_collection(request, analyses, "analyses_page")
    compare_target = compare_plot_id or effective_plot_id
    compare_analyses = (
        [
            analysis
            for analysis in repo.list_soil_analyses(plot_id=compare_target)
            if _within_scope(analysis.analysis_date, start_date, end_date)
        ]
        if compare_target
        else analyses[:6]
    )
    return templates.TemplateResponse(
        "soil_analyses.html",
        _base_context(
            request,
            user,
            csrf_token,
            "soil_analyses",
            _repo=repo,
            title="Analise de Solo",
            farms=repo.list_farms(),
            plots=plots,
            analyses=analyses_pagination["items"],
            analyses_pagination=analyses_pagination,
            edit_analysis=repo.get_soil_analysis(edit_id) if edit_id else None,
            filters={"farm_id": effective_farm_id, "plot_id": effective_plot_id, "compare_plot_id": compare_target},
            compare_chart=_soil_history_chart(compare_analyses),
            compare_analyses=compare_analyses,
            latest_recommendations=[item for item in analyses if item.ai_recommendation][:4],
        ),
    )


@router.post("/analise-solo")
def create_soil_analysis_action(
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    analysis_date: str = Form(...),
    laboratory: str = Form(...),
    ph: str | None = Form(None),
    organic_matter: str | None = Form(None),
    phosphorus: str | None = Form(None),
    potassium: str | None = Form(None),
    calcium: str | None = Form(None),
    magnesium: str | None = Form(None),
    aluminum: str | None = Form(None),
    h_al: str | None = Form(None),
    ctc: str | None = Form(None),
    base_saturation: str | None = Form(None),
    observations: str | None = Form(None),
    analysis_pdf: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, scope, denied = _resolve_plot_in_scope(request, repo, plot_id, "/analise-solo")
    if denied:
        return denied
    pdf_filename, pdf_content_type, pdf_data = _read_upload(analysis_pdf)
    analysis = create_soil_analysis(
        repo,
        _soil_payload(
            farm_id=scope["active_farm_id"],
            plot_id=plot.id,
            analysis_date=analysis_date,
            laboratory=laboratory,
            ph=ph,
            organic_matter=organic_matter,
            phosphorus=phosphorus,
            potassium=potassium,
            calcium=calcium,
            magnesium=magnesium,
            aluminum=aluminum,
            h_al=h_al,
            ctc=ctc,
            base_saturation=base_saturation,
            observations=observations,
            pdf_filename=pdf_filename,
            pdf_content_type=pdf_content_type,
            pdf_data=pdf_data,
        ),
    )
    analysis = _apply_soil_ai_recommendation(repo, analysis)
    if analysis.ai_status == "generated":
        _flash(request, "success", "Analise de solo salva e recomendacao da IA gerada com sucesso.")
    elif analysis.ai_status == "skipped":
        _flash(request, "success", "Analise de solo salva. Configure a OPENAI_API_KEY para gerar recomendacoes personalizadas.")
    else:
        _flash(request, "error", "Analise de solo salva, mas a recomendacao da IA falhou nesta tentativa.")
    return _redirect_with_query("/analise-solo", plot_id=plot.id, compare_plot_id=plot.id)


@router.post("/analise-solo/{analysis_id}/editar")
def update_soil_analysis_action(
    analysis_id: int,
    request: Request,
    csrf_token: str = Form(...),
    plot_id: int = Form(...),
    analysis_date: str = Form(...),
    laboratory: str = Form(...),
    ph: str | None = Form(None),
    organic_matter: str | None = Form(None),
    phosphorus: str | None = Form(None),
    potassium: str | None = Form(None),
    calcium: str | None = Form(None),
    magnesium: str | None = Form(None),
    aluminum: str | None = Form(None),
    h_al: str | None = Form(None),
    ctc: str | None = Form(None),
    base_saturation: str | None = Form(None),
    observations: str | None = Form(None),
    analysis_pdf: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot, scope, denied = _resolve_plot_in_scope(request, repo, plot_id, "/analise-solo")
    if denied:
        return denied
    analysis = repo.get_soil_analysis(analysis_id)
    if not analysis:
        _flash(request, "error", "Analise de solo nao encontrada.")
        return _redirect_for_request(request, "/analise-solo")
    if not _farm_matches_scope(analysis.farm_id, scope) or not _plot_matches_scope(analysis.plot, scope):
        _flash(request, "error", "Esta analise de solo nao pertence ao contexto ativo.")
        return _redirect_for_request(request, "/analise-solo")
    pdf_filename, pdf_content_type, pdf_data = _read_upload(analysis_pdf)
    updated = update_soil_analysis(
        repo,
        analysis,
        _soil_payload(
            farm_id=scope["active_farm_id"],
            plot_id=plot.id,
            analysis_date=analysis_date,
            laboratory=laboratory,
            ph=ph,
            organic_matter=organic_matter,
            phosphorus=phosphorus,
            potassium=potassium,
            calcium=calcium,
            magnesium=magnesium,
            aluminum=aluminum,
            h_al=h_al,
            ctc=ctc,
            base_saturation=base_saturation,
            observations=observations,
            pdf_filename=pdf_filename,
            pdf_content_type=pdf_content_type,
            pdf_data=pdf_data,
            current_analysis=analysis,
        ),
    )
    updated = _apply_soil_ai_recommendation(repo, updated)
    if updated.ai_status == "generated":
        _flash(request, "success", "Analise de solo atualizada e recomendacao regenerada.")
    elif updated.ai_status == "skipped":
        _flash(request, "success", "Analise de solo atualizada. Configure a OPENAI_API_KEY para gerar recomendacoes personalizadas.")
    else:
        _flash(request, "error", "Analise de solo atualizada, mas a recomendacao da IA falhou nesta tentativa.")
    return _redirect_for_request(request, "/analise-solo", plot_id=plot.id, compare_plot_id=plot.id)


@router.post("/analise-solo/{analysis_id}/regenerar")
def regenerate_soil_recommendation_action(
    analysis_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    analysis = repo.get_soil_analysis(analysis_id)
    if not analysis:
        _flash(request, "error", "Analise de solo nao encontrada.")
        return _redirect("/analise-solo")
    updated = _apply_soil_ai_recommendation(repo, analysis)
    if updated.ai_status == "generated":
        _flash(request, "success", "Recomendacao agronomica regenerada com sucesso.")
    elif updated.ai_status == "skipped":
        _flash(request, "error", "OPENAI_API_KEY nao configurada para gerar a recomendacao.")
    else:
        _flash(request, "error", "Nao foi possivel gerar a recomendacao agronomica nesta tentativa.")
    return _redirect_with_query("/analise-solo", edit_id=analysis_id, plot_id=analysis.plot_id, compare_plot_id=analysis.plot_id)


@router.post("/analise-solo/{analysis_id}/excluir")
def delete_soil_analysis_action(
    analysis_id: int,
    request: Request,
    csrf_token: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    analysis = repo.get_soil_analysis(analysis_id)
    if not analysis:
        _flash(request, "error", "Analise de solo nao encontrada.")
        return _redirect("/analise-solo")
    plot_id = analysis.plot_id
    repo.delete(analysis)
    _flash(request, "success", "Analise de solo excluida com sucesso.")
    return _redirect_with_query("/analise-solo", compare_plot_id=plot_id)


@router.get("/analise-solo/{analysis_id}/pdf")
def soil_analysis_pdf_download(
    analysis_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    analysis = _repository(db).get_soil_analysis(analysis_id)
    if not analysis or not analysis.pdf_data:
        return Response(status_code=status.HTTP_404_NOT_FOUND)
    headers = {"Content-Disposition": f'inline; filename="{analysis.pdf_filename or "analise-solo.pdf"}"'}
    return Response(content=analysis.pdf_data, media_type=analysis.pdf_content_type or "application/pdf", headers=headers)


@router.get("/mapa")
def map_page(
    request: Request,
    edit_plot_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    geojson = build_dashboard_context(
        repo,
        farm_id=scope["active_farm_id"],
        season=scope["active_season"],
    )["map_geojson"]
    edit_plot = repo.get_plot(edit_plot_id) if edit_plot_id else None
    return templates.TemplateResponse(
        "map.html",
        _base_context(
            request,
            user,
            csrf_token,
            "map",
            _repo=repo,
            map_geojson=geojson,
            plots=repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids),
            edit_plot=edit_plot,
            edit_plot_geometry=edit_plot.boundary_geojson if edit_plot and edit_plot.boundary_geojson else None,
        ),
    )


@router.post("/mapa/setores/{plot_id}/salvar")
def save_plot_map_geometry(
    plot_id: int,
    request: Request,
    csrf_token: str = Form(...),
    boundary_geojson: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
):
    del user
    validate_csrf(request, csrf_token)
    repo = _repository(db)
    plot = repo.get_plot(plot_id)
    if not plot:
        _flash(request, "error", "Setor nao encontrado para salvar a edicao no mapa.")
        return _redirect("/mapa")
    normalized = normalize_geojson(boundary_geojson)
    if not normalized:
        _flash(request, "error", "A geometria enviada nao e valida.")
        return _redirect_with_query("/mapa", edit_plot_id=plot_id)
    area_hectares = calculate_geojson_area_hectares(normalized)
    centroid_lat, centroid_lng = estimate_geojson_centroid(normalized)
    repo.update(
        plot,
        {
            "boundary_geojson": normalized,
            "area_hectares": area_hectares if area_hectares is not None else plot.area_hectares,
            "centroid_lat": centroid_lat if centroid_lat is not None else plot.centroid_lat,
            "centroid_lng": centroid_lng if centroid_lng is not None else plot.centroid_lng,
        },
    )
    try:
        generate_plot_preview_image(plot_id, normalized, _farm_boundary_for_plot_preview(plot))
    except Exception:
        logging.getLogger(__name__).exception("Falha ao gerar preview do setor %s apos mapa", plot_id)
    _flash(request, "success", "Poligono do setor atualizado com sucesso no mapa.")
    return _redirect_with_query("/mapa", edit_plot_id=plot_id)


@router.get("/mobile")
def mobile_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_web),
    csrf_token: str = Depends(get_csrf_token),
):
    repo = _repository(db)
    scope = _global_scope_context(request, repo)
    farm_ids, variety_ids = _scoped_plot_filters(request, scope["active_season"])
    start_date, end_date = _scoped_dates(scope["active_season"])
    plots = repo.list_plots(farm_ids=farm_ids, variety_ids=variety_ids)
    plot_ids = {plot.id for plot in plots}
    return templates.TemplateResponse(
        "mobile.html",
        _base_context(
            request,
            user,
            csrf_token,
            "mobile",
            _repo=repo,
            plots=plots,
            quick_irrigations=[
                irrigation
                for irrigation in repo.list_irrigations(limit=20)
                if irrigation.plot_id in plot_ids and _within_scope(irrigation.irrigation_date, start_date, end_date)
            ][:3],
            quick_incidents=[
                incident
                for incident in repo.list_pest_incidents(limit=20)
                if incident.plot_id in plot_ids and _within_scope(incident.occurrence_date, start_date, end_date)
            ][:3],
        ),
    )
